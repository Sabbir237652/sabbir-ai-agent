#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
🤖 Sabbir AI Agent — সম্পূর্ণ Advanced AI Agent, একটাই ফাইলে! (১০০% ফ্রি)

ক্ষমতা:
  ✅ Google Gemini (ফ্রি API) দিয়ে বাংলা/English chat
  ✅ ১৮টা tool: calculator, web search, note, document পড়া (RAG),
     planning, code execution, long-term memory... সব!
  ✅ ব্রাউজারে সুন্দর chat UI + ফাইল upload + chart display
  ✅ Restart করলেও সব মনে রাখে

চালানোর নিয়ম:
  ১. pip install requests pypdf matplotlib
  ২. এই ফাইলের পাশে ".env" নামে ফাইল বানিয়ে লিখুন:
       GEMINI_API_KEY=আপনার_key
     (ফ্রি key: https://aistudio.google.com/apikey)
     — না বানালেও চলবে, ব্রাউজার থেকেই key বসানো যায়।
  ৩. python sabbir_ai_agent.py
  ৪. ব্রাউজারে খুলুন: http://localhost:8000

Terminal-এ chat করতে চাইলে:  python sabbir_ai_agent.py --terminal
"""
import ast
import base64
import datetime
import html
import json
import math
import os
import re
import subprocess
import sys
import hashlib
import hmac
import threading
import time
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path

import requests

BASE = Path(__file__).parent

# ═══════════════════════════════════════════════════════════════════
# ১. CONFIG — settings, API key, system prompt
# ═══════════════════════════════════════════════════════════════════

def load_env():
    """.env ফাইল থেকে key লোড করা (কোনো extra library ছাড়াই)।"""
    env_path = BASE / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, value = line.partition("=")
                os.environ.setdefault(key.strip(), value.strip())

load_env()

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
MODEL_NAME = os.environ.get("MODEL_NAME", "gemini-3.6-flash")

# 🔐 ACCESS_CODE: লাইভ সার্ভারে অন্য কেউ যেন আপনার agent ব্যবহার করতে না পারে!
# Hugging Face Space-এর Settings → Variables and secrets-এ ACCESS_CODE সেট করুন।
# খালি রাখলে পাসওয়ার্ড ছাড়াই চলবে (নিরাপদ না)।
ACCESS_CODE = os.environ.get("ACCESS_CODE", "")

# মূল মডেলের ফ্রি কোটা শেষ হলে এগুলো একে একে চেষ্টা হবে (auto-fallback)
FALLBACK_MODELS = [
    "gemini-3.5-flash",
    "gemini-3.1-flash-lite",
    "gemini-flash-lite-latest",
]


# ╔═══════════════════════════════════════════════════════════════╗
# ║  🎨 BRAND IDENTITY — এখানে সব বদলান, নিজের মতো সাজান!          ║
# ║  শুধু নিচের লেখাগুলো বদলে save করলেই পুরো agent বদলে যাবে।    ║
# ╚═══════════════════════════════════════════════════════════════╝
BRAND = {
    # ── নাম ও পরিচয় ──
    "name": "Sabbir AI Agent",
    "tagline": "Personal Intelligence System",

    # ── Brand Colors (gradient-এর ৩টা রঙ, hex code) ──
    # রঙ বদলাতে চাইলে: https://htmlcolorcodes.com থেকে পছন্দের hex নিন
    "color1": "#7c3aed",   # Violet
    "color2": "#a78bfa",   # Light violet
    "color3": "#6366f1",   # Indigo

    # ── Personality (agent-এর চরিত্র — যা ইচ্ছা লিখুন) ──
    "personality": (
        "তুমি বন্ধুসুলভ, বুদ্ধিমান আর সাহায্য করতে সবসময় আগ্রহী। "
        "তুমি চট্টগ্রামের Sabbir-এর বানানো ব্যক্তিগত AI। "
        "তুমি আশাবাদী আর সমাধানমুখী — সমস্যা দেখলে ঘাবড়াও না, পথ খোঁজো।"
    ),

    # ── Voice / Tone (কথা বলার ধরন) ──
    "tone": (
        "উষ্ণ, ভদ্র আর প্রাণবন্ত। মাঝে মাঝে হালকা emoji ব্যবহার করো (বেশি না)। "
        "কঠিন জিনিস সহজ ভাষায় বোঝাও, যেন বন্ধুর সাথে কথা বলছো।"
    ),

    # ── AI Response Style (উত্তর সাজানোর নিয়ম) ──
    "response_style": (
        "ছোট প্রশ্নের ছোট উত্তর দাও, বড় প্রশ্নে গুছিয়ে heading/bullet ব্যবহার করো। "
        "গুরুত্বপূর্ণ শব্দ **bold** করো। তথ্য দিলে উৎস বলো। "
        "শেষে দরকার হলে ছোট্ট follow-up প্রশ্ন করো।"
    ),

    # ── Welcome Message (প্রথমবার খুললে যা দেখাবে) ──
    "welcome": (
        "আসসালামু আলাইকুম! 👋 আমি Sabbir AI Agent — আপনার নিজের AI সহকারী।\n\n"
        "আমি পারি: 🌐 ইন্টারনেটে খুঁজতে, 🧮 হিসাব করতে, 📊 chart বানাতে, "
        "📚 আপনার PDF পড়তে, 📝 note রাখতে — আর আপনার কথা মনেও রাখি!\n\n"
        "নিচের suggestion-এ চাপ দিন বা যা মনে চায় লিখুন — বাংলা বা English, দুটোই চলবে! ✨"
    ),

    # ── Loading Animation-এর লেখা (agent ভাবার সময়) ──
    "thinking_text": "Sabbir ভাবছে",

    # ── Suggestion Chips (chat-এর উপরের বাটনগুলো) ──
    "chips": [
        "👋 তুমি কী কী করতে পারো?",
        "🌦️ আজকের আবহাওয়া কেমন?",
        "🧮 2500 টাকার 15% কত?",
        "📊 একটা chart বানিয়ে দেখাও",
    ],
}

def ui_brand():
    """UI-তে পাঠানোর brand info।"""
    return {k: BRAND[k] for k in
            ("name", "tagline", "welcome", "thinking_text", "chips",
             "color1", "color2", "color3")}

def build_system_prompt():
    return f"""তুমি একজন AI agent। তোমার নাম "{BRAND['name']}"।
তোমার personality: {BRAND['personality']}
তোমার কথা বলার ধরন: {BRAND['tone']}
তোমার উত্তরের style: {BRAND['response_style']}

এখন তারিখ-সময়: {now_local().strftime("%Y-%m-%d %H:%M")} (বাংলাদেশ)। User সময় দিয়ে কিছু মনে করাতে বললে reminder_tool দিয়ে set করো।
তুমি বাংলা ও ইংরেজি দুই ভাষাতেই কথা বলতে পারো — user যে ভাষায় লেখে সেই ভাষায় উত্তর দাও।
তোমার কাছে কিছু tool আছে। যখন কোনো কাজের জন্য tool দরকার হয়, তখন tool ব্যবহার করো।
Tool-এর ফলাফল পাওয়ার পর user-কে সুন্দরভাবে উত্তর দাও।

গুরুত্বপূর্ণ: user যখন নিজের সম্পর্কে কিছু জানায় (নাম, ঠিকানা, পছন্দ, পেশা, লক্ষ্য ইত্যাদি),
তখন remember_fact tool দিয়ে সেটা স্থায়ীভাবে মনে রেখো — যাতে ভবিষ্যতে কাজে লাগাতে পারো।
তোমার "স্থায়ী স্মৃতি" অংশে যা লেখা আছে, তা তুমি আগে থেকেই জানো — সেগুলো ব্যবহার করে উত্তর দাও।

ভাবনা ও কাজের নিয়ম (Reasoning):
- জটিল প্রশ্নে: আগে সমস্যাটা ছোট অংশে ভাঙো, প্রতিটা অংশ সমাধান করো, তারপর মিলিয়ে উত্তর দাও।
- সিদ্ধান্তের প্রশ্নে: option গুলোর pros & cons তুলনা করে যুক্তিসহ সুপারিশ দাও।
- Tool-এর ফলাফল পাওয়ার পর যাচাই করো: ফলাফলটা কি প্রশ্নের উত্তর দেয়? না দিলে ভিন্নভাবে আবার চেষ্টা করো (যেমন: অন্য search query, অন্য পদ্ধতি)।
- কোনো tool error দিলে হাল ছেড়ো না — অন্তত ১ বার ভিন্ন উপায়ে চেষ্টা করো, তারপরও না হলে user-কে সৎভাবে জানাও কী চেষ্টা করেছো।
- অনিশ্চিত তথ্য অনুমান করে বলো না — web_search দিয়ে যাচাই করো বা "নিশ্চিত না" বলো।

বড়/জটিল কাজের নিয়ম (Multi-step Planning):
- কাজটায় যদি একাধিক ধাপ বা একাধিক tool লাগে, তাহলে আগে create_plan দিয়ে plan বানাও।
- তারপর এক এক করে ধাপ execute করো, প্রতিটি ধাপ শেষে complete_step দিয়ে টিক দাও।
- সব ধাপ শেষ হলে user-কে গুছিয়ে চূড়ান্ত ফলাফল দাও।
- সহজ প্রশ্নে (যেমন: সাধারণ কথা, একটামাত্র হিসাব) plan বানানোর দরকার নেই — সরাসরি উত্তর দাও।

নতুন feature চাইলে (Self-building):
- User নতুন feature/ক্ষমতা চাইলে আগে ভাবো: এটা কি একটা ছোট Python function দিয়ে করা যায়?
- গেলে: create_plugin দিয়ে নিজেই feature-টা বানিয়ে ফেলো, তারপর test করে দেখাও।
- API key/টাকা/বাইরের service লাগলে: বানিও না — বরং ধাপে ধাপে গাইড দাও (কোথায় account, কী key, কোথায় বসাবে) — ঠিক একজন শিক্ষকের মতো।
- জটিল বড় feature হলে: সৎভাবে বলো এটা developer দিয়ে করাতে হবে, আর কী কী লাগবে তার তালিকা দাও।

Autonomous workflow-এর নিয়ম:
- বড় কাজ পেলে: তথ্য সংগ্রহ → plan → ধাপে ধাপে execute → ফলাফল যাচাই → ভুলে retry → শেষে report।
- কোনো tool "✋ অনুমতির অপেক্ষায়" বললে সেটা স্বাভাবিক — user-কে বলো Control-এ গিয়ে approve করতে, জোর করে আবার চালিও না।
- টাকা, মোছা, বা code চালানোর মতো sensitive কাজের আগে user-এর ইচ্ছা স্পষ্ট না হলে জিজ্ঞেস করে নাও।
"""

SYSTEM_PROMPT = None  # runtime-এ build_system_prompt() ব্যবহার হয়
MAX_TOOL_ITERATIONS = 15


def set_api_key(key: str):
    """নতুন API key সেট করে .env ফাইলে save করে।"""
    global GEMINI_API_KEY
    GEMINI_API_KEY = key
    (BASE / ".env").write_text(
        f"GEMINI_API_KEY={key}\nMODEL_NAME={MODEL_NAME}\n", encoding="utf-8"
    )


# ═══════════════════════════════════════════════════════════════════
# ২. MEMORY — long-term memory (restart-প্রুফ)
# ═══════════════════════════════════════════════════════════════════

MEMORY_DIR = BASE / "memory"
MEMORY_DIR.mkdir(exist_ok=True)
HISTORY_FILE = MEMORY_DIR / "chat_history.json"
FACTS_FILE = MEMORY_DIR / "facts.json"
MAX_HISTORY_MESSAGES = 40


def load_history() -> list:
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []


def save_history(history: list):
    HISTORY_FILE.write_text(
        json.dumps(history, ensure_ascii=False, indent=1), encoding="utf-8"
    )


def clear_history():
    if HISTORY_FILE.exists():
        HISTORY_FILE.unlink()


def trim_history(history: list) -> list:
    """LLM-এ পাঠানোর জন্য history ছোট করে (টোকেন বাঁচাতে)।"""
    if len(history) <= MAX_HISTORY_MESSAGES:
        return history
    trimmed = history[-MAX_HISTORY_MESSAGES:]
    while trimmed:
        first = trimmed[0]
        parts = first.get("parts", [])
        if first.get("role") == "user" and parts and "text" in parts[0]:
            break
        trimmed.pop(0)
    return trimmed if trimmed else history[-2:]


def load_facts() -> list:
    if FACTS_FILE.exists():
        try:
            return json.loads(FACTS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []


def save_facts(facts: list):
    FACTS_FILE.write_text(
        json.dumps(facts, ensure_ascii=False, indent=1), encoding="utf-8"
    )


def add_fact(text: str) -> str:
    facts = load_facts()
    if any(f["text"] == text for f in facts):
        return "এই তথ্যটা আগে থেকেই মনে রাখা আছে।"
    facts.append({"text": text, "date": datetime.date.today().isoformat()})
    save_facts(facts)
    return f"মনে রাখলাম: \"{text}\" (মোট {len(facts)}টি তথ্য মনে আছে)"


def delete_fact(number: int) -> str:
    facts = load_facts()
    try:
        number = int(number)
    except (TypeError, ValueError):
        return "Error: number একটা সংখ্যা হতে হবে।"
    if 1 <= number <= len(facts):
        removed = facts.pop(number - 1)
        save_facts(facts)
        return f"ভুলে গেলাম: \"{removed['text']}\""
    return f"{number} নম্বর তথ্য পাওয়া যায়নি। মোট আছে {len(facts)}টি।"


def facts_as_text() -> str:
    facts = load_facts()
    if not facts:
        return ""
    lines = [f"{i+1}. {f['text']} (সংরক্ষিত: {f['date']})" for i, f in enumerate(facts)]
    return (
        "\n\n--- তোমার স্থায়ী স্মৃতি (user সম্পর্কে মনে রাখা তথ্য) ---\n"
        + "\n".join(lines)
        + "\n--- স্মৃতি শেষ ---"
    )


# ═══════════════════════════════════════════════════════════════════
# ৩. PLANNER — multi-step planning
# ═══════════════════════════════════════════════════════════════════

PLAN_FILE = MEMORY_DIR / "plan.json"


def _plan_load():
    if PLAN_FILE.exists():
        try:
            return json.loads(PLAN_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None
    return None


def _plan_save(plan):
    PLAN_FILE.write_text(json.dumps(plan, ensure_ascii=False, indent=1), encoding="utf-8")


def _plan_render(plan) -> str:
    done = sum(1 for s in plan["steps"] if s["done"])
    total = len(plan["steps"])
    lines = [f"📋 কাজ: {plan['goal']}  [{done}/{total} ধাপ শেষ]"]
    for i, s in enumerate(plan["steps"], 1):
        mark = "✅" if s["done"] else "⬜"
        lines.append(f"{mark} ধাপ {i}: {s['text']}")
    if done == total:
        lines.append("🎉 সব ধাপ শেষ! এবার user-কে চূড়ান্ত ফলাফল জানাও।")
    else:
        next_step = next(i for i, s in enumerate(plan["steps"], 1) if not s["done"])
        lines.append(f"👉 পরের কাজ: ধাপ {next_step}")
    return "\n".join(lines)


def create_plan(goal: str, steps: list) -> str:
    if not steps:
        return "Error: অন্তত একটা ধাপ দিতে হবে।"
    plan = {"goal": goal, "steps": [{"text": str(s), "done": False} for s in steps]}
    _plan_save(plan)
    return "নতুন plan তৈরি হলো:\n" + _plan_render(plan)


def complete_step(step_number: int, result_summary: str = "") -> str:
    plan = _plan_load()
    if not plan:
        return "কোনো plan নেই। আগে create_plan দিয়ে plan বানাও।"
    try:
        step_number = int(step_number)
    except (TypeError, ValueError):
        return "Error: step_number একটা সংখ্যা হতে হবে।"
    if not (1 <= step_number <= len(plan["steps"])):
        return f"ধাপ {step_number} নেই। মোট ধাপ: {len(plan['steps'])}"
    plan["steps"][step_number - 1]["done"] = True
    if result_summary:
        plan["steps"][step_number - 1]["result"] = result_summary
    _plan_save(plan)
    return _plan_render(plan)


def view_plan() -> str:
    plan = _plan_load()
    if not plan:
        return "এখন কোনো plan চালু নেই।"
    return _plan_render(plan)


# ═══════════════════════════════════════════════════════════════════
# ৪. RAG — নিজের document (PDF/TXT/MD) পড়ে উত্তর
# ═══════════════════════════════════════════════════════════════════

_LIBS = BASE / "libs"
if _LIBS.exists() and str(_LIBS) not in sys.path:
    sys.path.insert(0, str(_LIBS))

DOCS_DIR = BASE / "documents"
DOCS_DIR.mkdir(exist_ok=True)
INDEX_FILE = MEMORY_DIR / "rag_index.json"
CHUNK_SIZE = 900
CHUNK_OVERLAP = 150
SUPPORTED = {".pdf", ".txt", ".md", ".csv", ".py", ".html", ".json", ".xlsx"}


def extract_text(path: Path) -> str:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        # Excel → টেক্সট হিসেবে cell গুলো
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            parts = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    parts.append(" ".join(str(c) for c in row if c is not None))
            wb.close()
            return "\n".join(parts)
        except Exception:
            return ""
    if path.suffix.lower() == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError:
            return ""
        try:
            reader = PdfReader(str(path))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except Exception:
            return ""
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def make_chunks(text: str) -> list:
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return []
    chunks = []
    start = 0
    while start < len(text):
        end = start + CHUNK_SIZE
        if end < len(text):
            for sep in ("। ", ". ", "? ", "! ", "\n"):
                pos = text.rfind(sep, start + CHUNK_SIZE // 2, end)
                if pos != -1:
                    end = pos + len(sep)
                    break
        chunks.append(text[start:end].strip())
        start = max(end - CHUNK_OVERLAP, start + 1)
    return [c for c in chunks if c]


def _tokenize(text: str) -> list:
    return re.findall(r"[a-zA-Z0-9\u0980-\u09FF]+", text.lower())


def build_index() -> str:
    files = [f for f in DOCS_DIR.iterdir() if f.suffix.lower() in SUPPORTED]
    if not files:
        INDEX_FILE.write_text("[]", encoding="utf-8")
        return "documents ফোল্ডারে কোনো ফাইল নেই। আগে PDF/TXT/MD ফাইল যোগ করুন।"
    index, report = [], []
    for f in files:
        chunks = make_chunks(extract_text(f))
        for i, chunk in enumerate(chunks):
            index.append({"file": f.name, "chunk_id": i, "text": chunk,
                          "tokens": _tokenize(chunk)})
        report.append(f"• {f.name}: {len(chunks)} টুকরা")
    INDEX_FILE.write_text(json.dumps(index, ensure_ascii=False), encoding="utf-8")
    return (f"✅ Index তৈরি হলো — {len(files)}টি ফাইল, মোট {len(index)}টি টুকরা:\n"
            + "\n".join(report))


def _load_index() -> list:
    if INDEX_FILE.exists():
        try:
            return json.loads(INDEX_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []


def docs_search(query: str, top_k: int = 4) -> str:
    """BM25-ধাঁচের keyword scoring দিয়ে document-এ খোঁজা (সম্পূর্ণ ফ্রি)।"""
    index = _load_index()
    if not index:
        result = build_index()
        index = _load_index()
        if not index:
            return result
    q_tokens = _tokenize(query)
    if not q_tokens:
        return "প্রশ্নে কোনো শব্দ পাওয়া যায়নি।"
    n_docs = len(index)
    doc_freq = {}
    for item in index:
        for t in set(item["tokens"]):
            doc_freq[t] = doc_freq.get(t, 0) + 1
    scored = []
    for item in index:
        tokens = item["tokens"]
        if not tokens:
            continue
        score = 0.0
        for qt in q_tokens:
            tf = tokens.count(qt)
            if tf == 0:
                continue
            idf = math.log(1 + (n_docs / (1 + doc_freq.get(qt, 0))))
            score += idf * (tf / (tf + 1.5)) / math.sqrt(len(tokens) / 100)
        if score > 0:
            scored.append((score, item))
    if not scored:
        files = sorted({i["file"] for i in index})
        return (f"'{query}' এর সাথে মেলে এমন কিছু পাওয়া যায়নি।\n"
                f"Index-এ থাকা ফাইল: {', '.join(files)}")
    scored.sort(key=lambda x: -x[0])
    results = [f"[উৎস: {item['file']}, টুকরা #{item['chunk_id']}]\n{item['text']}"
               for _, item in scored[:top_k]]
    return (f"'{query}' এর জন্য {len(results)}টি প্রাসঙ্গিক অংশ পাওয়া গেল:\n\n"
            + "\n\n---\n\n".join(results))


def list_documents() -> str:
    files = [f for f in DOCS_DIR.iterdir() if f.is_file()]
    if not files:
        return "documents ফোল্ডারে এখনো কোনো ফাইল নেই।"
    indexed = {i["file"] for i in _load_index()}
    lines = []
    for f in sorted(files):
        status = "✅ indexed" if f.name in indexed else "⬜ index হয়নি"
        lines.append(f"• {f.name} ({f.stat().st_size / 1024:.1f} KB) — {status}")
    return "Documents:\n" + "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════
# ৫. CODE RUNNER — Sabbir নিজে Python code লিখে চালায়
# ═══════════════════════════════════════════════════════════════════

SANDBOX = BASE / "sandbox"
SANDBOX.mkdir(exist_ok=True)
TIMEOUT_SECONDS = 15
MAX_OUTPUT = 3000


def run_python(code: str) -> str:
    if not code.strip():
        return "Error: কোনো code দেওয়া হয়নি।"
    script = SANDBOX / "_run.py"
    script.write_text(code, encoding="utf-8")
    before_files = {f.name for f in SANDBOX.iterdir()}
    env = os.environ.copy()
    if _LIBS.exists():
        env["PYTHONPATH"] = str(_LIBS) + os.pathsep + env.get("PYTHONPATH", "")
    env["MPLBACKEND"] = "Agg"
    try:
        result = subprocess.run(
            [sys.executable, str(script)],
            capture_output=True, text=True,
            timeout=TIMEOUT_SECONDS, cwd=str(SANDBOX), env=env,
        )
    except subprocess.TimeoutExpired:
        return f"⏱️ Timeout! Code {TIMEOUT_SECONDS} সেকেন্ডের বেশি চলছিল, তাই থামিয়ে দেওয়া হয়েছে।"
    out = (result.stdout or "").strip()
    err = (result.stderr or "").strip()
    new_files = sorted({f.name for f in SANDBOX.iterdir()} - before_files - {"_run.py"})
    parts = []
    if out:
        parts.append("Output:\n" + out[:MAX_OUTPUT] + ("..." if len(out) > MAX_OUTPUT else ""))
    if err:
        parts.append("Errors:\n" + err[:1000])
    if new_files:
        parts.append("📁 নতুন ফাইল তৈরি হয়েছে: " + ", ".join(new_files) + " (sandbox ফোল্ডারে)")
    if not parts:
        parts.append("Code চলেছে, কিন্তু কোনো output দেয়নি। (print() ব্যবহার করতে ভুলো না!)")
    status = "✅ সফল" if result.returncode == 0 else f"❌ Exit code: {result.returncode}"
    return status + "\n\n" + "\n\n".join(parts)


def list_sandbox_files() -> str:
    files = [f for f in SANDBOX.iterdir() if f.is_file() and f.name != "_run.py"]
    if not files:
        return "Sandbox ফোল্ডার খালি।"
    return "Sandbox-এর ফাইল:\n" + "\n".join(
        f"• {f.name} ({f.stat().st_size / 1024:.1f} KB)" for f in sorted(files))


# ═══════════════════════════════════════════════════════════════════
# ৬. TOOLS — Sabbir-এর ১৮টা হাত-পা
# ═══════════════════════════════════════════════════════════════════

WORKSPACE = BASE / "agent_files"
WORKSPACE.mkdir(exist_ok=True)


def calculator(expression: str) -> str:
    allowed = {"__builtins__": {}}
    allowed.update({k: getattr(math, k) for k in dir(math) if not k.startswith("_")})
    try:
        return f"ফলাফল: {eval(expression, allowed, {})}"
    except Exception as e:
        return f"হিসাবে ভুল: {e}"


def get_current_time() -> str:
    return now_local().strftime("তারিখ: %Y-%m-%d, সময়: %H:%M:%S") + " (বাংলাদেশ)"


def save_note(filename: str, content: str) -> str:
    safe_name = Path(filename).name
    (WORKSPACE / safe_name).write_text(content, encoding="utf-8")
    return f"'{safe_name}' ফাইলে সংরক্ষণ করা হয়েছে ({len(content)} অক্ষর)।"


def read_note(filename: str) -> str:
    safe_name = Path(filename).name
    path = WORKSPACE / safe_name
    if not path.exists():
        files = [f.name for f in WORKSPACE.iterdir()] or ["(কোনো ফাইল নেই)"]
        return f"'{safe_name}' পাওয়া যায়নি। আছে: {', '.join(files)}"
    return path.read_text(encoding="utf-8")


def list_notes() -> str:
    files = [f.name for f in WORKSPACE.iterdir() if f.is_file()]
    return "সংরক্ষিত ফাইল: " + (", ".join(files) if files else "কোনো ফাইল নেই")


def web_search(query: str) -> str:
    """DuckDuckGo দিয়ে ইন্টারনেটে খোঁজা (ফ্রি, key লাগে না)।"""
    try:
        resp = requests.post(
            "https://html.duckduckgo.com/html/",
            data={"q": query},
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
            timeout=20,
        )
        resp.raise_for_status()
    except requests.RequestException as e:
        return f"Search error: {e}"
    titles = re.findall(
        r'<a[^>]+class="result__a"[^>]*href="([^"]+)"[^>]*>(.*?)</a>', resp.text, re.DOTALL)
    snippets = re.findall(
        r'<a[^>]+class="result__snippet"[^>]*>(.*?)</a>', resp.text, re.DOTALL)

    def clean(t):
        return html.unescape(re.sub(r"<[^>]+>", "", t)).strip()

    results = []
    for i, (link, title) in enumerate(titles[:5]):
        snippet = clean(snippets[i]) if i < len(snippets) else ""
        m = re.search(r"uddg=([^&]+)", link)
        if m:
            from urllib.parse import unquote
            link = unquote(m.group(1))
        results.append(f"{i+1}. {clean(title)}\n   URL: {link}\n   {snippet}")
    if not results:
        return f"'{query}' এর জন্য কোনো ফলাফল পাওয়া যায়নি।"
    return f"'{query}' এর search ফলাফল:\n\n" + "\n\n".join(results)


def fetch_webpage(url: str) -> str:
    try:
        resp = requests.get(
            url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"},
            timeout=20)
        resp.raise_for_status()
    except requests.RequestException as e:
        return f"পেজ আনতে সমস্যা: {e}"
    text = resp.text
    text = re.sub(r"<(script|style|noscript)[^>]*>.*?</\1>", " ", text,
                  flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return "পেজে কোনো লেখা পাওয়া যায়নি।"
    return text[:3000] + ("..." if len(text) > 3000 else "")


def _decl(name, desc, props=None, required=None):
    return {"name": name, "description": desc,
            "parameters": {"type": "object", "properties": props or {},
                           **({"required": required} if required else {})}}


TOOLS = {
    "calculator": {"func": calculator, "declaration": _decl(
        "calculator", "গাণিতিক হিসাব করে। যেমন: 2+2, sqrt(16), 2000*0.15",
        {"expression": {"type": "string", "description": "Python math expression"}},
        ["expression"])},
    "get_current_time": {"func": get_current_time, "declaration": _decl(
        "get_current_time", "বর্তমান তারিখ ও সময় জানায়।")},
    "save_note": {"func": save_note, "declaration": _decl(
        "save_note", "একটি note/text ফাইলে সংরক্ষণ করে।",
        {"filename": {"type": "string", "description": "ফাইলের নাম, যেমন 'todo.txt'"},
         "content": {"type": "string", "description": "যা সংরক্ষণ করতে হবে"}},
        ["filename", "content"])},
    "read_note": {"func": read_note, "declaration": _decl(
        "read_note", "আগে সংরক্ষণ করা note ফাইল পড়ে।",
        {"filename": {"type": "string", "description": "ফাইলের নাম"}}, ["filename"])},
    "list_notes": {"func": list_notes, "declaration": _decl(
        "list_notes", "সব সংরক্ষিত note ফাইলের তালিকা দেখায়।")},
    "remember_fact": {"func": add_fact, "declaration": _decl(
        "remember_fact",
        "গুরুত্বপূর্ণ তথ্য স্থায়ীভাবে মনে রাখে (restart করলেও ভুলবে না)। "
        "User যখন নিজের নাম, পছন্দ, জন্মদিন, পেশা, লক্ষ্য জানায় তখন ব্যবহার করো।",
        {"text": {"type": "string", "description": "যে তথ্যটা মনে রাখতে হবে (এক লাইনে)"}},
        ["text"])},
    "list_facts": {"func": lambda: facts_as_text() or "এখনো কোনো তথ্য স্থায়ীভাবে মনে রাখা হয়নি।",
                   "declaration": _decl("list_facts", "স্থায়ীভাবে মনে রাখা সব তথ্যের তালিকা দেখায়।")},
    "forget_fact": {"func": delete_fact, "declaration": _decl(
        "forget_fact", "স্থায়ী স্মৃতি থেকে একটি তথ্য মুছে ফেলে। আগে list_facts দিয়ে নম্বর দেখে নাও।",
        {"number": {"type": "integer", "description": "কত নম্বর তথ্য মুছতে হবে (1 থেকে শুরু)"}},
        ["number"])},
    "create_plan": {"func": create_plan, "declaration": _decl(
        "create_plan",
        "জটিল/বড় কাজের জন্য ধাপে ধাপে plan তৈরি করে। ৩টির বেশি ধাপ বা একাধিক tool "
        "লাগলে আগে plan বানাও, তারপর এক এক করে execute করো। সহজ প্রশ্নে দরকার নেই।",
        {"goal": {"type": "string", "description": "মূল লক্ষ্য, এক লাইনে"},
         "steps": {"type": "array", "items": {"type": "string"},
                   "description": "ধাপের তালিকা"}},
        ["goal", "steps"])},
    "complete_step": {"func": complete_step, "declaration": _decl(
        "complete_step", "Plan-এর একটি ধাপ সম্পন্ন হলে টিক দাও।",
        {"step_number": {"type": "integer", "description": "কত নম্বর ধাপ শেষ হলো"},
         "result_summary": {"type": "string", "description": "এই ধাপে কী পাওয়া গেল"}},
        ["step_number"])},
    "view_plan": {"func": view_plan, "declaration": _decl(
        "view_plan", "বর্তমান plan-এর অবস্থা দেখায় — কোন ধাপ শেষ, কোনটা বাকি।")},
    "search_documents": {"func": docs_search, "declaration": _decl(
        "search_documents",
        "User-এর নিজের document লাইব্রেরিতে (PDF/TXT/MD) খোঁজে। User যখন 'আমার "
        "ফাইলে/PDF-এ কী আছে' ধরনের প্রশ্ন করে তখন ব্যবহার করো। উত্তরে উৎস উল্লেখ করো।",
        {"query": {"type": "string", "description": "কী খুঁজতে হবে (মূল শব্দগুলো)"}},
        ["query"])},
    "list_documents": {"func": list_documents, "declaration": _decl(
        "list_documents", "User-এর document লাইব্রেরির ফাইলের তালিকা দেখায়।")},
    "reindex_documents": {"func": build_index, "declaration": _decl(
        "reindex_documents", "Document লাইব্রেরির index নতুন করে বানায়। নতুন ফাইল যোগ হলে চালাও।")},
    "run_python_code": {"func": run_python, "declaration": _decl(
        "run_python_code",
        "Python code লিখে চালায়, output ফেরত দেয়। জটিল হিসাব, data analysis, "
        "chart তৈরি — programming লাগলেই ব্যবহার করো। ফলাফল অবশ্যই print() করবে। "
        "Chart: matplotlib দিয়ে plt.savefig('name.png')। সময়সীমা ১৫ সেকেন্ড।",
        {"code": {"type": "string", "description": "সম্পূর্ণ Python code (print সহ)"}},
        ["code"])},
    "list_sandbox_files": {"func": list_sandbox_files, "declaration": _decl(
        "list_sandbox_files", "Code দিয়ে তৈরি ফাইলগুলোর তালিকা (sandbox ফোল্ডার)।")},
    "web_search": {"func": web_search, "declaration": _decl(
        "web_search",
        "ইন্টারনেটে খোঁজে (DuckDuckGo)। সাম্প্রতিক খবর, দাম, আবহাওয়া বা অজানা তথ্যের "
        "জন্য ব্যবহার করো। English query দিলে ভালো ফলাফল আসে।",
        {"query": {"type": "string", "description": "Search query"}}, ["query"])},
    "fetch_webpage": {"func": fetch_webpage, "declaration": _decl(
        "fetch_webpage",
        "একটি ওয়েবপেজের সম্পূর্ণ লেখা পড়ে আনে। web_search-এর URL থেকে বিস্তারিত জানতে ব্যবহার করো।",
        {"url": {"type": "string", "description": "সম্পূর্ণ URL"}}, ["url"])},
}


def get_declarations():
    return [t["declaration"] for t in TOOLS.values()]


def execute_tool(name: str, args: dict) -> str:
    if name not in TOOLS:
        return f"Error: '{name}' নামে কোনো tool নেই।"
    try:
        return str(TOOLS[name]["func"](**args))
    except Exception as e:
        return f"Tool error: {e}"


# ═══════════════════════════════════════════════════════════════════
# ৭. LLM CLIENT — Gemini API (auto-retry + auto-fallback সহ)
# ═══════════════════════════════════════════════════════════════════

API_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"


class LLMError(Exception):
    pass


def llm_chat(messages, tools=None):
    if not GEMINI_API_KEY:
        raise LLMError(
            "GEMINI_API_KEY পাওয়া যায়নি!\n"
            "১. https://aistudio.google.com/apikey এ যান (সম্পূর্ণ ফ্রি)\n"
            "২. 'Create API key' চাপুন\n"
            "৩. .env ফাইলে লিখুন:  GEMINI_API_KEY=আপনার_key"
        )
    # non-Gemini provider বাছা থাকলে OpenAI-compatible adapter-এ যাই
    if get_engine()["provider"] != "gemini":
        return _openai_chat(messages, tools)
    system_text = build_system_prompt() + facts_as_text()
    payload = {
        "system_instruction": {"parts": [{"text": system_text}]},
        "contents": messages,
        "generationConfig": {"temperature": 0.7},
    }
    if tools:
        payload["tools"] = [{"function_declarations": tools}]

    _primary = get_engine()["model"] if get_engine()["provider"] == "gemini" else MODEL_NAME
    models_to_try = [_primary] + [m for m in FALLBACK_MODELS if m != _primary]
    resp = None
    for model in models_to_try:
        url = API_URL.format(model=model)
        for attempt in range(2):
            try:
                resp = requests.post(url, params={"key": GEMINI_API_KEY},
                                     json=payload, timeout=120)
            except requests.Timeout:
                raise LLMError("⏳ Google-এর server-এ ভিড় (timeout)। সম্ভবত আজকের ফ্রি কোটা শেষ বা সাময়িক সমস্যা — কয়েক মিনিট পরে আবার চেষ্টা করুন।")
            except requests.RequestException as e:
                raise LLMError(f"Network error: {e}")
            if resp.status_code == 200:
                track_usage(model)
                if model != _primary:
                    print(f"  🔀 [fallback] '{model}' মডেল ব্যবহার হলো (মূল মডেলের কোটা শেষ)")
                break
            if resp.status_code == 429 and attempt == 0:
                print(f"  ⏳ '{model}' rate limited — ২০ সেকেন্ড অপেক্ষা...")
                time.sleep(20)
            else:
                break
        if resp.status_code == 200:
            break
        if resp.status_code in (429, 404):
            continue
        break

    if resp.status_code == 429:
        raise LLMError("সব মডেলের ফ্রি কোটা আপাতত শেষ! কিছুক্ষণ (বা আগামীকাল) পরে আবার চেষ্টা করুন।")
    if resp.status_code != 200:
        raise LLMError(f"API error {resp.status_code}: {resp.text[:300]}")
    data = resp.json()
    try:
        return data["candidates"][0]["content"]
    except (KeyError, IndexError):
        raise LLMError(f"Unexpected API response: {str(data)[:300]}")




# ══════════ AI USAGE TRACKING ══════════
USAGE_FILE = MEMORY_DIR / "usage.json"

def _load_usage():
    if USAGE_FILE.exists():
        try:
            return json.loads(USAGE_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}

def track_usage(model):
    """প্রতিটা LLM call গোনা হয় — দিনভিত্তিক।"""
    u = _load_usage()
    today = datetime.date.today().isoformat()
    day = u.setdefault(today, {"calls": 0, "models": {}})
    day["calls"] += 1
    day["models"][model] = day["models"].get(model, 0) + 1
    u["total"] = u.get("total", 0) + 1
    # ৩০ দিনের বেশি পুরনো এন্ট্রি মুছে ফেলা
    for k in [k for k in u if k not in ("total",) and k < (datetime.date.today() - datetime.timedelta(days=30)).isoformat()]:
        del u[k]
    USAGE_FILE.write_text(json.dumps(u, ensure_ascii=False), encoding="utf-8")

# ══════════ AGENT DASHBOARD DATA ══════════
def _read_note_lines(fname, limit=5):
    p = WORKSPACE / fname
    if not p.exists():
        return []
    lines = [l.strip() for l in p.read_text(encoding="utf-8", errors="ignore").splitlines() if l.strip()]
    return lines[-limit:] if len(lines) > limit else lines

def dashboard_data():
    """Dashboard-এর ৯টা card-এর data তৈরি করে।"""
    now = datetime.datetime.now()
    cards = []

    # ১. Today's Tasks
    cards.append({"icon": "✅", "title": "Today's Tasks",
        "lines": _read_note_lines("tasks.txt", 6),
        "empty": "কোনো কাজ নেই। Tasks mode-এ গিয়ে যোগ করুন!"})

    # ২. Upcoming Events
    cards.append({"icon": "📅", "title": "Upcoming Events",
        "lines": _read_note_lines("events.txt", 5),
        "empty": "কোনো event নেই। Tasks mode-এ 'নতুন event যোগ করো' বলুন।"})

    # ৩. Recent Conversations
    recent = []
    try:
        hist = load_history()
        for m in reversed(hist):
            if m.get("role") == "user" and m.get("parts") and "text" in m["parts"][0]:
                t = m["parts"][0]["text"]
                t = re.sub(r"^\[[^\]]*\]\s*", "", t)  # mode prefix বাদ
                recent.append(("💬 " + t[:60] + ("…" if len(t) > 60 else "")))
                if len(recent) >= 5:
                    break
    except Exception:
        pass
    cards.append({"icon": "🗨️", "title": "Recent Conversations",
        "lines": recent, "empty": "এখনো কোনো কথোপকথন হয়নি।"})

    # ৪. Memory (facts)
    facts = load_facts()
    cards.append({"icon": "🧠", "title": "Memory",
        "lines": [f"• {f['text']}" for f in facts[-6:]],
        "empty": "এখনো কিছু মনে রাখা হয়নি। নিজের সম্পর্কে কিছু বলুন!"})

    # ৫. Running Automations (plan)
    plan_lines = []
    plan = _plan_load()
    if plan:
        done = sum(1 for s in plan["steps"] if s["done"])
        total = len(plan["steps"])
        plan_lines.append(f"📋 {plan['goal']}")
        plan_lines.append(f"অগ্রগতি: {done}/{total} ধাপ {'🎉 সম্পন্ন!' if done == total else '⏳ চলছে'}")
        for i, s in enumerate(plan["steps"][:4], 1):
            plan_lines.append(("✅" if s["done"] else "⬜") + f" {s['text'][:45]}")
    cards.append({"icon": "⚡", "title": "Running Automations",
        "lines": plan_lines, "empty": "কোনো automation চলছে না। Agents mode-এ বড় কাজ দিন!"})

    # ৬. Important Notifications
    notif = []
    u = _load_usage()
    today_calls = u.get(datetime.date.today().isoformat(), {}).get("calls", 0)
    if today_calls >= 60:
        notif.append("⚠️ আজ AI ব্যবহার অনেক বেশি — ফ্রি কোটা শেষ হতে পারে")
    if not GEMINI_API_KEY:
        notif.append("🔑 API key সেট করা হয়নি!")
    tasks_n = len(_read_note_lines("tasks.txt", 100))
    if tasks_n > 0:
        notif.append(f"📌 {tasks_n}টা কাজ তালিকায় আছে")
    if plan and sum(1 for s in plan["steps"] if s["done"]) < len(plan["steps"]):
        notif.append("⏳ একটা automation এখনো অসম্পূর্ণ")
    docs_n = len([f for f in DOCS_DIR.iterdir() if f.is_file()]) if DOCS_DIR.exists() else 0
    if docs_n:
        notif.append(f"📚 {docs_n}টা document লাইব্রেরিতে আছে")
    cards.append({"icon": "🔔", "title": "Important Notifications",
        "lines": notif, "empty": "সব ঠিকঠাক — কোনো জরুরি বিজ্ঞপ্তি নেই ✨"})

    # ৭. Email Summary
    email_lines = _read_note_lines("emails.txt", 4)
    n_drafts = len([l for l in _read_note_lines("emails.txt", 1000) if l.lower().startswith("subject")])
    summary = ([f"📮 মোট draft: {n_drafts}"] if n_drafts else []) + email_lines
    cards.append({"icon": "📧", "title": "Email Summary",
        "lines": summary, "empty": "কোনো email draft নেই। Email mode-এ লিখিয়ে নিন!"})

    # ৮. Business Statistics
    notes_n = len([f for f in WORKSPACE.iterdir() if f.is_file()])
    sandbox_n = len([f for f in SANDBOX.iterdir() if f.is_file() and f.name != "_run.py"])
    hist_n = 0
    try:
        hist_n = sum(1 for m in load_history() if m.get("role") == "user" and m.get("parts") and "text" in m["parts"][0])
    except Exception:
        pass
    _orders = _biz_load("orders")
    _tickets = _biz_load("tickets")
    _pending = sum(1 for o in _orders if o.get("status") == "pending")
    _open_t = sum(1 for t in _tickets if t.get("status") == "open")
    cards.append({"icon": "📈", "title": "Business Statistics", "lines": [
        f"🧾 মোট order: {len(_orders)} (pending: {_pending})",
        f"👥 Customer: {len(_biz_load('customers'))} জন",
        f"📦 পণ্য: {len(_biz_load('products'))}টা",
        f"🎫 Open ticket: {_open_t}",
        f"📝 Notes: {notes_n} | 📚 Docs: {docs_n} | 💬 প্রশ্ন: {hist_n}",
    ], "empty": ""})

    # ৯. AI Usage
    models_today = u.get(datetime.date.today().isoformat(), {}).get("models", {})
    usage_lines = [f"🔥 আজকের AI call: {today_calls}",
                   f"♾️ সর্বমোট: {u.get('total', 0)}"]
    for m, n in sorted(models_today.items(), key=lambda x: -x[1])[:3]:
        usage_lines.append(f"• {m}: {n}")
    est = max(0, 80 - today_calls)
    usage_lines.append(f"🎫 আনুমানিক ফ্রি কোটা বাকি: ~{est}")
    cards.append({"icon": "🤖", "title": "AI Usage", "lines": usage_lines, "empty": ""})

    return {"generated": now.strftime("%H:%M:%S"), "cards": cards}




# ══════════ ACTIVITY TRACKING (Command Center-এর জন্য) ══════════
ACTIVITY_FILE = MEMORY_DIR / "activity.json"
# প্রতিটা message ≈ কত মিনিটের কাজ ধরা হবে (আনুমানিক)
MINUTES_PER_MESSAGE = 6

def _load_activity():
    if ACTIVITY_FILE.exists():
        try:
            return json.loads(ACTIVITY_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}

def track_activity(mode):
    """কোন mode-এ কাজ হলো তা দিনভিত্তিক গোনা হয়।"""
    a = _load_activity()
    today = datetime.date.today().isoformat()
    day = a.setdefault(today, {})
    day[mode] = day.get(mode, 0) + 1
    cutoff = (datetime.date.today() - datetime.timedelta(days=30)).isoformat()
    for k in [k for k in a if k < cutoff]:
        del a[k]
    ACTIVITY_FILE.write_text(json.dumps(a, ensure_ascii=False), encoding="utf-8")

# Mode → Weekly Overview-এর কোন খাতে যাবে
MODE_CATEGORY = {
    "study": "Study",
    "business": "Business", "email": "Business",
    "code": "Projects", "agents": "Projects", "files": "Projects", "research": "Projects",
    "chat": "General", "think": "General", "web": "General", "tasks": "General",
}

def command_data():
    """Command Center-এর data: গত ৭ দিনের হিসাব।"""
    a = _load_activity()
    today = datetime.date.today()
    week_days = [(today - datetime.timedelta(days=i)).isoformat() for i in range(7)]

    cat_msgs = {}
    day_lines = []
    mode_totals = {}
    bn_days = {"Mon": "সোম", "Tue": "মঙ্গল", "Wed": "বুধ", "Thu": "বৃহঃ",
               "Fri": "শুক্র", "Sat": "শনি", "Sun": "রবি"}
    for d in week_days:
        day_data = a.get(d, {})
        total = sum(day_data.values())
        if total:
            dt = datetime.date.fromisoformat(d)
            dname = bn_days.get(dt.strftime("%a"), dt.strftime("%a"))
            day_lines.append(f"{dname} {dt.strftime('%d/%m')}: {total}টা কাজ")
        for mode, n in day_data.items():
            cat = MODE_CATEGORY.get(mode, "General")
            cat_msgs[cat] = cat_msgs.get(cat, 0) + n
            mode_totals[mode] = mode_totals.get(mode, 0) + n

    overview = []
    for cat in ("Study", "Business", "Projects", "General"):
        n = cat_msgs.get(cat, 0)
        hours = round(n * MINUTES_PER_MESSAGE / 60, 1)
        overview.append({"label": cat, "hours": hours, "msgs": n})

    # Tasks completed: tasks.txt-র ✅/[done] + plan-এর সম্পন্ন ধাপ
    completed = 0
    tp = WORKSPACE / "tasks.txt"
    if tp.exists():
        for l in tp.read_text(encoding="utf-8", errors="ignore").splitlines():
            low = l.strip().lower()
            if low and ("✅" in l or low.startswith("[done]") or low.startswith("done")):
                completed += 1
    plan = _plan_load()
    if plan:
        completed += sum(1 for s in plan["steps"] if s["done"])
    # গত ৭ দিনের মোট প্রশ্নও একটা কাজ হিসেবে
    week_msgs = sum(sum(a.get(d, {}).values()) for d in week_days)

    highlights = []
    if mode_totals:
        top = max(mode_totals.items(), key=lambda x: x[1])
        highlights.append(f"🏆 সবচেয়ে বেশি ব্যবহার: {top[0]} mode ({top[1]} বার)")
    highlights.append(f"💬 এই সপ্তাহে মোট কাজ: {week_msgs}")
    u = _load_usage()
    week_calls = sum(u.get(d, {}).get("calls", 0) for d in week_days)
    highlights.append(f"🤖 AI call: {week_calls}")
    active_days = sum(1 for d in week_days if a.get(d))
    highlights.append(f"🔥 Active দিন: {active_days}/7")

    return {
        "generated": datetime.datetime.now().strftime("%H:%M:%S"),
        "overview": overview,
        "tasks_completed": completed + week_msgs,
        "days": day_lines[:7],
        "highlights": highlights,
    }






# ╔═══════════════════════════════════════════════════════════════╗
# ║  🛒 BUSINESS SUITE — Products, Customers, Orders, Invoice,    ║
# ║     Support Tickets, Finance (সব JSON database-এ save হয়)     ║
# ╚═══════════════════════════════════════════════════════════════╝
BIZ_DIR = MEMORY_DIR / "business"
BIZ_DIR.mkdir(exist_ok=True)

def _biz_load(name):
    p = BIZ_DIR / f"{name}.json"
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def _biz_save(name, data):
    (BIZ_DIR / f"{name}.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

# ── 📦 PRODUCTS ──
def product_tool(action, name="", price=None, stock=None):
    items = _biz_load("products")
    action = (action or "list").lower()
    if action == "add":
        if not name.strip():
            return "Error: পণ্যের নাম দরকার।"
        items.append({"name": name.strip(), "price": float(price or 0), "stock": int(stock or 0)})
        _biz_save("products", items)
        return f"📦 পণ্য যোগ হলো: {name} — দাম {price or 0} টাকা, stock {stock or 0}টা (মোট পণ্য: {len(items)})"
    if action == "update":
        for it in items:
            if name.strip().lower() in it["name"].lower():
                if price is not None: it["price"] = float(price)
                if stock is not None: it["stock"] = int(stock)
                _biz_save("products", items)
                return f"📦 আপডেট: {it['name']} — দাম {it['price']} টাকা, stock {it['stock']}টা"
        return f"'{name}' নামে পণ্য পাওয়া যায়নি।"
    if action == "check":
        matches = [it for it in items if name.strip().lower() in it["name"].lower()]
        if not matches:
            return f"'{name}' পাওয়া যায়নি। মোট পণ্য: {len(items)}টা।"
        return "\n".join(f"📦 {it['name']}: দাম {it['price']} টাকা, stock {it['stock']}টা"
                          + (" ⚠️ stock কম!" if it['stock'] <= 3 else "") for it in matches)
    # list
    if not items:
        return "কোনো পণ্য নেই। 'add' action দিয়ে পণ্য যোগ করুন।"
    return "📦 পণ্য তালিকা:\n" + "\n".join(
        f"{i+1}. {it['name']} — {it['price']} টাকা (stock: {it['stock']})"
        for i, it in enumerate(items))

# ── 👥 CUSTOMERS ──
def customer_tool(action, name="", phone="", note=""):
    custs = _biz_load("customers")
    action = (action or "list").lower()
    if action == "add":
        if not name.strip():
            return "Error: customer-এর নাম দরকার।"
        for c in custs:
            if c["name"].lower() == name.strip().lower():
                if phone: c["phone"] = phone
                if note: c["note"] = note
                _biz_save("customers", custs)
                return f"👥 {name} আগে থেকেই আছে — তথ্য আপডেট হলো।"
        custs.append({"name": name.strip(), "phone": phone, "note": note,
                      "since": datetime.date.today().isoformat()})
        _biz_save("customers", custs)
        return f"👥 নতুন customer: {name} (মোট: {len(custs)} জন)"
    if action == "find":
        q = (name or phone).strip().lower()
        matches = [c for c in custs if q in c["name"].lower() or q in c.get("phone", "")]
        if not matches:
            return f"'{name or phone}' পাওয়া যায়নি।"
        orders = _biz_load("orders")
        out = []
        for c in matches:
            his = [o for o in orders if o["customer"].lower() == c["name"].lower()]
            total = sum(o.get("total", 0) for o in his)
            out.append(f"👥 {c['name']} | 📞 {c.get('phone','—')} | সদস্য: {c.get('since','—')}"
                       f"\n   Order: {len(his)}টা, মোট কেনা: {total} টাকা"
                       + (f"\n   📝 {c['note']}" if c.get('note') else ""))
        return "\n".join(out)
    if not custs:
        return "কোনো customer নেই।"
    orders = _biz_load("orders")
    lines = []
    for i, c in enumerate(custs, 1):
        n = sum(1 for o in orders if o["customer"].lower() == c["name"].lower())
        lines.append(f"{i}. {c['name']} ({c.get('phone','—')}) — {n}টা order" + (" ⭐ repeat!" if n >= 2 else ""))
    return f"👥 Customer তালিকা ({len(custs)} জন):\n" + "\n".join(lines)

# ── 🧾 ORDERS ──
def order_tool(action, customer="", items="", total=None, order_id="", status=""):
    orders = _biz_load("orders")
    action = (action or "list").lower()
    if action == "create":
        if not customer.strip() or not items.strip():
            return "Error: customer-এর নাম আর পণ্যের বিবরণ দরকার।"
        oid = f"ORD-{datetime.date.today().strftime('%y%m%d')}-{len(orders)+1:03d}"
        order = {"id": oid, "customer": customer.strip(), "items": items.strip(),
                 "total": float(total or 0), "status": "pending",
                 "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}
        orders.append(order)
        _biz_save("orders", orders)
        # customer auto-register
        customer_tool("add", name=customer)
        log_activity("order", f"নতুন order {oid}: {customer} — {total or 0} টাকা")
        return (f"✅ Order তৈরি!\n🧾 Order ID: {oid}\n👥 Customer: {customer}"
                f"\n📦 পণ্য: {items}\n💰 মোট: {total or 0} টাকা\n📌 Status: pending")
    if action == "status":
        for o in orders:
            if o["id"] == order_id.strip().upper() or (order_id and order_id in o["id"]):
                if status:
                    o["status"] = status.strip().lower()
                    _biz_save("orders", orders)
                    log_activity("order", f"{o['id']} → {status}")
                    return f"🧾 {o['id']} এর status এখন: {o['status']}"
                return (f"🧾 {o['id']}\n👥 {o['customer']}\n📦 {o['items']}"
                        f"\n💰 {o['total']} টাকা\n📌 Status: {o['status']}\n📅 {o['date']}")
        return f"'{order_id}' order পাওয়া যায়নি।"
    if action == "cancel":
        return order_tool("status", order_id=order_id, status="cancelled")
    # list
    if not orders:
        return "কোনো order নেই।"
    recent = orders[-10:]
    return f"🧾 Order তালিকা (শেষ {len(recent)}টা / মোট {len(orders)}):\n" + "\n".join(
        f"• {o['id']} | {o['customer']} | {o['total']} টাকা | {o['status']}" for o in reversed(recent))

# ── 🧾 INVOICE ──
def make_invoice(order_id):
    orders = _biz_load("orders")
    for o in orders:
        if o["id"] == order_id.strip().upper() or order_id in o["id"]:
            inv = (f"{'='*38}\n          🧾 INVOICE\n       {BRAND['name']}\n{'='*38}\n"
                   f"Invoice No: INV-{o['id'][4:]}\nতারিখ: {datetime.date.today().isoformat()}\n"
                   f"Customer: {o['customer']}\n{'-'*38}\nবিবরণ: {o['items']}\n{'-'*38}\n"
                   f"মোট: {o['total']} টাকা\nStatus: {o['status']}\n{'='*38}\n"
                   f"ধন্যবাদ! আবার আসবেন 🙏")
            fname = f"invoice_{o['id']}.txt"
            (WORKSPACE / fname).write_text(inv, encoding="utf-8")
            log_activity("order", f"invoice তৈরি: {o['id']}")
            return f"🧾 Invoice তৈরি হয়ে '{fname}' ফাইলে save হলো:\n\n{inv}"
    return f"'{order_id}' order পাওয়া যায়নি।"

# ── 🎫 SUPPORT TICKETS ──
def ticket_tool(action, customer="", issue="", ticket_id="", status=""):
    tickets = _biz_load("tickets")
    action = (action or "list").lower()
    if action == "create":
        if not issue.strip():
            return "Error: সমস্যার বিবরণ দরকার।"
        tid = f"TKT-{len(tickets)+1:04d}"
        prio = "high" if any(w in issue.lower() for w in
            ["জরুরি", "urgent", "ভাঙা", "নষ্ট", "refund", "টাকা ফেরত"]) else "normal"
        tickets.append({"id": tid, "customer": customer.strip() or "অজানা", "issue": issue.strip(),
                        "status": "open", "priority": prio,
                        "date": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")})
        _biz_save("tickets", tickets)
        log_activity("ticket", f"{tid} ({prio}): {issue[:40]}")
        return f"🎫 Ticket তৈরি: {tid}\nPriority: {'🔴 High' if prio=='high' else '🟡 Normal'}\nStatus: open"
    if action == "update":
        for t in tickets:
            if t["id"] == ticket_id.strip().upper():
                t["status"] = status.strip().lower() or t["status"]
                _biz_save("tickets", tickets)
                return f"🎫 {t['id']} → {t['status']}"
        return f"'{ticket_id}' পাওয়া যায়নি।"
    open_t = [t for t in tickets if t["status"] == "open"]
    if not tickets:
        return "কোনো ticket নেই।"
    return f"🎫 Tickets (open: {len(open_t)}/{len(tickets)}):\n" + "\n".join(
        f"• {t['id']} [{'🔴' if t['priority']=='high' else '🟡'}] {t['customer']}: {t['issue'][:45]} ({t['status']})"
        for t in reversed(tickets[-10:]))

# ── 💰 FINANCE ──
def finance_tool(action, amount=None, note="", category=""):
    txns = _biz_load("finance")
    action = (action or "report").lower()
    if action in ("income", "expense"):
        if amount is None:
            return "Error: টাকার পরিমাণ দরকার।"
        txns.append({"type": action, "amount": float(amount), "note": note,
                     "category": category or ("বিক্রি" if action == "income" else "খরচ"),
                     "date": datetime.date.today().isoformat()})
        _biz_save("finance", txns)
        log_activity("finance", f"{action}: {amount} টাকা ({note[:30]})")
        return f"💰 {'আয়' if action=='income' else 'খরচ'} লেখা হলো: {amount} টাকা — {note}"
    # report
    if not txns:
        return "কোনো লেনদেন নেই। 'income'/'expense' action দিয়ে লিখুন।"
    today = datetime.date.today()
    month = today.strftime("%Y-%m")
    m_txns = [t for t in txns if t["date"].startswith(month)]
    inc = sum(t["amount"] for t in m_txns if t["type"] == "income")
    exp = sum(t["amount"] for t in m_txns if t["type"] == "expense")
    all_inc = sum(t["amount"] for t in txns if t["type"] == "income")
    all_exp = sum(t["amount"] for t in txns if t["type"] == "expense")
    cats = {}
    for t in m_txns:
        if t["type"] == "expense":
            cats[t["category"]] = cats.get(t["category"], 0) + t["amount"]
    cat_lines = "\n".join(f"   • {k}: {v} টাকা" for k, v in sorted(cats.items(), key=lambda x: -x[1])[:5])
    return (f"💰 আর্থিক রিপোর্ট ({month}):\n"
            f"📈 এ মাসের আয়: {inc} টাকা\n📉 এ মাসের খরচ: {exp} টাকা\n"
            f"{'✅ লাভ' if inc-exp >= 0 else '❌ ক্ষতি'}: {abs(inc-exp)} টাকা\n"
            + (f"খরচের খাত:\n{cat_lines}\n" if cat_lines else "")
            + f"\n♾️ সর্বমোট: আয় {all_inc} / খরচ {all_exp} / নিট {all_inc-all_exp} টাকা"
            f"\n🧾 মোট লেনদেন: {len(txns)}টা")

TOOLS.update({
    "product_tool": {"func": product_tool, "declaration": _decl(
        "product_tool",
        "পণ্যের database: action='add' (নতুন পণ্য), 'list' (সব দেখা), 'check' (দাম/stock দেখা), 'update' (দাম/stock বদলানো)। দাম/stock প্রশ্নে এটা ব্যবহার করো।",
        {"action": {"type": "string", "description": "add / list / check / update"},
         "name": {"type": "string", "description": "পণ্যের নাম"},
         "price": {"type": "number", "description": "দাম (টাকা)"},
         "stock": {"type": "integer", "description": "কয়টা আছে"}},
        ["action"])},
    "customer_tool": {"func": customer_tool, "declaration": _decl(
        "customer_tool",
        "Customer database: action='add' (নতুন/আপডেট), 'list' (সবাই + repeat customer), 'find' (খোঁজা + purchase history)।",
        {"action": {"type": "string", "description": "add / list / find"},
         "name": {"type": "string", "description": "নাম"},
         "phone": {"type": "string", "description": "ফোন নম্বর"},
         "note": {"type": "string", "description": "নোট"}},
        ["action"])},
    "order_tool": {"func": order_tool, "declaration": _decl(
        "order_tool",
        "Order management: action='create' (নতুন order, auto Order-ID), 'list', 'status' (দেখা/বদলানো: pending/confirmed/shipped/delivered), 'cancel'। কেউ কিছু কিনতে চাইলে create করো।",
        {"action": {"type": "string", "description": "create / list / status / cancel"},
         "customer": {"type": "string", "description": "customer-এর নাম"},
         "items": {"type": "string", "description": "কী কী পণ্য, কয়টা"},
         "total": {"type": "number", "description": "মোট টাকা"},
         "order_id": {"type": "string", "description": "Order ID (যেমন ORD-260826-001)"},
         "status": {"type": "string", "description": "নতুন status"}},
        ["action"])},
    "make_invoice": {"func": make_invoice, "declaration": _decl(
        "make_invoice", "Order ID দিয়ে invoice/রসিদ বানায় এবং ফাইলে save করে।",
        {"order_id": {"type": "string", "description": "Order ID"}}, ["order_id"])},
    "ticket_tool": {"func": ticket_tool, "declaration": _decl(
        "ticket_tool",
        "Customer support ticket: action='create' (অভিযোগ/সমস্যা এলে, জরুরি শব্দ থাকলে auto high-priority), 'list', 'update' (status: open/working/solved)।",
        {"action": {"type": "string", "description": "create / list / update"},
         "customer": {"type": "string", "description": "customer-এর নাম"},
         "issue": {"type": "string", "description": "সমস্যার বিবরণ"},
         "ticket_id": {"type": "string", "description": "Ticket ID"},
         "status": {"type": "string", "description": "open / working / solved"}},
        ["action"])},
    "finance_tool": {"func": finance_tool, "declaration": _decl(
        "finance_tool",
        "আয়-ব্যয়ের হিসাব: action='income' (আয় লেখা), 'expense' (খরচ লেখা), 'report' (মাসিক লাভ-ক্ষতি রিপোর্ট)।",
        {"action": {"type": "string", "description": "income / expense / report"},
         "amount": {"type": "number", "description": "টাকার পরিমাণ"},
         "note": {"type": "string", "description": "কীসের জন্য"},
         "category": {"type": "string", "description": "খাত (যেমন: বিক্রি, ভাড়া, বিদ্যুৎ)"}},
        ["action"])},
})


# ╔═══════════════════════════════════════════════════════════════╗
# ║  🧩 LONG-TERM FEATURES: Goals, Permissions, Log, Autos, Backup ║
# ╚═══════════════════════════════════════════════════════════════╝
FAILED_ATTEMPTS = {}  # brute-force সুরক্ষা: ip -> [timestamps]

# ══════════ 📜 ACTIVITY LOG ══════════
LOG_FILE = MEMORY_DIR / "activity_log.json"

def log_activity(kind, text):
    """Agent-এর প্রতিটা কাজের history রাখে (শেষ ৩০০টা)।"""
    try:
        log = []
        if LOG_FILE.exists():
            log = json.loads(LOG_FILE.read_text(encoding="utf-8"))
        log.append({"t": datetime.datetime.now().strftime("%d/%m %H:%M"),
                    "kind": kind, "text": str(text)[:160]})
        LOG_FILE.write_text(json.dumps(log[-300:], ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass

def load_log():
    if LOG_FILE.exists():
        try:
            return json.loads(LOG_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

# ══════════ 🛡️ PERMISSION SYSTEM ══════════
PERMS_FILE = MEMORY_DIR / "permissions.json"
DEFAULT_PERMS = {"web": True, "code": True, "files": True, "memory": True, "business": True, "approval_mode": False}
# কোন tool-এর জন্য কোন অনুমতি লাগে
TOOL_PERMISSION = {
    "web_search": "web", "fetch_webpage": "web",
    "run_python_code": "code",
    "save_note": "files", "read_note": "files", "list_notes": "files",
    "search_documents": "files", "reindex_documents": "files",
    "remember_fact": "memory", "forget_fact": "memory",
    "product_tool": "business", "customer_tool": "business", "order_tool": "business",
    "make_invoice": "business", "ticket_tool": "business", "finance_tool": "business",
}

def get_perms():
    if PERMS_FILE.exists():
        try:
            p = json.loads(PERMS_FILE.read_text(encoding="utf-8"))
            return {**DEFAULT_PERMS, **p}
        except (json.JSONDecodeError, OSError):
            pass
    return dict(DEFAULT_PERMS)

def set_perm(key, value):
    p = get_perms()
    if key in DEFAULT_PERMS:
        p[key] = bool(value)
        PERMS_FILE.write_text(json.dumps(p, ensure_ascii=False), encoding="utf-8")
        log_activity("perm", f"{key} -> {'ON' if value else 'OFF'}")
    return p

# Permission-check + logging সহ execute_tool wrapper
_orig_execute_tool = execute_tool
def execute_tool(name, args):
    perm = TOOL_PERMISSION.get(name)
    if perm and not get_perms().get(perm, True):
        log_activity("blocked", f"{name} — '{perm}' permission বন্ধ")
        return f"⛔ অনুমতি নেই! '{perm}' permission বন্ধ করা আছে। Settings (⚙️)-এ গিয়ে চালু করুন।"
    result = _orig_execute_tool(name, args)
    log_activity("tool", f"{name}({json.dumps(args, ensure_ascii=False)[:60]}) → {str(result)[:70]}")
    return result

# ══════════ 🎯 GOAL SYSTEM ══════════
GOALS_FILE = MEMORY_DIR / "goals.json"

def load_goals():
    if GOALS_FILE.exists():
        try:
            return json.loads(GOALS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def save_goals(goals):
    GOALS_FILE.write_text(json.dumps(goals, ensure_ascii=False, indent=1), encoding="utf-8")

def goal_add(title, deadline=""):
    goals = load_goals()
    goals.append({"title": title.strip(), "progress": 0, "deadline": deadline.strip(),
                  "done": False, "created": datetime.date.today().isoformat()})
    save_goals(goals)
    log_activity("goal", f"নতুন লক্ষ্য: {title[:50]}")
    return f"🎯 লক্ষ্য যোগ হলো: \"{title}\" (মোট {len(goals)}টি)"

def goal_update(number, progress=None, done=None):
    goals = load_goals()
    try:
        number = int(number)
    except (TypeError, ValueError):
        return "Error: number একটা সংখ্যা হতে হবে।"
    i = number - 1
    if not (0 <= i < len(goals)):
        return f"লক্ষ্য #{number} নেই। মোট {len(goals)}টি আছে।"
    if progress is not None:
        goals[i]["progress"] = max(0, min(100, int(progress)))
        if goals[i]["progress"] >= 100:
            goals[i]["done"] = True
    if done is not None:
        goals[i]["done"] = bool(done)
        if done:
            goals[i]["progress"] = 100
    save_goals(goals)
    log_activity("goal", f"লক্ষ্য #{number} আপডেট: {goals[i]['progress']}%")
    return f"🎯 \"{goals[i]['title']}\" এখন {goals[i]['progress']}% {'✅ সম্পন্ন!' if goals[i]['done'] else ''}"

def goals_as_text():
    goals = load_goals()
    if not goals:
        return "কোনো লক্ষ্য সেট করা নেই।"
    lines = []
    for i, g in enumerate(goals, 1):
        mark = "✅" if g.get("done") else f"{g.get('progress', 0)}%"
        dl = f" (⏰ {g['deadline']})" if g.get("deadline") else ""
        lines.append(f"{i}. [{mark}] {g['title']}{dl}")
    return "🎯 আমার লক্ষ্য:\n" + "\n".join(lines)

# Agent নিজেও goal manage করতে পারবে (নতুন ৩টা tool)
TOOLS.update({
    "add_goal": {"func": goal_add, "declaration": _decl(
        "add_goal", "User-এর নতুন লক্ষ্য (goal) যোগ করে। User লক্ষ্য/টার্গেটের কথা বললে ব্যবহার করো।",
        {"title": {"type": "string", "description": "লক্ষ্যের নাম"},
         "deadline": {"type": "string", "description": "শেষ তারিখ YYYY-MM-DD (ঐচ্ছিক)"}},
        ["title"])},
    "update_goal": {"func": goal_update, "declaration": _decl(
        "update_goal", "লক্ষ্যের অগ্রগতি (%) আপডেট করে বা সম্পন্ন চিহ্নিত করে।",
        {"number": {"type": "integer", "description": "কত নম্বর লক্ষ্য (1 থেকে)"},
         "progress": {"type": "integer", "description": "অগ্রগতি 0-100"}},
        ["number"])},
    "list_goals": {"func": goals_as_text, "declaration": _decl(
        "list_goals", "User-এর সব লক্ষ্য ও অগ্রগতি দেখায়।")},
})

# ══════════ 🔄 AUTOMATION BUILDER ══════════
AUTOS_FILE = MEMORY_DIR / "automations.json"

def load_autos():
    if AUTOS_FILE.exists():
        try:
            return json.loads(AUTOS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def save_autos(autos):
    AUTOS_FILE.write_text(json.dumps(autos, ensure_ascii=False, indent=1), encoding="utf-8")

# ══════════ 💾 BACKUP & EXPORT ══════════
def backup_bundle():
    notes = {}
    for f in WORKSPACE.iterdir():
        if f.is_file():
            try:
                notes[f.name] = f.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                pass
    log_activity("backup", "backup export করা হলো")
    return {
        "version": 1, "app": BRAND["name"],
        "date": datetime.datetime.now().isoformat(),
        "facts": load_facts(), "history": load_history(),
        "goals": load_goals(), "autos": load_autos(),
        "perms": get_perms(), "notes": notes,
        "activity": _load_activity(), "usage": _load_usage(),
        "business": {name: _biz_load(name) for name in
                     ("products", "customers", "orders", "tickets", "finance")},
        "watchlist": _watch_load(),
        "team": _team_load(),
        "config": _custom_load(),
        "reminders": _rem_load(),
        "plugins": {f.stem: {"code": f.read_text(encoding="utf-8"),
                             "meta": (PLUGINS_DIR / (f.stem + ".json")).read_text(encoding="utf-8")
                                     if (PLUGINS_DIR / (f.stem + ".json")).exists() else "{}"}
                    for f in PLUGINS_DIR.glob("*.py")},
    }

def restore_bundle(data):
    if not isinstance(data, dict):
        return False
    if isinstance(data.get("facts"), list):
        save_facts(data["facts"])
    if isinstance(data.get("history"), list):
        save_history(data["history"])
    if isinstance(data.get("goals"), list):
        save_goals(data["goals"])
    if isinstance(data.get("autos"), list):
        save_autos(data["autos"])
    if isinstance(data.get("perms"), dict):
        PERMS_FILE.write_text(json.dumps({**DEFAULT_PERMS, **data["perms"]}, ensure_ascii=False), encoding="utf-8")
    for name, txt in (data.get("notes") or {}).items():
        try:
            (WORKSPACE / Path(name).name).write_text(str(txt), encoding="utf-8")
        except OSError:
            pass
    if isinstance(data.get("activity"), dict):
        ACTIVITY_FILE.write_text(json.dumps(data["activity"], ensure_ascii=False), encoding="utf-8")
    if isinstance(data.get("watchlist"), list):
        _watch_save(data["watchlist"])
    if isinstance(data.get("team"), list):
        _team_save(data["team"])
    if isinstance(data.get("config"), dict):
        _custom_save(data["config"])
    if isinstance(data.get("reminders"), list):
        _rem_save(data["reminders"])
    if isinstance(data.get("plugins"), dict):
        for pname, pdata in data["plugins"].items():
            safe = re.sub(r"[^a-z0-9_]", "", str(pname).lower())
            if safe and isinstance(pdata, dict) and pdata.get("code"):
                (PLUGINS_DIR / f"{safe}.py").write_text(str(pdata["code"]), encoding="utf-8")
                (PLUGINS_DIR / f"{safe}.json").write_text(str(pdata.get("meta", "{}")), encoding="utf-8")
        load_plugins()
    if isinstance(data.get("business"), dict):
        for bname, bdata in data["business"].items():
            if isinstance(bdata, list) and bname in ("products", "customers", "orders", "tickets", "finance"):
                _biz_save(bname, bdata)
    log_activity("backup", "backup restore করা হলো")
    return True

# ══════════ ⚡ COMMAND CENTER EXTRA (priorities, alerts...) ══════════
def command_extra():
    urgent, imp, later = [], [], []
    tp = WORKSPACE / "tasks.txt"
    if tp.exists():
        for l in tp.read_text(encoding="utf-8", errors="ignore").splitlines():
            s = l.strip()
            if not s:
                continue
            if "🔴" in s or s.startswith("!!!"):
                urgent.append(s)
            elif "🟢" in s or "[later]" in s.lower():
                later.append(s)
            else:
                imp.append(s)
    upcoming = _read_note_lines("events.txt", 4)
    emails = [l for l in _read_note_lines("emails.txt", 100)
              if l.lower().startswith("subject")][:3]
    alerts = []
    u = _load_usage()
    today_calls = u.get(datetime.date.today().isoformat(), {}).get("calls", 0)
    if today_calls >= 60:
        alerts.append("⚠️ আজ AI কোটা প্রায় শেষের পথে")
    plan = _plan_load()
    if plan and not all(s["done"] for s in plan["steps"]):
        alerts.append("⏳ একটা automation/plan অসম্পূর্ণ")
    for g in load_goals():
        if not g.get("done") and g.get("deadline"):
            try:
                dl = datetime.date.fromisoformat(g["deadline"])
                if dl <= datetime.date.today() + datetime.timedelta(days=3):
                    alerts.append(f"🎯 Deadline কাছে: {g['title'][:35]}")
            except ValueError:
                pass
    if not get_perms().get("web", True):
        alerts.append("🛡️ Web permission বন্ধ আছে")
    _w = _watch_load()
    for _r in [r for r in _rem_load() if not r.get("notified")][:3]:
        alerts.append(f"⏰ {_r['when'][5:]} — {_r['text'][:30]}")
    _stale = [w for w in _w if not w.get("last_checked")
              or w["last_checked"][:10] < datetime.date.today().isoformat()]
    if _stale:
        alerts.append(f"👁️ {len(_stale)}টা watchlist item আজ check হয়নি")
    return {"priorities": {"urgent": urgent[:5], "important": imp[:5], "later": later[:5]},
            "upcoming": upcoming, "emails": emails, "alerts": alerts[:5]}



# ╔═══════════════════════════════════════════════════════════════╗
# ║  📊 DATA & DOCUMENTS — CSV/Excel analysis, compare, report    ║
# ╚═══════════════════════════════════════════════════════════════╝
import csv as _csv

def _find_data_file(filename):
    """documents/ বা agent_files/ থেকে ফাইল খোঁজে।"""
    name = Path(filename).name
    for d in (DOCS_DIR, WORKSPACE, SANDBOX):
        p = d / name
        if p.exists():
            return p
    # আংশিক নাম মিললেও চলবে
    for d in (DOCS_DIR, WORKSPACE, SANDBOX):
        for f in d.iterdir():
            if f.is_file() and name.lower() in f.name.lower():
                return f
    return None

def _load_table(path):
    """CSV/Excel → (headers, rows) হিসেবে পড়ে।"""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            return None, "Excel পড়তে openpyxl লাগবে (pip install openpyxl)"
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if c is None else c) for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        text = path.read_text(encoding="utf-8", errors="ignore")
        delim = "\t" if "\t" in text.split("\n")[0] else ","
        rows = list(_csv.reader(text.splitlines(), delimiter=delim))
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return None, "ফাইলটা খালি।"
    return (rows[0], rows[1:]), None

def analyze_data(filename, question=""):
    """CSV/Excel ফাইলের সম্পূর্ণ বিশ্লেষণ: গঠন, পরিসংখ্যান, duplicate।"""
    p = _find_data_file(filename)
    if not p:
        files = [f.name for d in (DOCS_DIR, WORKSPACE) for f in d.iterdir()
                 if f.suffix.lower() in (".csv", ".xlsx", ".txt")]
        return f"'{filename}' পাওয়া যায়নি। আছে: {', '.join(files) or 'কিছু নেই'}\n📎 বাটন দিয়ে upload করুন।"
    table, err = _load_table(p)
    if err:
        return err
    headers, rows = table
    n = len(rows)
    out = [f"📊 {p.name} — {n}টা row, {len(headers)}টা column",
           f"Columns: {', '.join(str(h) for h in headers)}", ""]
    # প্রতি column-এর ধরন + পরিসংখ্যান
    for i, h in enumerate(headers):
        vals = [r[i] for r in rows if i < len(r) and str(r[i]).strip()]
        nums = []
        for v in vals:
            try:
                nums.append(float(str(v).replace(",", "")))
            except ValueError:
                pass
        if nums and len(nums) >= len(vals) * 0.7:
            out.append(f"🔢 {h}: যোগফল={sum(nums):,.0f}, গড়={sum(nums)/len(nums):,.1f}, "
                       f"সর্বোচ্চ={max(nums):,.0f}, সর্বনিম্ন={min(nums):,.0f}")
        else:
            uniq = {}
            for v in vals:
                uniq[str(v)] = uniq.get(str(v), 0) + 1
            top = sorted(uniq.items(), key=lambda x: -x[1])[:3]
            out.append(f"🏷️ {h}: {len(uniq)}টা ভিন্ন মান | বেশি: " +
                       ", ".join(f"{k}({v})" for k, v in top))
    # duplicate detection
    seen, dups = {}, 0
    for r in rows:
        key = tuple(str(c).strip().lower() for c in r)
        if key in seen:
            dups += 1
        seen[key] = 1
    out.append("")
    out.append(f"🔁 Duplicate row: {dups}টা" + (" ⚠️" if dups else " ✓"))
    # নমুনা
    out.append("\nপ্রথম ৩ row:")
    for r in rows[:3]:
        out.append("  " + " | ".join(str(c)[:18] for c in r[:6]))
    if question:
        out.append(f"\n(প্রশ্ন: '{question}' — উপরের data দেখে উত্তর দাও, দরকারে query_data দিয়ে filter/sort করো)")
    log_activity("data", f"analyze: {p.name} ({n} rows)")
    return "\n".join(out)

def query_data(filename, column="", action="top", value="", limit=10):
    """Data filter/sort/গোনা: action='top' (বড়→ছোট), 'bottom', 'filter' (মান মিলিয়ে), 'count'।"""
    p = _find_data_file(filename)
    if not p:
        return f"'{filename}' পাওয়া যায়নি।"
    table, err = _load_table(p)
    if err:
        return err
    headers, rows = table
    try:
        limit = max(1, min(30, int(limit)))
    except (TypeError, ValueError):
        limit = 10
    hnames = [str(h).lower() for h in headers]
    ci = 0
    if column:
        cl = str(column).lower()
        matches = [i for i, h in enumerate(hnames) if cl in h]
        if not matches:
            return f"'{column}' column নেই। আছে: {', '.join(str(h) for h in headers)}"
        ci = matches[0]
    action = (action or "top").lower()
    if action == "count":
        return f"📊 {p.name}: মোট {len(rows)}টা row।"
    if action == "filter":
        hit = [r for r in rows if ci < len(r) and str(value).strip().lower() in str(r[ci]).lower()]
        if not hit:
            return f"'{value}' মিলে এমন row নেই।"
        out = [f"📊 '{value}' মিলেছে {len(hit)}টা row (দেখাচ্ছি {min(limit, len(hit))}):",
               " | ".join(str(h)[:15] for h in headers[:6])]
        out += ["  " + " | ".join(str(c)[:15] for c in r[:6]) for r in hit[:limit]]
        return "\n".join(out)
    # top/bottom — সংখ্যা ধরে sort
    def keyf(r):
        try:
            return float(str(r[ci]).replace(",", "")) if ci < len(r) else 0
        except ValueError:
            return float("-inf")
    s = sorted(rows, key=keyf, reverse=(action != "bottom"))
    out = [f"📊 {headers[ci]} অনুযায়ী {'সবচেয়ে বেশি' if action != 'bottom' else 'সবচেয়ে কম'} {limit}টা:",
           " | ".join(str(h)[:15] for h in headers[:6])]
    out += ["  " + " | ".join(str(c)[:15] for c in r[:6]) for r in s[:limit]]
    return "\n".join(out)

def make_chart(filename, label_column, value_column, chart_type="bar", title=""):
    """CSV/Excel data থেকে chart বানিয়ে PNG-তে save করে।"""
    p = _find_data_file(filename)
    if not p:
        return f"'{filename}' পাওয়া যায়নি।"
    table, err = _load_table(p)
    if err:
        return err
    headers, rows = table
    hnames = [str(h).lower() for h in headers]
    def col_idx(cname):
        cl = str(cname).lower()
        m = [i for i, h in enumerate(hnames) if cl in h]
        return m[0] if m else None
    li, vi = col_idx(label_column), col_idx(value_column)
    if li is None or vi is None:
        return f"Column পাওয়া যায়নি। আছে: {', '.join(str(h) for h in headers)}"
    labels, values = [], []
    for r in rows[:15]:
        if li < len(r) and vi < len(r):
            try:
                values.append(float(str(r[vi]).replace(",", "")))
                labels.append(str(r[li])[:14])
            except ValueError:
                pass
    if not values:
        return f"'{value_column}' column-এ সংখ্যা পাওয়া যায়নি।"
    safe = re.sub(r"[^A-Za-z0-9_]", "_", p.stem)[:20]
    fname = f"chart_{safe}.png"
    code = (
        "import matplotlib\nmatplotlib.use('Agg')\nimport matplotlib.pyplot as plt\n"
        f"labels = {labels!r}\nvalues = {values!r}\n"
        "plt.figure(figsize=(9,5))\n"
        + (f"plt.pie(values, labels=labels, autopct='%1.0f%%')\n" if chart_type == "pie"
           else f"plt.plot(labels, values, marker='o', color='#6366f1')\nplt.xticks(rotation=30, ha='right')\n" if chart_type == "line"
           else f"plt.bar(labels, values, color='#6366f1')\nplt.xticks(rotation=30, ha='right')\n")
        + f"plt.title({(title or p.stem)!r})\nplt.tight_layout()\nplt.savefig({fname!r})\nprint('saved')\n")
    r = run_python(code)
    if "saved" in r:
        log_activity("data", f"chart: {fname}")
        return f"📈 Chart তৈরি!\n\n📁 নতুন ফাইল তৈরি হয়েছে: {fname} (sandbox ফোল্ডারে)"
    return "Chart বানাতে সমস্যা:\n" + r[:300]

def compare_documents(file1, file2):
    """দুটো document/ফাইলের পার্থক্য তুলনা করে।"""
    p1, p2 = _find_data_file(file1), _find_data_file(file2)
    if not p1 or not p2:
        return f"ফাইল পাওয়া যায়নি: {file1 if not p1 else file2}"
    t1 = extract_text(p1)[:4000] if p1.suffix.lower() == ".pdf" else p1.read_text(encoding="utf-8", errors="ignore")[:4000]
    t2 = extract_text(p2)[:4000] if p2.suffix.lower() == ".pdf" else p2.read_text(encoding="utf-8", errors="ignore")[:4000]
    import difflib
    l1, l2 = t1.splitlines(), t2.splitlines()
    diff = list(difflib.unified_diff(l1, l2, lineterm="", n=1))[:60]
    sm = difflib.SequenceMatcher(None, t1, t2)
    sim = sm.ratio() * 100
    out = [f"📄 তুলনা: {p1.name} ↔ {p2.name}",
           f"মিল: {sim:.0f}%", ""]
    added = [l[1:].strip() for l in diff if l.startswith("+") and not l.startswith("+++")][:10]
    removed = [l[1:].strip() for l in diff if l.startswith("-") and not l.startswith("---")][:10]
    if added:
        out.append(f"➕ {p2.name}-এ নতুন/ভিন্ন ({len(added)}):")
        out += [f"   {a[:70]}" for a in added if a]
    if removed:
        out.append(f"➖ {p1.name}-এ ছিল কিন্তু {p2.name}-এ নেই ({len(removed)}):")
        out += [f"   {r[:70]}" for r in removed if r]
    if not added and not removed:
        out.append("✅ উল্লেখযোগ্য পার্থক্য নেই।")
    return "\n".join(out)

TOOLS.update({
    "analyze_data": {"func": analyze_data, "declaration": _decl(
        "analyze_data",
        "CSV/Excel ফাইলের সম্পূর্ণ বিশ্লেষণ: row/column, যোগফল-গড়-max-min, duplicate detection, নমুনা। "
        "User data/spreadsheet নিয়ে প্রশ্ন করলে আগে এটা চালাও।",
        {"filename": {"type": "string", "description": "ফাইলের নাম (documents-এ upload করা)"},
         "question": {"type": "string", "description": "user-এর মূল প্রশ্ন (ঐচ্ছিক)"}},
        ["filename"])},
    "query_data": {"func": query_data, "declaration": _decl(
        "query_data",
        "Data filter/sort: action='top' (কোনো column-এ সবচেয়ে বেশি), 'bottom' (কম), 'filter' (value মিলিয়ে খোঁজা), 'count'।",
        {"filename": {"type": "string", "description": "ফাইলের নাম"},
         "column": {"type": "string", "description": "কোন column"},
         "action": {"type": "string", "description": "top / bottom / filter / count"},
         "value": {"type": "string", "description": "filter-এর মান"},
         "limit": {"type": "integer", "description": "কয়টা row (default 10)"}},
        ["filename"])},
    "make_chart": {"func": make_chart, "declaration": _decl(
        "make_chart",
        "CSV/Excel data থেকে chart (bar/line/pie) বানিয়ে ছবি হিসেবে দেখায়। label_column=নামের column, value_column=সংখ্যার column।",
        {"filename": {"type": "string", "description": "data ফাইলের নাম"},
         "label_column": {"type": "string", "description": "label-এর column নাম"},
         "value_column": {"type": "string", "description": "সংখ্যার column নাম"},
         "chart_type": {"type": "string", "description": "bar / line / pie"},
         "title": {"type": "string", "description": "chart-এর শিরোনাম"}},
        ["filename", "label_column", "value_column"])},
    "compare_documents": {"func": compare_documents, "declaration": _decl(
        "compare_documents",
        "দুটো ফাইল/document তুলনা করে: কত % মিল, কী যোগ হয়েছে, কী বাদ গেছে।",
        {"file1": {"type": "string", "description": "প্রথম ফাইল"},
         "file2": {"type": "string", "description": "দ্বিতীয় ফাইল"}},
        ["file1", "file2"])},
})
TOOL_PERMISSION.update({
    "analyze_data": "files", "query_data": "files",
    "make_chart": "code", "compare_documents": "files",
})




# ╔═══════════════════════════════════════════════════════════════╗
# ║  📰 MONITORING & RESEARCH — Watchlist, site-change detection  ║
# ╚═══════════════════════════════════════════════════════════════╝
WATCH_FILE = MEMORY_DIR / "watchlist.json"

def _watch_load():
    if WATCH_FILE.exists():
        try:
            return json.loads(WATCH_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def _watch_save(items):
    WATCH_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")

def watch_tool(action, name="", watch_type="topic", query=""):
    """Watchlist manage: add / list / remove।"""
    items = _watch_load()
    action = (action or "list").lower()
    if action == "add":
        if not name.strip() or not query.strip():
            return "Error: name (ডাকনাম) আর query (কী নজরে রাখবো) দুটোই দরকার।"
        wt = (watch_type or "topic").lower()
        if wt not in ("topic", "news", "price", "competitor", "website"):
            wt = "topic"
        if wt == "website" and not query.strip().lower().startswith("http"):
            return "Error: website type-এ query-তে সম্পূর্ণ URL দিন (https:// সহ)।"
        # একই নাম থাকলে replace
        items = [i for i in items if i["name"].lower() != name.strip().lower()]
        items.append({"name": name.strip(), "type": wt, "query": query.strip(),
                      "added": datetime.date.today().isoformat(),
                      "last_checked": "", "last_hash": ""})
        _watch_save(items)
        log_activity("watch", f"নতুন watch: {name} ({wt})")
        return f"👁️ Watchlist-এ যোগ হলো: \"{name}\" ({wt}) — 'check করো' বললেই খবর আনবো। (মোট {len(items)}টা)"
    if action == "remove":
        before = len(items)
        items = [i for i in items if i["name"].lower() != name.strip().lower()]
        if len(items) == before:
            return f"'{name}' watchlist-এ নেই।"
        _watch_save(items)
        return f"👁️ '{name}' watchlist থেকে বাদ দেওয়া হলো। (বাকি {len(items)}টা)"
    if not items:
        return "Watchlist খালি। 'add' action দিয়ে topic/price/website যোগ করুন।"
    type_icon = {"topic": "📌", "news": "📰", "price": "💰", "competitor": "🏪", "website": "🌐"}
    return f"👁️ Watchlist ({len(items)}টা):\n" + "\n".join(
        f"{i+1}. {type_icon.get(it['type'],'📌')} {it['name']} [{it['type']}] — {it['query'][:45]}"
        + (f" (শেষ check: {it['last_checked']})" if it.get('last_checked') else " (কখনো check হয়নি)")
        for i, it in enumerate(items))

def check_watchlist(name=""):
    """Watchlist-এর item check করে — web search / site change detection।"""
    items = _watch_load()
    if not items:
        return "Watchlist খালি — আগে watch_tool দিয়ে কিছু যোগ করুন।"
    targets = [i for i in items if not name.strip()
               or name.strip().lower() in i["name"].lower()]
    if not targets:
        return f"'{name}' watchlist-এ পাওয়া যায়নি।"
    targets = targets[:3]  # কোটা বাঁচাতে একবারে সর্বোচ্চ ৩টা
    results = []
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for it in targets:
        try:
            if it["type"] == "website":
                page = fetch_webpage(it["query"])
                h = hashlib.md5(page[:2500].encode("utf-8", "ignore")).hexdigest()
                if it.get("last_hash") and it["last_hash"] != h:
                    status = "🔔 পরিবর্তন ধরা পড়েছে! পেজের শুরুর অংশ:\n" + page[:400]
                elif not it.get("last_hash"):
                    status = "প্রথম snapshot নেওয়া হলো। পেজের শুরু:\n" + page[:300]
                else:
                    status = "✓ কোনো পরিবর্তন নেই।"
                it["last_hash"] = h
                results.append(f"🌐 {it['name']} ({it['query'][:40]}):\n{status}")
            else:
                q = it["query"]
                if it["type"] == "price":
                    q += " price দাম"
                elif it["type"] == "news":
                    q += " latest news today"
                elif it["type"] == "competitor":
                    q += " latest updates"
                r = web_search(q)
                results.append(f"{'📰' if it['type']=='news' else '👁️'} {it['name']}:\n{r[:700]}")
            it["last_checked"] = now
        except Exception as e:
            results.append(f"⚠️ {it['name']}: check করতে সমস্যা — {e}")
    _watch_save(items)
    log_activity("watch", f"check: {', '.join(t['name'] for t in targets)}")
    return ("👁️ Watchlist check (" + now + "):\n\n" + "\n\n---\n\n".join(results)
            + "\n\n(এই ফলাফল দেখে user-কে সংক্ষেপে গুরুত্বপূর্ণ পয়েন্টগুলো জানাও)")

def save_report(title, content):
    """Research report ফাইলে save করে (reports ফোল্ডারে, তারিখসহ)।"""
    if not title.strip() or not content.strip():
        return "Error: title আর content দুটোই দরকার।"
    safe = re.sub(r"[^A-Za-z0-9\u0980-\u09FF_ -]", "", title)[:40].strip().replace(" ", "_")
    fname = f"report_{safe}_{datetime.date.today().isoformat()}.md"
    body = (f"# {title}\n\n📅 তারিখ: {datetime.date.today().isoformat()}"
            f"\n🤖 তৈরি: {BRAND['name']}\n\n---\n\n{content}")
    (WORKSPACE / fname).write_text(body, encoding="utf-8")
    log_activity("report", f"report save: {fname}")
    return f"📑 Report save হলো: '{fname}' ({len(content)} অক্ষর)। read_note দিয়ে পরে পড়া যাবে।"

TOOLS.update({
    "watch_tool": {"func": watch_tool, "declaration": _decl(
        "watch_tool",
        "Watchlist manage: action='add' (নতুন নজরদারি: type=topic/news/price/competitor/website), "
        "'list' (সব দেখা), 'remove' (বাদ)। User কিছু 'নজরে রাখতে/monitor করতে' চাইলে এটা।",
        {"action": {"type": "string", "description": "add / list / remove"},
         "name": {"type": "string", "description": "ডাকনাম, যেমন 'ডলারের দাম'"},
         "watch_type": {"type": "string", "description": "topic / news / price / competitor / website"},
         "query": {"type": "string", "description": "কী search হবে, বা website হলে URL"}},
        ["action"])},
    "check_watchlist": {"func": check_watchlist, "declaration": _decl(
        "check_watchlist",
        "Watchlist-এর item গুলো এখনই check করে (web search / website change detection)। "
        "name দিলে শুধু সেটা, না দিলে প্রথম ৩টা। ফলাফল পেয়ে সংক্ষেপে summary দাও।",
        {"name": {"type": "string", "description": "নির্দিষ্ট item-এর নাম (ঐচ্ছিক)"}}, [])},
    "save_report": {"func": save_report, "declaration": _decl(
        "save_report",
        "Research/analysis report ফাইলে save করে (তারিখসহ)। বড় research শেষে ফলাফল সংরক্ষণে ব্যবহার করো।",
        {"title": {"type": "string", "description": "report-এর শিরোনাম"},
         "content": {"type": "string", "description": "সম্পূর্ণ report (markdown)"}},
        ["title", "content"])},
})
TOOL_PERMISSION.update({
    "watch_tool": "web", "check_watchlist": "web", "save_report": "files",
})




# ╔═══════════════════════════════════════════════════════════════╗
# ║  ✋ ACTION APPROVAL — sensitive কাজে user-এর অনুমতি লাগে       ║
# ║  🔌 INTEGRATION SLOTS — ভবিষ্যতের paid API-র প্রস্তুত জায়গা    ║
# ╚═══════════════════════════════════════════════════════════════╝
APPROVALS_FILE = MEMORY_DIR / "approvals.json"

# Approval mode চালু থাকলে এই tool গুলো execute-এর আগে অনুমতি চাইবে
SENSITIVE_TOOLS = {"run_python_code", "forget_fact", "finance_tool", "make_invoice"}

def _approvals_load():
    if APPROVALS_FILE.exists():
        try:
            return json.loads(APPROVALS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def _approvals_save(items):
    APPROVALS_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")

def approval_mode_on():
    return get_perms().get("approval_mode", False)

# execute_tool-কে আবার wrap: approval layer (permission layer-এর উপরে)
_perm_execute_tool = execute_tool
def execute_tool(name, args):
    if approval_mode_on() and name in SENSITIVE_TOOLS:
        items = _approvals_load()
        aid = f"AP-{len(items)+1:03d}"
        items.append({"id": aid, "tool": name,
                      "args": args, "status": "pending",
                      "time": datetime.datetime.now().strftime("%d/%m %H:%M")})
        _approvals_save(items[-30:])
        log_activity("approval", f"অনুমতির অপেক্ষায়: {name} ({aid})")
        return (f"✋ '{name}' একটা sensitive কাজ — Approval mode চালু থাকায় এখনই চালাইনি। "
                f"অনুরোধ #{aid} জমা হয়েছে। User-কে বলো: Control (🛠️) → Pending Approvals-এ "
                f"গিয়ে ✅ Approve করলেই কাজটা হবে।")
    return _perm_execute_tool(name, args)

def run_approval(aid, approve=True):
    items = _approvals_load()
    for it in items:
        if it["id"] == aid and it["status"] == "pending":
            if not approve:
                it["status"] = "rejected"
                _approvals_save(items)
                log_activity("approval", f"বাতিল: {it['tool']} ({aid})")
                return {"ok": True, "result": f"❌ {aid} বাতিল করা হলো।"}
            result = _perm_execute_tool(it["tool"], it["args"])
            it["status"] = "approved"
            _approvals_save(items)
            log_activity("approval", f"অনুমোদিত ও চালানো হলো: {it['tool']} ({aid})")
            return {"ok": True, "result": f"✅ {aid} অনুমোদিত।\n\n{result}"}
    return {"ok": False, "result": f"'{aid}' pending অবস্থায় পাওয়া যায়নি।"}

# ── 🔌 INTEGRATION SLOTS (ভবিষ্যতের জন্য প্রস্তুত কাঠামো) ──
# এখন খরচ নেই বলে বন্ধ। ভবিষ্যতে চালু করতে:
#   ১. Secrets-এ token/key যোগ করুন (যেমন TELEGRAM_BOT_TOKEN)
#   ২. এখানে "enabled" check নিজে থেকেই সত্য হবে
#   ৩. সংশ্লিষ্ট send/read function লিখে TOOLS-এ register করুন
INTEGRATIONS = {
    "telegram": {"label": "📨 Telegram Bot", "env": "TELEGRAM_BOT_TOKEN",
                 "note": "ফ্রি! BotFather থেকে token নিয়ে Secrets-এ TELEGRAM_BOT_TOKEN দিন"},
    "gmail":    {"label": "📧 Gmail (আসল inbox)", "env": "GMAIL_API_CREDS",
                 "note": "Google Cloud OAuth লাগবে"},
    "sheets":   {"label": "📊 Google Sheets sync", "env": "GSHEETS_API_CREDS",
                 "note": "Google Cloud service account লাগবে"},
    "whatsapp": {"label": "💬 WhatsApp Business", "env": "WHATSAPP_API_TOKEN",
                 "note": "Meta Business API (paid)"},
    "payment":  {"label": "💳 Payment gateway", "env": "PAYMENT_API_KEY",
                 "note": "bKash/SSLCommerz merchant account লাগবে"},
}

def integrations_status():
    out = []
    for key, cfg in INTEGRATIONS.items():
        configured = bool(os.environ.get(cfg["env"], ""))
        out.append({"key": key, "label": cfg["label"], "env": cfg["env"],
                    "configured": configured, "note": cfg["note"]})
    return out




# ╔═══════════════════════════════════════════════════════════════╗
# ║  🔧 ADMIN SYSTEM — runtime customize + multi-AI + key vault   ║
# ╚═══════════════════════════════════════════════════════════════╝
DEFAULT_BRAND = dict(BRAND)  # reset-এর জন্য মূল কপি
CUSTOM_FILE = MEMORY_DIR / "custom_config.json"
KEYS_FILE = MEMORY_DIR / "api_keys.json"   # runtime key vault (backup-এ যায় না)

# 🤖 AI PROVIDERS — Gemini এখন চালু; বাকিগুলো key বসালেই চলবে
PROVIDERS = {
    "gemini": {"label": "Google Gemini (ফ্রি ✓ চালু)", "env": "GEMINI_API_KEY", "type": "gemini",
               "models": ["gemini-3.6-flash", "gemini-3.5-flash", "gemini-3.1-flash-lite", "gemini-flash-lite-latest"],
               "note": "এখনকার engine — সম্পূর্ণ ফ্রি"},
    "groq": {"label": "Groq (ফ্রি tier আছে!)", "env": "GROQ_API_KEY", "type": "openai",
             "base": "https://api.groq.com/openai/v1",
             "models": ["llama-3.3-70b-versatile", "llama-3.1-8b-instant"],
             "note": "console.groq.com থেকে ফ্রি key"},
    "openai": {"label": "OpenAI (ChatGPT)", "env": "OPENAI_API_KEY", "type": "openai",
               "base": "https://api.openai.com/v1",
               "models": ["gpt-4o-mini", "gpt-4o"],
               "note": "platform.openai.com — paid"},
    "openrouter": {"label": "OpenRouter (সব মডেল একসাথে)", "env": "OPENROUTER_API_KEY", "type": "openai",
                   "base": "https://openrouter.ai/api/v1",
                   "models": ["deepseek/deepseek-chat", "anthropic/claude-3.5-haiku", "google/gemini-2.0-flash-001"],
                   "note": "openrouter.ai — Claude/DeepSeek সহ সব, কিছু ফ্রি মডেলও আছে"},
    "deepseek": {"label": "DeepSeek", "env": "DEEPSEEK_API_KEY", "type": "openai",
                 "base": "https://api.deepseek.com/v1",
                 "models": ["deepseek-chat"],
                 "note": "platform.deepseek.com — সস্তা paid"},
}

def _custom_load():
    if CUSTOM_FILE.exists():
        try:
            return json.loads(CUSTOM_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}

def _custom_save(cfg):
    CUSTOM_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=1), encoding="utf-8")

def _keys_load():
    if KEYS_FILE.exists():
        try:
            return json.loads(KEYS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}

def save_runtime_key(env_name, value):
    """API key নিরাপদে save (শুধু server-side; UI-তে কখনো ফেরত যায় না)।"""
    allowed = {p["env"] for p in get_all_providers().values()} | {i["env"] for i in INTEGRATIONS.values()}
    if env_name not in allowed:
        return False
    keys = _keys_load()
    if value.strip():
        keys[env_name] = value.strip()
        os.environ[env_name] = value.strip()
    else:
        keys.pop(env_name, None)
        os.environ.pop(env_name, None)
    KEYS_FILE.write_text(json.dumps(keys), encoding="utf-8")
    log_activity("admin", f"API key আপডেট: {env_name}")
    return True

# startup-এ save করা key গুলো environment-এ লোড
for _k, _v in _keys_load().items():
    os.environ.setdefault(_k, _v)

# BRAND override apply (আগের save করা customize)
_cfg = _custom_load()
if isinstance(_cfg.get("brand"), dict):
    for _bk, _bv in _cfg["brand"].items():
        if _bk in BRAND:
            BRAND[_bk] = _bv


def get_all_providers():
    """Built-in + user-এর যোগ করা custom provider সব একসাথে।"""
    provs = dict(PROVIDERS)
    cfg = _custom_load()
    for cp in cfg.get("custom_providers", []):
        key = cp.get("key")
        if key and key not in provs and cp.get("base") and cp.get("env"):
            provs[key] = {"label": cp.get("label", key), "env": cp["env"], "type": "openai",
                          "base": cp["base"], "models": cp.get("models", ["default"]),
                          "note": "আপনার যোগ করা custom provider", "custom": True}
    return provs

def admin_add_provider(label, base, model):
    """নতুন AI provider যোগ (ভবিষ্যতের ChatGPT-6/Claude-5/যেকোনো OpenAI-compatible API)।"""
    label = str(label).strip()[:40]
    base = str(base).strip().rstrip("/")
    model = str(model).strip()[:80]
    if not label or not model or not base.startswith("http"):
        return False
    slug = re.sub(r"[^a-z0-9]", "", label.lower())[:16] or "provider"
    key = "custom_" + slug
    env = "CUSTOM_" + slug.upper() + "_KEY"
    cfg = _custom_load()
    cps = [p for p in cfg.get("custom_providers", []) if p.get("key") != key]
    cps.append({"key": key, "label": label, "base": base, "models": [model], "env": env})
    cfg["custom_providers"] = cps[:10]  # সর্বোচ্চ ১০টা custom
    _custom_save(cfg)
    log_activity("admin", f"নতুন AI provider: {label}")
    return True

def admin_remove_provider(key):
    cfg = _custom_load()
    cfg["custom_providers"] = [p for p in cfg.get("custom_providers", []) if p.get("key") != key]
    if cfg.get("provider") == key:
        cfg["provider"] = "gemini"
        cfg.pop("model", None)
    _custom_save(cfg)
    log_activity("admin", f"provider মুছে ফেলা: {key}")
    return True

def get_engine():
    cfg = _custom_load()
    provider = cfg.get("provider", "gemini")
    _all = get_all_providers()
    if provider not in _all:
        provider = "gemini"
    model = cfg.get("model") or (MODEL_NAME if provider == "gemini" else _all[provider]["models"][0])
    return {"provider": provider, "model": model}

def get_max_iters():
    cfg = _custom_load()
    try:
        n = int(cfg.get("system", {}).get("max_iterations", MAX_TOOL_ITERATIONS))
        return max(3, min(25, n))
    except (TypeError, ValueError):
        return MAX_TOOL_ITERATIONS

def admin_get():
    eng = get_engine()
    provs = []
    for key, p in get_all_providers().items():
        provs.append({"key": key, "label": p["label"], "models": p["models"],
                      "configured": bool(os.environ.get(p["env"], "")),
                      "env": p["env"], "note": p["note"],
                      "custom": p.get("custom", False),
                      "active": key == eng["provider"]})
    return {"brand": {k: BRAND[k] for k in BRAND},
            "engine": eng, "providers": provs,
            "integrations": integrations_status(),
            "system": {"max_iterations": get_max_iters()}}

def admin_save_brand(new_brand):
    if not isinstance(new_brand, dict):
        return False
    cfg = _custom_load()
    saved = cfg.get("brand", {})
    for k, v in new_brand.items():
        if k in DEFAULT_BRAND:
            if k == "chips" and isinstance(v, list):
                BRAND[k] = [str(x)[:80] for x in v][:8]
                saved[k] = BRAND[k]
            elif isinstance(v, str) and v.strip():
                BRAND[k] = v.strip()[:1200]
                saved[k] = BRAND[k]
    cfg["brand"] = saved
    _custom_save(cfg)
    log_activity("admin", "branding আপডেট")
    return True

def admin_reset_brand():
    for k, v in DEFAULT_BRAND.items():
        BRAND[k] = v
    cfg = _custom_load()
    cfg.pop("brand", None)
    _custom_save(cfg)
    log_activity("admin", "branding reset")
    return True

def admin_set_engine(provider, model):
    _all = get_all_providers()
    if provider not in _all:
        return False
    cfg = _custom_load()
    cfg["provider"] = provider
    cfg["model"] = str(model or "").strip()[:80] or _all[provider]["models"][0]
    _custom_save(cfg)
    log_activity("admin", f"AI engine: {provider} / {cfg['model']}")
    return True

def admin_set_system(max_iterations=None):
    cfg = _custom_load()
    sysd = cfg.setdefault("system", {})
    if max_iterations is not None:
        try:
            sysd["max_iterations"] = max(3, min(25, int(max_iterations)))
        except (TypeError, ValueError):
            pass
    _custom_save(cfg)
    return True

# ── OpenAI-compatible adapter (Groq/OpenAI/OpenRouter/DeepSeek) — BETA ──
def _openai_chat(messages, tools=None):
    eng = get_engine()
    p = get_all_providers()[eng["provider"]]
    key = os.environ.get(p["env"], "")
    if not key:
        raise LLMError(f"{p['label']} এর API key নেই! Admin (🔧) → API Keys-এ key বসান, "
                       f"অথবা Gemini-তে ফিরে যান।")
    # Gemini-format history → OpenAI format
    msgs = [{"role": "system", "content": build_system_prompt() + facts_as_text()}]
    call_id = 0
    for m in messages:
        role = "assistant" if m.get("role") == "model" else "user"
        texts, tcalls, tresults = [], [], []
        for part in m.get("parts", []):
            if "text" in part:
                texts.append(part["text"])
            elif "functionCall" in part:
                call_id += 1
                tcalls.append({"id": f"call_{call_id}", "type": "function",
                               "function": {"name": part["functionCall"]["name"],
                                            "arguments": json.dumps(part["functionCall"].get("args", {}), ensure_ascii=False)}})
            elif "functionResponse" in part:
                tresults.append(part["functionResponse"])
        if tcalls:
            msgs.append({"role": "assistant", "content": " ".join(texts) or None, "tool_calls": tcalls})
        elif tresults:
            base_id = call_id - len(tresults)
            for j, tr in enumerate(tresults):
                msgs.append({"role": "tool", "tool_call_id": f"call_{base_id + j + 1}",
                             "content": json.dumps(tr.get("response", {}), ensure_ascii=False)[:8000]})
        elif texts:
            msgs.append({"role": role, "content": " ".join(texts)})
    payload = {"model": eng["model"], "messages": msgs, "temperature": 0.7}
    if tools:
        payload["tools"] = [{"type": "function", "function": t} for t in tools]
    try:
        resp = requests.post(p["base"] + "/chat/completions",
                             headers={"Authorization": f"Bearer {key}"},
                             json=payload, timeout=120)
    except requests.RequestException as e:
        raise LLMError(f"Network error ({p['label']}): {e}")
    if resp.status_code != 200:
        raise LLMError(f"{p['label']} error {resp.status_code}: {resp.text[:200]}")
    track_usage(f"{eng['provider']}:{eng['model']}")
    msg = resp.json()["choices"][0]["message"]
    parts = []
    if msg.get("content"):
        parts.append({"text": msg["content"]})
    for tc in msg.get("tool_calls") or []:
        try:
            args = json.loads(tc["function"].get("arguments") or "{}")
        except json.JSONDecodeError:
            args = {}
        parts.append({"functionCall": {"name": tc["function"]["name"], "args": args}})
    return {"role": "model", "parts": parts or [{"text": ""}]}




# ╔═══════════════════════════════════════════════════════════════╗
# ║  👥 AGENT TEAM — মূল agent-এর অধীনে multiple sub-agent        ║
# ║  🌐 BUILDER — website/project ফাইল বানিয়ে দেয়                 ║
# ╚═══════════════════════════════════════════════════════════════╝
TEAM_FILE = MEMORY_DIR / "team.json"

DEFAULT_TEAM = [
    {"name": "ওয়েব ডেভেলপার", "emoji": "🌐",
     "role": "তুমি একজন expert web developer। সুন্দর, responsive, modern HTML/CSS/JS website বানাও। সম্পূর্ণ কোড এক ফাইলে দাও, inline CSS/JS সহ। বাংলা content সাপোর্ট করো।"},
    {"name": "ডিজিটাল মার্কেটার", "emoji": "📣",
     "role": "তুমি একজন expert digital marketer। Ad copy, social media strategy, SEO, funnel, campaign — বাস্তবসম্মত ও ফলদায়ক পরামর্শ দাও। বাংলাদেশের বাজার বোঝো।"},
    {"name": "কনটেন্ট রাইটার", "emoji": "✍️",
     "role": "তুমি একজন professional content writer। Blog, script, caption, email — আকর্ষণীয় ও নির্ভুল লেখা দাও।"},
    {"name": "AI Agent বিশেষজ্ঞ", "emoji": "🤖",
     "role": "তুমি AI agent বানানোর বিশেষজ্ঞ। Python-এ agent-এর architecture, tool system, prompt design — সম্পূর্ণ working code সহ বুঝিয়ে দাও।"},
]

def _team_load():
    if TEAM_FILE.exists():
        try:
            return json.loads(TEAM_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return [dict(a) for a in DEFAULT_TEAM]

def _team_save(team):
    TEAM_FILE.write_text(json.dumps(team, ensure_ascii=False, indent=1), encoding="utf-8")

def team_tool(action, name="", role="", emoji="🤝"):
    """Sub-agent team manage: add/list/remove।"""
    team = _team_load()
    action = (action or "list").lower()
    if action == "add":
        if not name.strip() or not role.strip():
            return "Error: name আর role (দায়িত্বের বর্ণনা) দুটোই দরকার।"
        team = [a for a in team if a["name"].lower() != name.strip().lower()]
        team.append({"name": name.strip()[:40], "emoji": (emoji or "🤝")[:4],
                     "role": role.strip()[:600]})
        _team_save(team)
        log_activity("team", f"নতুন sub-agent: {name}")
        return f"👥 Sub-agent যোগ হলো: {emoji} {name} (মোট {len(team)} জন)"
    if action == "remove":
        before = len(team)
        team = [a for a in team if a["name"].lower() != name.strip().lower()]
        if len(team) == before:
            return f"'{name}' team-এ নেই।"
        _team_save(team)
        return f"👥 '{name}' team থেকে বাদ। (বাকি {len(team)} জন)"
    if not team:
        return "Team খালি। 'add' দিয়ে sub-agent বানান।"
    return f"👥 আমার Agent Team ({len(team)} জন):\n" + "\n".join(
        f"{i+1}. {a['emoji']} {a['name']} — {a['role'][:60]}…" for i, a in enumerate(team))

def delegate_to_agent(agent_name, task):
    """কাজটা নির্দিষ্ট sub-agent-কে দিয়ে করায় (একই AI brain, ভিন্ন expertise)।"""
    if not task.strip():
        return "Error: কী কাজ করাতে হবে সেটা দরকার।"
    team = _team_load()
    match = [a for a in team if agent_name.strip().lower() in a["name"].lower()]
    if not match:
        return (f"'{agent_name}' নামে sub-agent নেই। আছে: "
                + ", ".join(a["name"] for a in team))
    ag = match[0]
    sub_prompt = (f"{ag['role']}\n\nতুমি '{BRAND['name']}' team-এর sub-agent \"{ag['emoji']} {ag['name']}\"। "
                  f"নিচের কাজটা সম্পূর্ণ করে দাও। শুধু কাজের ফলাফল দাও, অতিরিক্ত কথা না।\n\nকাজ: {task}")
    log_activity("team", f"{ag['name']}-কে কাজ: {task[:50]}")
    # tool ছাড়া single-shot call — sub-agent শুধু expertise দেয়
    response = llm_chat([{"role": "user", "parts": [{"text": sub_prompt}]}], tools=None)
    text = "".join(p.get("text", "") for p in response.get("parts", []))
    return f"{ag['emoji']} [{ag['name']} এর কাজ]:\n\n{text.strip()[:6000]}"

def build_website(filename, html_content):
    """সম্পূর্ণ HTML website ফাইল বানিয়ে live link দেয়।"""
    if not html_content.strip():
        return "Error: HTML content দরকার।"
    name = re.sub(r"[^A-Za-z0-9_-]", "_", Path(filename or "site").stem)[:30] or "site"
    fname = f"{name}.html"
    content = html_content
    if "<html" not in content.lower():
        content = ("<!DOCTYPE html><html lang='bn'><head><meta charset='UTF-8'>"
                   "<meta name='viewport' content='width=device-width, initial-scale=1'>"
                   f"<title>{name}</title></head><body>{content}</body></html>")
    (SANDBOX / fname).write_text(content, encoding="utf-8")
    log_activity("builder", f"website: {fname} ({len(content)} chars)")
    return (f"🌐 Website তৈরি!\n📁 নতুন ফাইল তৈরি হয়েছে: {fname} (sandbox ফোল্ডারে)\n"
            f"🔗 দেখতে: /sandbox/{fname} — chat-এর লিংক বাটনে ক্লিক করুন। "
            f"Download করে যেকোনো hosting-এ (Netlify/GitHub Pages) ফ্রি তোলা যাবে।")

TOOLS.update({
    "team_tool": {"func": team_tool, "declaration": _decl(
        "team_tool",
        "Sub-agent team manage: action='list' (কারা আছে), 'add' (নতুন বিশেষজ্ঞ sub-agent বানানো), 'remove'। User নতুন agent বানাতে/দেখতে চাইলে এটা।",
        {"action": {"type": "string", "description": "add / list / remove"},
         "name": {"type": "string", "description": "sub-agent-এর নাম"},
         "role": {"type": "string", "description": "দায়িত্ব/expertise-এর বর্ণনা"},
         "emoji": {"type": "string", "description": "একটা emoji"}},
        ["action"])},
    "delegate_to_agent": {"func": delegate_to_agent, "declaration": _decl(
        "delegate_to_agent",
        "কাজ নির্দিষ্ট sub-agent-কে দিয়ে করাও (যেমন: ওয়েব ডেভেলপার, ডিজিটাল মার্কেটার, কনটেন্ট রাইটার, AI Agent বিশেষজ্ঞ)। "
        "বিশেষজ্ঞ-মানের কাজ দরকার হলে delegate করো, ফলাফল user-কে দাও।",
        {"agent_name": {"type": "string", "description": "কোন sub-agent (নামের অংশ হলেও চলবে)"},
         "task": {"type": "string", "description": "সম্পূর্ণ কাজের বর্ণনা"}},
        ["agent_name", "task"])},
    "build_website": {"func": build_website, "declaration": _decl(
        "build_website",
        "সম্পূর্ণ HTML website/landing page ফাইল বানিয়ে live link দেয়। User website চাইলে: "
        "সুন্দর responsive HTML (inline CSS/JS সহ) লিখে এটা দিয়ে save করো। বড় site হলে আগে delegate_to_agent দিয়ে ওয়েব ডেভেলপারের থেকে কোড নাও।",
        {"filename": {"type": "string", "description": "ফাইলের নাম (English, যেমন 'shop_landing')"},
         "html_content": {"type": "string", "description": "সম্পূর্ণ HTML কোড"}},
        ["filename", "html_content"])},
})
TOOL_PERMISSION.update({
    "team_tool": "memory", "delegate_to_agent": "web", "build_website": "code",
})




# ╔═══════════════════════════════════════════════════════════════╗
# ║  🪄 CREATOR TOOLS — image generation, file manager, shell,    ║
# ║     multi-file project + ZIP (Arena-agent-এর মতো ক্ষমতা)      ║
# ╚═══════════════════════════════════════════════════════════════╝
IMAGE_MODELS = ["gemini-3.1-flash-image", "gemini-2.5-flash-image"]

def generate_image(prompt, filename="generated"):
    """AI দিয়ে ছবি বানায় (Gemini image model, ফ্রি কোটায়)।"""
    if not prompt.strip():
        return "Error: ছবির বর্ণনা (prompt) দরকার।"
    if not GEMINI_API_KEY:
        return "Error: GEMINI_API_KEY নেই।"
    name = re.sub(r"[^A-Za-z0-9_-]", "_", Path(filename or "generated").stem)[:30] or "img"
    last_err = ""
    for model in IMAGE_MODELS:
        try:
            resp = requests.post(
                API_URL.format(model=model),
                params={"key": GEMINI_API_KEY},
                json={"contents": [{"role": "user", "parts": [{"text": prompt}]}],
                      "generationConfig": {"responseModalities": ["TEXT", "IMAGE"]}},
                timeout=120)
            if resp.status_code == 429:
                last_err = "কোটা শেষ"
                continue
            if resp.status_code != 200:
                last_err = f"HTTP {resp.status_code}"
                continue
            for part in resp.json()["candidates"][0]["content"]["parts"]:
                if "inlineData" in part:
                    data = base64.b64decode(part["inlineData"]["data"])
                    fname = f"{name}.png"
                    (SANDBOX / fname).write_bytes(data)
                    track_usage(model)
                    log_activity("image", f"ছবি তৈরি: {fname}")
                    return f"🎨 ছবি তৈরি!\n📁 নতুন ফাইল তৈরি হয়েছে: {fname} (sandbox ফোল্ডারে)"
            last_err = "ছবি আসেনি"
        except Exception as e:
            last_err = str(e)[:80]
    return f"ছবি বানাতে পারলাম না ({last_err})। ফ্রি image-কোটা সীমিত — পরে আবার চেষ্টা করুন।"

def file_manager(action, filename="", content="", find_text="", replace_text=""):
    """ফাইল ব্যবস্থাপনা: list/read/write/append/edit/delete।"""
    action = (action or "list").lower()
    if action == "list":
        out = []
        for label, d in (("sandbox", SANDBOX), ("notes", WORKSPACE), ("documents", DOCS_DIR)):
            files = [f for f in d.iterdir() if f.is_file() and f.name != "_run.py"]
            if files:
                out.append(f"📂 {label}/: " + ", ".join(
                    f"{f.name} ({max(1, f.stat().st_size // 1024)}KB)" for f in sorted(files)[:20]))
        return "\n".join(out) or "কোনো ফাইল নেই।"
    safe = Path(filename).name
    if not safe:
        return "Error: filename দরকার।"
    target = None
    for d in (SANDBOX, WORKSPACE, DOCS_DIR):
        if (d / safe).exists():
            target = d / safe
            break
    if action == "read":
        if not target:
            return f"'{safe}' পাওয়া যায়নি। file_manager list দিয়ে দেখুন।"
        if target.suffix.lower() in (".png", ".jpg", ".jpeg", ".zip", ".xlsx"):
            return f"'{safe}' binary ফাইল — text হিসেবে পড়া যায় না।"
        text = target.read_text(encoding="utf-8", errors="ignore")
        return f"📄 {safe} ({len(text)} chars):\n{text[:4000]}" + ("…" if len(text) > 4000 else "")
    if action in ("write", "append"):
        if not content:
            return "Error: content দরকার।"
        p = target or (SANDBOX / safe)
        if action == "append" and p.exists():
            content = p.read_text(encoding="utf-8", errors="ignore") + "\n" + content
        p.write_text(content, encoding="utf-8")
        log_activity("file", f"{action}: {safe}")
        link = f"\n🔗 sandbox/{safe}" if safe.endswith(".html") else ""
        return f"💾 '{safe}' সংরক্ষণ হলো ({len(content)} chars)।{link}"
    if action == "edit":
        if not target:
            return f"'{safe}' পাওয়া যায়নি।"
        if not find_text:
            return "Error: find_text দরকার।"
        text = target.read_text(encoding="utf-8", errors="ignore")
        if find_text not in text:
            return f"'{find_text[:40]}' লেখাটা '{safe}'-এ পাওয়া যায়নি।"
        n = text.count(find_text)
        target.write_text(text.replace(find_text, replace_text or ""), encoding="utf-8")
        log_activity("file", f"edit: {safe} ({n} জায়গায়)")
        return f"✏️ '{safe}'-এ {n} জায়গায় বদলানো হলো।"
    if action == "delete":
        if not target:
            return f"'{safe}' পাওয়া যায়নি।"
        target.unlink()
        log_activity("file", f"delete: {safe}")
        return f"🗑️ '{safe}' মুছে ফেলা হলো।"
    return "Error: action হবে list/read/write/append/edit/delete।"

def run_shell(command):
    """নিরাপদ shell command (sandbox-এ, ১৫ সেকেন্ড, whitelist)।"""
    if not command.strip():
        return "Error: command দরকার।"
    first = command.strip().split()[0]
    ALLOWED = {"ls", "cat", "head", "tail", "wc", "grep", "find", "du", "date",
               "echo", "sort", "uniq", "diff", "python3", "pip", "zip", "unzip", "curl"}
    if first not in ALLOWED:
        return (f"⛔ '{first}' অনুমোদিত না। চলবে: {', '.join(sorted(ALLOWED))}। "
                f"জটিল কাজে run_python_code ব্যবহার করো।")
    banned = ["rm ", "sudo", ">", "&", ";", "|", "`", "$("]
    if any(b in command for b in banned):
        return "⛔ pipe/redirect/chain করা যাবে না — একটাই সরল command দিন।"
    try:
        result = subprocess.run(command.split(), capture_output=True, text=True,
                                timeout=15, cwd=str(SANDBOX))
    except subprocess.TimeoutExpired:
        return "⏱️ Timeout (১৫ সেকেন্ড)।"
    except FileNotFoundError:
        return f"'{first}' পাওয়া যায়নি।"
    out = (result.stdout or "").strip()[:3000]
    err = (result.stderr or "").strip()[:800]
    log_activity("shell", command[:60])
    return (f"$ {command}\n{out}" + (f"\n[stderr] {err}" if err else "")) or "কোনো output নেই।"

def package_project(zip_name, file_list=""):
    """Sandbox-এর ফাইলগুলো ZIP করে download link দেয় (multi-file project)।"""
    import zipfile
    name = re.sub(r"[^A-Za-z0-9_-]", "_", Path(zip_name or "project").stem)[:30] or "project"
    wanted = [f.strip() for f in (file_list or "").split(",") if f.strip()]
    files = []
    for f in SANDBOX.iterdir():
        if f.is_file() and f.name != "_run.py" and not f.name.endswith(".zip"):
            if not wanted or any(w in f.name for w in wanted):
                files.append(f)
    if not files:
        return "ZIP করার মতো ফাইল নেই।"
    zpath = SANDBOX / f"{name}.zip"
    with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as z:
        for f in files[:40]:
            z.write(f, f.name)
    log_activity("project", f"ZIP: {name}.zip ({len(files)} ফাইল)")
    return (f"📦 Project ZIP তৈরি: {name}.zip ({len(files)}টা ফাইল)\n"
            f"🔗 Download: sandbox/{name}.zip")

TOOLS.update({
    "generate_image": {"func": generate_image, "declaration": _decl(
        "generate_image",
        "AI দিয়ে ছবি/logo/illustration বানায় (Gemini image)। User ছবি বানাতে চাইলে এটা। প্রম্পট English-এ বিস্তারিত লিখলে ভালো ফল।",
        {"prompt": {"type": "string", "description": "ছবির বিস্তারিত বর্ণনা (English ভালো)"},
         "filename": {"type": "string", "description": "ফাইলের নাম (English)"}},
        ["prompt"])},
    "file_manager": {"func": file_manager, "declaration": _decl(
        "file_manager",
        "ফাইল ব্যবস্থাপনা: action='list' (সব ফাইল), 'read' (পড়া), 'write' (নতুন/overwrite), 'append', 'edit' (find_text→replace_text), 'delete'। Website/code ফাইল এডিটেও এটা।",
        {"action": {"type": "string", "description": "list/read/write/append/edit/delete"},
         "filename": {"type": "string", "description": "ফাইলের নাম"},
         "content": {"type": "string", "description": "write/append-এর লেখা"},
         "find_text": {"type": "string", "description": "edit: যা খুঁজবে"},
         "replace_text": {"type": "string", "description": "edit: যা বসাবে"}},
        ["action"])},
    "run_shell": {"func": run_shell, "declaration": _decl(
        "run_shell",
        "নিরাপদ shell command চালায় (ls, cat, grep, python3, pip, zip, curl ইত্যাদি — sandbox-এ)। ফাইল দেখা/খোঁজায় কাজে লাগে।",
        {"command": {"type": "string", "description": "একটা সরল command"}},
        ["command"])},
    "package_project": {"func": package_project, "declaration": _decl(
        "package_project",
        "Sandbox-এর ফাইলগুলো এক ZIP-এ বেঁধে download link দেয়। Multi-file website/project বানানোর শেষে এটা দিয়ে প্যাকেজ করো।",
        {"zip_name": {"type": "string", "description": "ZIP-এর নাম (English)"},
         "file_list": {"type": "string", "description": "কমা দিয়ে নির্দিষ্ট ফাইল (খালি=সব)"}},
        ["zip_name"])},
})
TOOL_PERMISSION.update({
    "generate_image": "web", "file_manager": "files",
    "run_shell": "code", "package_project": "files",
})
SENSITIVE_TOOLS.add("run_shell")
SENSITIVE_TOOLS.add("file_manager")




# ╔═══════════════════════════════════════════════════════════════╗
# ║  ⏰ REMINDER SYSTEM — সময়মতো notification দিয়ে মনে করায়       ║
# ╚═══════════════════════════════════════════════════════════════╝
TZ_OFFSET_HOURS = float(os.environ.get("TZ_OFFSET_HOURS", "6"))  # বাংলাদেশ = UTC+6

def now_local():
    """Server যেখানেই থাকুক (Render=UTC), সবসময় বাংলাদেশ সময়।"""
    return datetime.datetime.now(datetime.timezone.utc).replace(tzinfo=None) + datetime.timedelta(hours=TZ_OFFSET_HOURS)

REMINDERS_FILE = MEMORY_DIR / "reminders.json"

def _rem_load():
    if REMINDERS_FILE.exists():
        try:
            return json.loads(REMINDERS_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return []
    return []

def _rem_save(items):
    REMINDERS_FILE.write_text(json.dumps(items, ensure_ascii=False, indent=1), encoding="utf-8")

def reminder_tool(action, text="", when="", number=None):
    """Reminder: add / list / delete।"""
    items = _rem_load()
    action = (action or "list").lower()
    if action == "add":
        if not text.strip():
            return "Error: কী মনে করাতে হবে সেটা দরকার।"
        try:
            dt = datetime.datetime.strptime(when.strip(), "%Y-%m-%d %H:%M")
        except (ValueError, AttributeError):
            return ("Error: when অবশ্যই 'YYYY-MM-DD HH:MM' format-এ দিতে হবে (24-ঘণ্টা)। "
                    f"এখন সময়: {now_local().strftime('%Y-%m-%d %H:%M')}")
        if dt <= now_local():
            return f"Error: সময়টা অতীতের! এখন {now_local().strftime('%Y-%m-%d %H:%M')} — ভবিষ্যতের সময় দাও।"
        items.append({"text": text.strip()[:200], "when": dt.strftime("%Y-%m-%d %H:%M"),
                      "notified": False, "created": now_local().strftime("%Y-%m-%d %H:%M")})
        items.sort(key=lambda r: r["when"])
        _rem_save(items[:50])
        log_activity("reminder", f"set: {when} — {text[:40]}")
        return (f"⏰ Reminder set হলো!\n📌 {text.strip()}\n🕐 {dt.strftime('%d/%m/%Y %H:%M')} "
                f"(বাংলাদেশ সময়)\nসময় হলেই notification দিয়ে মনে করিয়ে দেবো। "
                f"(⚠️ app-টা browser-এ খোলা থাকতে হবে)")
    if action == "delete":
        try:
            i = int(number) - 1
        except (TypeError, ValueError):
            return "Error: কত নম্বর reminder মুছবো সেটা দরকার।"
        active = [r for r in items if not r.get("notified")]
        if not (0 <= i < len(active)):
            return f"{number} নম্বর reminder নেই। মোট active: {len(active)}টা।"
        target = active[i]
        items.remove(target)
        _rem_save(items)
        return f"🗑️ মুছে ফেলা হলো: {target['text'][:50]}"
    # list
    active = [r for r in items if not r.get("notified")]
    if not active:
        return "কোনো active reminder নেই। 'add' দিয়ে সেট করুন।"
    return f"⏰ Active reminders ({len(active)}টা):\n" + "\n".join(
        f"{i+1}. 🕐 {r['when']} — {r['text']}" for i, r in enumerate(active))

def due_reminders():
    """যেসব reminder-এর সময় হয়ে গেছে — ফেরত দেয় ও notified মার্ক করে।"""
    items = _rem_load()
    now_str = now_local().strftime("%Y-%m-%d %H:%M")
    due = [r for r in items if not r.get("notified") and r["when"] <= now_str]
    if due:
        for r in due:
            r["notified"] = True
            log_activity("reminder", f"🔔 fired: {r['text'][:40]}")
        _rem_save(items)
    return due

TOOLS.update({
    "reminder_tool": {"func": reminder_tool, "declaration": _decl(
        "reminder_tool",
        "Reminder set/list/delete করে। User 'মনে করিয়ে দিও' বললে: action='add', text=কী মনে করাতে হবে, "
        "when='YYYY-MM-DD HH:MM' (24-ঘণ্টা, বাংলাদেশ সময়)। System prompt-এ এখনকার তারিখ-সময় দেওয়া আছে — "
        "'কালকে ১টা' মানে আগামীকালের 13:00 হিসাব করে দাও।",
        {"action": {"type": "string", "description": "add / list / delete"},
         "text": {"type": "string", "description": "কী মনে করাতে হবে"},
         "when": {"type": "string", "description": "YYYY-MM-DD HH:MM"},
         "number": {"type": "integer", "description": "delete-এর জন্য কত নম্বর"}},
        ["action"])},
})
TOOL_PERMISSION["reminder_tool"] = "memory"




# ╔═══════════════════════════════════════════════════════════════╗
# ║  🧬 SELF-BUILDING — agent নিজেই নতুন tool/feature বানায়        ║
# ╚═══════════════════════════════════════════════════════════════╝
PLUGINS_DIR = Path(__file__).parent / "plugins"
PLUGINS_DIR.mkdir(exist_ok=True)

def _plugin_env():
    """Plugin-এর ভেতরে যা যা ব্যবহার করা যাবে।"""
    return {
        "requests": requests, "json": json, "re": re, "math": math,
        "datetime": datetime, "Path": Path, "os": os, "base64": base64,
        "now_local": now_local, "SANDBOX": SANDBOX, "WORKSPACE": WORKSPACE,
        "log_activity": log_activity, "web_search": web_search,
        "fetch_webpage": fetch_webpage, "_biz_load": _biz_load, "_biz_save": _biz_save,
    }

def _validate_plugin_code(code):
    """Plugin code নিরাপদ ও সঠিক কিনা যাচাই। ফেরত: (fn_name, error)।"""
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return None, f"Python syntax ভুল: {e}"
    fns = [n for n in tree.body if isinstance(n, ast.FunctionDef)]
    if len(fns) != 1 or len(tree.body) != 1:
        return None, "ঠিক ১টা function definition দাও (আর কিছু না — import-ও না, সব আগে থেকে দেওয়া আছে)।"
    banned = {"exec", "eval", "compile", "__import__", "open"}
    for node in ast.walk(tree):
        if isinstance(node, ast.Name) and node.id in banned:
            return None, f"'{node.id}' plugin-এ ব্যবহার করা যাবে না। ফাইল লিখতে SANDBOX/WORKSPACE Path ব্যবহার করো (write_text/read_text)।"
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            return None, "import করা যাবে না — requests/json/re/math/datetime/Path আগে থেকেই আছে।"
    name = fns[0].name
    if not re.match(r"^[a-z][a-z0-9_]{2,30}$", name):
        return None, "Function-এর নাম ছোট হাতের english + underscore হতে হবে।"
    if name in TOOLS and not (PLUGINS_DIR / f"{name}.py").exists():
        return None, f"'{name}' নামে built-in tool আগে থেকেই আছে — অন্য নাম দাও।"
    return name, None

def _register_plugin(fn_name, code, description, params=None):
    """Code চালিয়ে function বের করে TOOLS-এ register করে।"""
    env = _plugin_env()
    exec(compile(code, f"<plugin:{fn_name}>", "exec"), env)  # validated code only
    fn = env[fn_name]
    props = {}
    required = []
    if isinstance(params, dict):
        for pname, pdesc in list(params.items())[:6]:
            props[str(pname)] = {"type": "string", "description": str(pdesc)[:120]}
            required.append(str(pname))
    def safe_wrapper(**kwargs):
        try:
            return str(fn(**kwargs))[:4000]
        except Exception as e:
            return f"Plugin error ({fn_name}): {e}"
    TOOLS[fn_name] = {"func": safe_wrapper, "declaration": _decl(
        fn_name, (description or fn_name)[:250], props, required or None)}
    TOOL_PERMISSION[fn_name] = "code"

def create_plugin(name_description, code, params_json=""):
    """নতুন tool/feature বানিয়ে agent-এর মধ্যে স্থায়ীভাবে যোগ করে।"""
    if not code.strip():
        return "Error: Python function code দরকার।"
    params = None
    if params_json and params_json.strip():
        try:
            params = json.loads(params_json)
        except json.JSONDecodeError:
            return 'Error: params_json সঠিক JSON না। format: {"param_name": "কী কাজে লাগে"}'
    fn_name, err = _validate_plugin_code(code)
    if err:
        return f"⛔ Plugin বানানো গেল না: {err}"
    try:
        _register_plugin(fn_name, code, name_description, params)
    except Exception as e:
        return f"⛔ Plugin চালু করতে সমস্যা: {e}"
    meta = {"description": (name_description or "")[:250], "params": params or {},
            "created": now_local().strftime("%Y-%m-%d %H:%M")}
    (PLUGINS_DIR / f"{fn_name}.py").write_text(code, encoding="utf-8")
    (PLUGINS_DIR / f"{fn_name}.json").write_text(json.dumps(meta, ensure_ascii=False), encoding="utf-8")
    log_activity("plugin", f"নতুন feature: {fn_name}")
    return (f"🧬 নতুন feature যোগ হয়ে গেছে!\n🔧 Tool: {fn_name}\n📝 {meta['description'][:80]}\n"
            f"এখন থেকে এই tool ব্যবহার করা যাবে (restart-এও থাকবে)। "
            f"Test করতে: user-কে বলো tool-টা চালিয়ে দেখাতে।")

def list_plugins():
    files = sorted(PLUGINS_DIR.glob("*.py"))
    if not files:
        return ("কোনো custom plugin নেই। user নতুন feature চাইলে create_plugin দিয়ে বানাও, "
                "বা জটিল হলে ধাপে ধাপে গাইড করো।")
    out = []
    for f in files:
        meta = {}
        mf = PLUGINS_DIR / (f.stem + ".json")
        if mf.exists():
            try:
                meta = json.loads(mf.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                pass
        out.append(f"• {f.stem} — {meta.get('description','')[:60]} ({meta.get('created','')})")
    return f"🧬 Custom plugins ({len(files)}টা):\n" + "\n".join(out)

def delete_plugin(name):
    safe = re.sub(r"[^a-z0-9_]", "", str(name).lower())
    pf = PLUGINS_DIR / f"{safe}.py"
    if not pf.exists():
        return f"'{name}' নামে plugin নেই। list_plugins দিয়ে দেখো।"
    pf.unlink()
    (PLUGINS_DIR / f"{safe}.json").unlink(missing_ok=True)
    TOOLS.pop(safe, None)
    log_activity("plugin", f"plugin মুছে ফেলা: {safe}")
    return f"🗑️ Plugin '{safe}' মুছে ফেলা হলো।"

def load_plugins():
    """Startup-এ save করা সব plugin আবার চালু করে।"""
    loaded = 0
    for f in sorted(PLUGINS_DIR.glob("*.py")):
        try:
            code = f.read_text(encoding="utf-8")
            fn_name, err = _validate_plugin_code(code)
            if err or fn_name != f.stem:
                continue
            meta = {}
            mf = PLUGINS_DIR / (f.stem + ".json")
            if mf.exists():
                try:
                    meta = json.loads(mf.read_text(encoding="utf-8"))
                except json.JSONDecodeError:
                    pass
            _register_plugin(fn_name, code, meta.get("description", fn_name), meta.get("params"))
            loaded += 1
        except Exception:
            continue
    return loaded

TOOLS.update({
    "create_plugin": {"func": create_plugin, "declaration": _decl(
        "create_plugin",
        "নিজের মধ্যে নতুন tool/feature স্থায়ীভাবে যোগ করে! User নতুন feature চাইলে: "
        "১টা Python function লেখো (import ছাড়া — requests/json/re/math/datetime/Path/web_search আগে থেকেই আছে), "
        "সব parameter string ও default value সহ। code-এ শুধু function-টাই থাকবে। "
        "params_json-এ parameter-দের বর্ণনা দাও।",
        {"name_description": {"type": "string", "description": "feature-টা কী করে (বাংলায়)"},
         "code": {"type": "string", "description": "সম্পূর্ণ Python function (def দিয়ে শুরু)"},
         "params_json": {"type": "string", "description": 'JSON: {"param": "বর্ণনা"} (ঐচ্ছিক)'}},
        ["name_description", "code"])},
    "list_plugins": {"func": list_plugins, "declaration": _decl(
        "list_plugins", "নিজের বানানো custom plugin/feature-দের তালিকা দেখায়।")},
    "delete_plugin": {"func": delete_plugin, "declaration": _decl(
        "delete_plugin", "একটা custom plugin মুছে ফেলে।",
        {"name": {"type": "string", "description": "plugin-এর নাম"}}, ["name"])},
})
TOOL_PERMISSION.update({"create_plugin": "code", "delete_plugin": "code"})
SENSITIVE_TOOLS.add("create_plugin")
SENSITIVE_TOOLS.add("delete_plugin")
_PLUGINS_LOADED = load_plugins()




# ╔═══════════════════════════════════════════════════════════════╗
# ║  🦸 SUPER TOOLS — ফ্রিগুলো এখনই চলে; paid গুলো key বসালেই     ║
# ║  চলবে, key না থাকলে agent নিজেই কেনার গাইড দেয়                ║
# ╚═══════════════════════════════════════════════════════════════╝

def _need_key(env, service, how):
    return (f"🔑 এই কাজের জন্য {service}-এর key দরকার (এখনো বসানো হয়নি)।\n"
            f"কীভাবে পাবেন: {how}\n"
            f"Key পেলে Admin (🔧) → API Keys-এ '{env}' ঘরে বসান — তারপর আমি এই কাজ নিজেই করবো!")

# ── 🌦️ WEATHER (সম্পূর্ণ ফ্রি — key লাগে না) ──
def get_weather(city):
    if not str(city).strip():
        return "Error: শহরের নাম দরকার।"
    try:
        g = requests.get("https://geocoding-api.open-meteo.com/v1/search",
                         params={"name": str(city).strip(), "count": 1}, timeout=15).json()
        if not g.get("results"):
            return f"'{city}' পাওয়া যায়নি — English বানানে দিন (যেমন Chattogram)।"
        loc = g["results"][0]
        w = requests.get("https://api.open-meteo.com/v1/forecast",
            params={"latitude": loc["latitude"], "longitude": loc["longitude"],
                    "current": "temperature_2m,relative_humidity_2m,weather_code,wind_speed_10m",
                    "daily": "temperature_2m_max,temperature_2m_min,precipitation_probability_max",
                    "forecast_days": 3, "timezone": "auto"}, timeout=15).json()
        cur = w.get("current", {})
        daily = w.get("daily", {})
        codes = {0:"পরিষ্কার ☀️",1:"মোটামুটি পরিষ্কার 🌤️",2:"আংশিক মেঘলা ⛅",3:"মেঘলা ☁️",
                 45:"কুয়াশা 🌫️",51:"গুঁড়ি বৃষ্টি 🌦️",61:"হালকা বৃষ্টি 🌧️",63:"বৃষ্টি 🌧️",
                 65:"ভারী বৃষ্টি ⛈️",80:"বৃষ্টির ঝাপটা 🌧️",95:"বজ্রঝড় ⛈️"}
        out = [f"🌦️ {loc['name']}, {loc.get('country','')}:",
               f"🌡️ {cur.get('temperature_2m','?')}°C | 💧 {cur.get('relative_humidity_2m','?')}% | 💨 {cur.get('wind_speed_10m','?')} km/h",
               f"☁️ {codes.get(cur.get('weather_code'), 'অজানা')}", "আগামী ৩ দিন:"]
        for i, d in enumerate(daily.get("time", [])[:3]):
            out.append(f"• {d}: {daily['temperature_2m_min'][i]}–{daily['temperature_2m_max'][i]}°C, বৃষ্টি {daily['precipitation_probability_max'][i]}%")
        log_activity("tool", f"weather: {city}")
        return "\n".join(out)
    except Exception as e:
        return f"আবহাওয়া আনতে সমস্যা: {e}"

# ── 💱 CURRENCY (সম্পূর্ণ ফ্রি) ──
def currency_convert(amount, from_currency="USD", to_currency="BDT"):
    try:
        amt = float(str(amount).replace(",", ""))
    except (TypeError, ValueError):
        return "Error: amount সংখ্যা হতে হবে।"
    f, t = str(from_currency).upper().strip()[:3], str(to_currency).upper().strip()[:3]
    try:
        r = requests.get(f"https://open.er-api.com/v6/latest/{f}", timeout=15).json()
        if r.get("result") != "success":
            return f"'{f}' currency পাওয়া যায়নি।"
        rate = r["rates"].get(t)
        if not rate:
            return f"'{t}' currency পাওয়া যায়নি।"
        log_activity("tool", f"currency {f}->{t}")
        return f"💱 1 {f} = {rate:,.2f} {t}\n➡️ {amt:,.2f} {f} = {amt*rate:,.2f} {t}"
    except Exception as e:
        return f"রেট আনতে সমস্যা: {e}"

# ── 📨 TELEGRAM (ফ্রি token — BotFather) ──
def send_telegram(chat_id, message="", file_name=""):
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if not token:
        return _need_key("TELEGRAM_BOT_TOKEN", "Telegram Bot",
                         "Telegram-এ @BotFather → /newbot → ফ্রি token (টাকা লাগে না)")
    base = f"https://api.telegram.org/bot{token}"
    try:
        if str(file_name).strip():
            p = None
            for d in (SANDBOX, WORKSPACE):
                cand = d / Path(file_name).name
                if cand.exists():
                    p = cand
                    break
            if not p:
                return f"'{file_name}' ফাইল পাওয়া যায়নি।"
            with p.open("rb") as fh:
                r = requests.post(f"{base}/sendDocument", data={"chat_id": str(chat_id)},
                                  files={"document": fh}, timeout=45)
        else:
            if not str(message).strip():
                return "Error: message বা file_name অন্তত একটা দরকার।"
            r = requests.post(f"{base}/sendMessage",
                              json={"chat_id": str(chat_id), "text": str(message)[:4000]}, timeout=30)
        if r.status_code == 200:
            log_activity("send", f"telegram → {chat_id}")
            return "✅ Telegram-এ পাঠানো হয়েছে!"
        return f"Telegram error: {r.text[:120]}"
    except Exception as e:
        return f"পাঠাতে সমস্যা: {e}"

# ── 📧 EMAIL (Resend — দিনে ১০০ ফ্রি) ──
def send_email(to, subject, body):
    key = os.environ.get("EMAIL_API_KEY", "")
    if not key:
        return _need_key("EMAIL_API_KEY", "Resend Email",
                         "resend.com-এ ফ্রি sign up → API Keys → Create (দিনে ১০০ email ফ্রি)")
    if not (str(to).strip() and "@" in str(to)):
        return "Error: সঠিক email ঠিকানা দরকার।"
    try:
        r = requests.post("https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {key}"},
            json={"from": os.environ.get("EMAIL_FROM", "onboarding@resend.dev"),
                  "to": [str(to).strip()], "subject": str(subject)[:200] or "(no subject)",
                  "text": str(body)[:10000]}, timeout=30)
        if r.status_code in (200, 201):
            log_activity("send", f"email → {to}")
            return f"✅ Email পাঠানো হয়েছে: {to}"
        return f"Email error {r.status_code}: {r.text[:120]}"
    except Exception as e:
        return f"Email সমস্যা: {e}"

# ── 🚀 DEPLOY WEBSITE (Netlify — ফ্রি token) ──
def deploy_website(file_name):
    token = os.environ.get("NETLIFY_TOKEN", "")
    if not token:
        return _need_key("NETLIFY_TOKEN", "Netlify (ফ্রি hosting)",
                         "netlify.com-এ ফ্রি sign up → User settings → Applications → New access token")
    p = None
    for d in (SANDBOX, WORKSPACE):
        cand = d / Path(file_name).name
        if cand.exists():
            p = cand
            break
    if not p or not p.name.endswith(".html"):
        return f"'{file_name}' HTML ফাইল পাওয়া যায়নি। আগে build_website দিয়ে বানাও।"
    try:
        import zipfile, io
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            z.write(p, "index.html")
        r = requests.post("https://api.netlify.com/api/v1/sites",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/zip"},
            data=buf.getvalue(), timeout=90)
        if r.status_code in (200, 201):
            j = r.json()
            url = j.get("ssl_url") or j.get("url", "")
            log_activity("deploy", f"netlify: {url}")
            return f"🚀 Website LIVE হয়ে গেছে!\n🔗 {url}\nযে কেউ এই লিংক খুলতে পারবে!"
        return f"Netlify error {r.status_code}: {r.text[:120]}"
    except Exception as e:
        return f"Deploy সমস্যা: {e}"

# ── 📱 SMS (Twilio — paid) ──
def send_sms(phone, message):
    sid = os.environ.get("TWILIO_SID", "")
    tok = os.environ.get("TWILIO_TOKEN", "")
    from_no = os.environ.get("TWILIO_FROM", "")
    if not (sid and tok and from_no):
        return _need_key("TWILIO_SID + TWILIO_TOKEN + TWILIO_FROM", "Twilio SMS (paid)",
                         "twilio.com-এ account → SID, Auth Token, Phone number কিনুন (৩টা ঘরেই বসাতে হবে)")
    try:
        r = requests.post(f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json",
            auth=(sid, tok),
            data={"To": str(phone), "From": from_no, "Body": str(message)[:1500]}, timeout=30)
        if r.status_code in (200, 201):
            log_activity("send", f"sms → {phone}")
            return f"✅ SMS পাঠানো হয়েছে: {phone}"
        return f"SMS error: {r.text[:120]}"
    except Exception as e:
        return f"SMS সমস্যা: {e}"

TOOLS.update({
    "get_weather": {"func": get_weather, "declaration": _decl(
        "get_weather", "যেকোনো শহরের লাইভ আবহাওয়া + ৩ দিনের পূর্বাভাস (ফ্রি, নির্ভুল)। আবহাওয়ার প্রশ্নে web_search-এর বদলে এটাই ব্যবহার করো।",
        {"city": {"type": "string", "description": "শহরের নাম English-এ (যেমন Chattogram, Dhaka)"}}, ["city"])},
    "currency_convert": {"func": currency_convert, "declaration": _decl(
        "currency_convert", "লাইভ currency রেট + রূপান্তর (ফ্রি)। ডলার/টাকা/রিয়াল প্রশ্নে এটাই ব্যবহার করো।",
        {"amount": {"type": "number", "description": "কত টাকা/ডলার"},
         "from_currency": {"type": "string", "description": "কোন currency থেকে (USD/BDT/SAR/EUR...)"},
         "to_currency": {"type": "string", "description": "কোনটায়"}}, ["amount"])},
    "send_telegram": {"func": send_telegram, "declaration": _decl(
        "send_telegram", "Telegram-এ message বা ফাইল পাঠায়। Key না থাকলে user-কে ফ্রি token নেওয়ার গাইড দেয়।",
        {"chat_id": {"type": "string", "description": "প্রাপকের chat ID"},
         "message": {"type": "string", "description": "message"},
         "file_name": {"type": "string", "description": "sandbox-এর ফাইল (ঐচ্ছিক)"}}, ["chat_id"])},
    "send_email": {"func": send_email, "declaration": _decl(
        "send_email", "আসল email পাঠায় (Resend, দিনে ১০০ ফ্রি)। Key না থাকলে গাইড দেয়।",
        {"to": {"type": "string", "description": "প্রাপকের email"},
         "subject": {"type": "string", "description": "বিষয়"},
         "body": {"type": "string", "description": "email-এর লেখা"}}, ["to", "subject", "body"])},
    "deploy_website": {"func": deploy_website, "declaration": _decl(
        "deploy_website", "বানানো HTML website-কে internet-এ LIVE করে (Netlify ফ্রি)। build_website-এর পরে user চাইলে এটা।",
        {"file_name": {"type": "string", "description": "sandbox-এর .html ফাইলের নাম"}}, ["file_name"])},
    "send_sms": {"func": send_sms, "declaration": _decl(
        "send_sms", "ফোনে SMS পাঠায় (Twilio, paid)। Key না থাকলে কেনার গাইড দেয়।",
        {"phone": {"type": "string", "description": "নম্বর (+8801... format)"},
         "message": {"type": "string", "description": "SMS লেখা"}}, ["phone", "message"])},
})
TOOL_PERMISSION.update({
    "get_weather": "web", "currency_convert": "web",
    "send_telegram": "business", "send_email": "business",
    "deploy_website": "code", "send_sms": "business",
})
SENSITIVE_TOOLS.update({"send_telegram", "send_email", "send_sms", "deploy_website"})

# INTEGRATIONS তালিকা আপডেট (Admin key vault-এ ঘর আসবে)
INTEGRATIONS.update({
    "email_resend": {"label": "📧 Resend Email (১০০/দিন ফ্রি)", "env": "EMAIL_API_KEY",
                     "note": "resend.com → ফ্রি API key"},
    "netlify": {"label": "🚀 Netlify Deploy (ফ্রি)", "env": "NETLIFY_TOKEN",
                "note": "netlify.com → access token"},
    "twilio_sid": {"label": "📱 Twilio SID (paid SMS)", "env": "TWILIO_SID", "note": "twilio.com"},
    "twilio_token": {"label": "📱 Twilio Token", "env": "TWILIO_TOKEN", "note": "twilio.com"},
    "twilio_from": {"label": "📱 Twilio নম্বর", "env": "TWILIO_FROM", "note": "+1... নম্বর"},
})




# ╔═══════════════════════════════════════════════════════════════╗
# ║  🛠️ UTILITY TOOLS — QR code, password, converter, export      ║
# ╚═══════════════════════════════════════════════════════════════╝

def make_qr_code(text, filename="qr"):
    """QR code বানায় (লিংক/bKash নম্বর/যেকোনো লেখা) — ছবি হিসেবে।"""
    if not str(text).strip():
        return "Error: QR-এ কী থাকবে সেটা দরকার (লিংক/নম্বর/লেখা)।"
    try:
        import qrcode
    except ImportError:
        return "Error: qrcode library নেই (requirements.txt-এ আছে, deploy-এ পাওয়া যাবে)।"
    name = re.sub(r"[^A-Za-z0-9_-]", "_", Path(filename or "qr").stem)[:30] or "qr"
    try:
        img = qrcode.make(str(text).strip()[:1000])
        fname = f"{name}.png"
        img.save(str(SANDBOX / fname))
        log_activity("tool", f"QR: {str(text)[:40]}")
        return (f"📱 QR code তৈরি!\n📁 নতুন ফাইল তৈরি হয়েছে: {fname} (sandbox ফোল্ডারে)\n"
                f"ভেতরে আছে: {str(text).strip()[:80]}")
    except Exception as e:
        return f"QR বানাতে সমস্যা: {e}"

def generate_password(length="16", count="3"):
    """শক্তিশালী random password বানায়।"""
    import secrets, string
    try:
        n = max(8, min(64, int(length)))
        cnt = max(1, min(10, int(count)))
    except (TypeError, ValueError):
        n, cnt = 16, 3
    chars = string.ascii_letters + string.digits + "!@#$%^&*-_+="
    pws = []
    for _ in range(cnt):
        pw = (secrets.choice(string.ascii_uppercase) + secrets.choice(string.ascii_lowercase)
              + secrets.choice(string.digits) + secrets.choice("!@#$%-_")
              + "".join(secrets.choice(chars) for _ in range(n - 4)))
        pws.append("".join(secrets.SystemRandom().sample(pw, len(pw))))
    return (f"🔐 {cnt}টা শক্তিশালী password ({n} অক্ষর):\n"
            + "\n".join(f"{i+1}. {p}" for i, p in enumerate(pws))
            + "\n\n⚠️ নিরাপদ জায়গায় রাখুন — এগুলো আর কোথাও save হয়নি।")

def unit_convert(value, from_unit, to_unit):
    """একক রূপান্তর: দৈর্ঘ্য/ওজন/তাপমাত্রা/জমি (বাংলাদেশি একক সহ!)।"""
    try:
        v = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return "Error: value সংখ্যা হতে হবে।"
    f = str(from_unit).lower().strip()
    t = str(to_unit).lower().strip()
    # তাপমাত্রা আলাদা
    temps = {"c": "celsius", "f": "fahrenheit", "celsius": "celsius", "fahrenheit": "fahrenheit"}
    if f in temps and t in temps:
        if temps[f] == temps[t]:
            return f"{v} — একই একক!"
        r = v * 9 / 5 + 32 if temps[f] == "celsius" else (v - 32) * 5 / 9
        return f"🌡️ {v}° {temps[f]} = {r:.1f}° {temps[t]}"
    # মিটারে base (দৈর্ঘ্য), কেজিতে (ওজন), বর্গফুটে (জমি)
    U = {
        "km": ("length", 1000), "kilometer": ("length", 1000), "m": ("length", 1), "meter": ("length", 1),
        "cm": ("length", .01), "mm": ("length", .001), "mile": ("length", 1609.34),
        "ft": ("length", .3048), "feet": ("length", .3048), "foot": ("length", .3048),
        "inch": ("length", .0254), "yard": ("length", .9144), "hat": ("length", .4572), "হাত": ("length", .4572),
        "kg": ("weight", 1), "kilogram": ("weight", 1), "g": ("weight", .001), "gram": ("weight", .001),
        "lb": ("weight", .4536), "pound": ("weight", .4536), "ton": ("weight", 1000),
        "mon": ("weight", 37.324), "মণ": ("weight", 37.324), "maund": ("weight", 37.324),
        "ser": ("weight", .9331), "সের": ("weight", .9331), "tola": ("weight", .011664), "ভরি": ("weight", .011664), "vori": ("weight", .011664),
        "sqft": ("area", 1), "square feet": ("area", 1), "katha": ("area", 720), "কাঠা": ("area", 720),
        "bigha": ("area", 14400), "বিঘা": ("area", 14400), "shotok": ("area", 435.6), "শতক": ("area", 435.6),
        "decimal": ("area", 435.6), "acre": ("area", 43560), "একর": ("area", 43560),
    }
    if f not in U or t not in U:
        return ("Error: একক চিনলাম না। পারি: km/m/cm/ft/inch/mile/হাত | kg/g/মণ/সের/ভরি/lb | "
                "কাঠা/বিঘা/শতক/একর/sqft | C/F")
    if U[f][0] != U[t][0]:
        return f"Error: {f} ({U[f][0]}) আর {t} ({U[t][0]}) ভিন্ন ধরনের একক!"
    r = v * U[f][1] / U[t][1]
    return f"📏 {v:,.4g} {from_unit} = {r:,.4g} {to_unit}"

def export_chat(last_n="50"):
    """কথোপকথনের সারাংশ text ফাইলে save করে (share/print-এর জন্য)।"""
    try:
        n = max(5, min(200, int(last_n)))
    except (TypeError, ValueError):
        n = 50
    hist = load_history()
    lines = [f"═══ {BRAND['name']} — কথোপকথন Export ═══",
             f"তারিখ: {now_local().strftime('%Y-%m-%d %H:%M')} | শেষ {n}টা message", ""]
    count = 0
    for msg in hist[-n*2:]:
        parts = msg.get("parts", [])
        text = " ".join(p.get("text", "") for p in parts if "text" in p).strip()
        if not text:
            continue
        text = re.sub(r"^\[[^\]]*\]\s*", "", text)
        who = "👤 আপনি" if msg.get("role") == "user" else "🤖 Agent"
        lines.append(f"{who}: {text[:600]}")
        lines.append("")
        count += 1
    if count == 0:
        return "Export করার মতো কথোপকথন নেই।"
    fname = f"chat_export_{now_local().strftime('%Y%m%d_%H%M')}.txt"
    (SANDBOX / fname).write_text("\n".join(lines), encoding="utf-8")
    log_activity("tool", f"chat export: {count} messages")
    return f"💾 কথোপকথন export হলো!\n📁 নতুন ফাইল তৈরি হয়েছে: {fname} (sandbox ফোল্ডারে)\n{count}টা message।"

def text_stats(text):
    """লেখার পরিসংখ্যান: শব্দ, অক্ষর, বাক্য, পড়ার সময়।"""
    s = str(text).strip()
    if not s:
        return "Error: লেখা দরকার।"
    words = len(re.findall(r"[\w\u0980-\u09FF]+", s))
    chars = len(s)
    chars_ns = len(re.sub(r"\s", "", s))
    sents = max(1, len(re.findall(r"[.!?।]+", s)))
    read_min = max(1, round(words / 200))
    return (f"📊 লেখার পরিসংখ্যান:\n• শব্দ: {words:,}\n• অক্ষর: {chars:,} (স্পেস ছাড়া {chars_ns:,})\n"
            f"• বাক্য: {sents:,}\n• গড় শব্দ/বাক্য: {words/sents:.1f}\n• পড়ার সময়: ~{read_min} মিনিট")

TOOLS.update({
    "make_qr_code": {"func": make_qr_code, "declaration": _decl(
        "make_qr_code", "QR code বানায় (লিংক/bKash নম্বর/দোকানের ঠিকানা/যেকোনো লেখা) — ছবি হিসেবে দেখায়। ব্যবসার visiting card/পোস্টারে দারুণ কাজের!",
        {"text": {"type": "string", "description": "QR-এ যা থাকবে (লিংক/লেখা)"},
         "filename": {"type": "string", "description": "ফাইলের নাম (English)"}}, ["text"])},
    "generate_password": {"func": generate_password, "declaration": _decl(
        "generate_password", "শক্তিশালী random password বানায়।",
        {"length": {"type": "integer", "description": "কত অক্ষর (৮-৬৪)"},
         "count": {"type": "integer", "description": "কয়টা (১-১০)"}}, [])},
    "unit_convert": {"func": unit_convert, "declaration": _decl(
        "unit_convert", "একক রূপান্তর — বাংলাদেশি একক সহ! কাঠা/বিঘা/শতক/একর, মণ/সের/ভরি, km/ft/inch, C/F।",
        {"value": {"type": "number", "description": "মান"},
         "from_unit": {"type": "string", "description": "কোন একক থেকে"},
         "to_unit": {"type": "string", "description": "কোন এককে"}}, ["value", "from_unit", "to_unit"])},
    "export_chat": {"func": export_chat, "declaration": _decl(
        "export_chat", "পুরো কথোপকথন text ফাইলে save করে — share বা রেকর্ড রাখার জন্য।",
        {"last_n": {"type": "integer", "description": "শেষ কয়টা message (default 50)"}}, [])},
    "text_stats": {"func": text_stats, "declaration": _decl(
        "text_stats", "লেখার শব্দ/অক্ষর/বাক্য গোনা + পড়ার সময়। লেখালেখি/assignment-এ কাজের।",
        {"text": {"type": "string", "description": "যে লেখা বিশ্লেষণ হবে"}}, ["text"])},
})
TOOL_PERMISSION.update({
    "make_qr_code": "code", "export_chat": "files",
})


# ══════════ 🧩 PLUGIN SYSTEM — নিজের tool যোগ করুন! ══════════
# নিচের মতো ৩ লাইনেই নতুন plugin/tool যোগ করা যায় — agent নিজেই শিখে নেবে:
#
# def amar_plugin(text: str) -> str:
#     return "ফলাফল: " + text.upper()
# TOOLS["amar_plugin"] = {"func": amar_plugin, "declaration": _decl(
#     "amar_plugin", "কী কাজ করে তার বর্ণনা (agent এটা পড়ে বোঝে কখন ব্যবহার করবে)",
#     {"text": {"type": "string", "description": "input"}}, ["text"])}


# ═══════════════════════════════════════════════════════════════════
# ৮. AGENT — মূল মস্তিষ্কের লুপ (ভাবা → tool চালানো → উত্তর)
# ═══════════════════════════════════════════════════════════════════

class Agent:
    def __init__(self):
        self.history = load_history()  # restart করলেও আগের কথা মনে থাকে
        self._lock = threading.Lock()  # একসাথে দুটো chat এলে সিরিয়ালি চলবে

    def run(self, user_message: str, verbose: bool = True, tool_log=None, image=None) -> str:
        with self._lock:
            return self._run_locked(user_message, verbose, tool_log, image)

    def _sanitize_history(self):
        """ছবির base64 history থেকে সরিয়ে ছোট marker রাখে।"""
        for msg in self.history:
            parts = msg.get("parts", [])
            for i, p in enumerate(parts):
                if "inline_data" in p:
                    parts[i] = {"text": "[📷 এখানে user একটা ছবি পাঠিয়েছিল]"}

    def _run_locked(self, user_message: str, verbose: bool = True, tool_log=None, image=None) -> str:
        parts = [{"text": user_message}]
        if image and image.get("data"):
            parts.append({"inline_data": {
                "mime_type": image.get("mime", "image/jpeg"),
                "data": image["data"],
            }})
        self.history.append({"role": "user", "parts": parts})
        for _ in range(get_max_iters()):
            response = llm_chat(trim_history(self.history), tools=get_declarations())
            self.history.append(response)
            function_calls = [p["functionCall"] for p in response.get("parts", [])
                              if "functionCall" in p]
            if not function_calls:
                text_parts = [p.get("text", "") for p in response.get("parts", [])]
                self._sanitize_history()
                save_history(self.history)
                return "".join(text_parts).strip()
            result_parts = []
            for call in function_calls:
                name = call["name"]
                args = call.get("args", {})
                if verbose:
                    print(f"  🔧 [tool] {name}({args})")
                result = execute_tool(name, args)
                if verbose:
                    print(f"  ✅ [result] {result[:120]}")
                if tool_log is not None:
                    tool_log.append(
                        f"{name}({json.dumps(args, ensure_ascii=False)[:150]}) → {result[:100]}")
                result_parts.append({"functionResponse": {
                    "name": name, "response": {"result": result}}})
            self.history.append({"role": "user", "parts": result_parts})
        self._sanitize_history()
        save_history(self.history)
        return "দুঃখিত, অনেকবার চেষ্টা করেও কাজটা শেষ করতে পারিনি।"

    def reset(self):
        self.history = []
        clear_history()


# ═══════════════════════════════════════════════════════════════════
# ৯. WEB UI — ব্রাউজারে chat (upload + chart display সহ)
# ═══════════════════════════════════════════════════════════════════

HTML_PAGE = """<!DOCTYPE html>
<html lang="bn">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover">
<meta name="theme-color" content="#7c3aed">
<link rel="manifest" href="manifest.json">
<link rel="icon" href="logo.png">
<link rel="apple-touch-icon" href="logo.png">
<title>Sabbir AI Agent</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800&family=Hind+Siliguri:wght@400;500;600;700&family=Playfair+Display:wght@700;800&display=swap" rel="stylesheet">
<style>
* { margin:0; padding:0; box-sizing:border-box; -webkit-tap-highlight-color:transparent; }

:root {
  --bg-base:#f3f1fb;
  --blob1:#ddd3fb; --blob2:#e9e2fd; --blob3:#c4b5fd;
  --panel:rgba(255,254,255,.65);
  --stroke:rgba(124,58,237,.18);
  --text:#1e1440; --dim:#6d6493;
  --g1:#7c3aed; --g2:#a78bfa; --g3:#6366f1;
  --bot-bg:rgba(255,255,255,.8); --bot-stroke:rgba(124,58,237,.15);
  --tool-bg:rgba(16,185,129,.1); --tool-stroke:rgba(16,185,129,.3); --tool-text:#059669;
  --err-bg:rgba(239,68,68,.08); --err-stroke:rgba(239,68,68,.3); --err-text:#dc2626;
  --field:rgba(255,255,255,.85);
  --glow:0 10px 40px -10px rgba(124,58,237,.45);
  --card-shadow:0 8px 32px rgba(60,30,140,.12);
  --chip-bg:rgba(255,255,255,.75);
  --side-bg:rgba(252,250,255,.6);
}
[data-theme="dark"] {
  --bg-base:#0d0521;
  --blob1:#3b1d8f; --blob2:#6d28d9; --blob3:#1e1b4b;
  --panel:rgba(20,12,45,.6);
  --stroke:rgba(167,139,250,.18);
  --text:#ece9fe; --dim:#9d94c8;
  --bot-bg:rgba(30,20,60,.7); --bot-stroke:rgba(167,139,250,.16);
  --tool-bg:rgba(16,185,129,.09); --tool-stroke:rgba(16,185,129,.25); --tool-text:#34d399;
  --err-bg:rgba(239,68,68,.09); --err-stroke:rgba(239,68,68,.28); --err-text:#f87171;
  --field:rgba(13,7,32,.7);
  --glow:0 10px 45px -8px rgba(139,92,246,.55);
  --card-shadow:0 10px 36px rgba(0,0,0,.5);
  --chip-bg:rgba(35,24,70,.65);
  --side-bg:rgba(18,10,40,.55);
}

html,body { height:100%; }
body {
  font-family:'Inter','Hind Siliguri','Segoe UI',system-ui,sans-serif;
  background:var(--bg-base); color:var(--text);
  display:flex; flex-direction:column;
  transition:background .6s ease, color .3s ease;
  overflow:hidden;
}

/* Animated background blobs */
.blob { position:fixed; border-radius:50%; filter:blur(90px); opacity:.55; z-index:0;
        pointer-events:none; transition:background 1s ease; }
#blob1 { width:48vw; height:48vw; background:var(--blob1); top:-14vw; left:-12vw;
         animation:drift1 19s ease-in-out infinite alternate; }
#blob2 { width:40vw; height:40vw; background:var(--blob2); bottom:-12vw; right:-10vw;
         animation:drift2 23s ease-in-out infinite alternate; }
#blob3 { width:32vw; height:32vw; background:var(--blob3); top:38%; left:55%;
         animation:drift3 27s ease-in-out infinite alternate; }
@keyframes drift1 { to { transform:translate(7vw,5vh) scale(1.18); } }
@keyframes drift2 { to { transform:translate(-6vw,-7vh) scale(1.12); } }
@keyframes drift3 { to { transform:translate(-9vw,6vh) scale(.88); } }

/* ══════════ HEADER ══════════ */
header {
  position:relative; z-index:6;
  display:flex; align-items:center; gap:13px;
  padding:12px 20px;
  background:var(--panel);
  backdrop-filter:blur(22px); -webkit-backdrop-filter:blur(22px);
  border-bottom:1px solid var(--stroke);
}
.logo {
  width:42px; height:42px; border-radius:13px; flex-shrink:0;
  display:flex; align-items:center; justify-content:center;
  background:linear-gradient(135deg,var(--g1),var(--g2),var(--g3));
  box-shadow:var(--glow); font-size:21px; overflow:hidden;
  animation:pulse 3.5s ease-in-out infinite;
}
@keyframes pulse { 0%,100% { transform:scale(1); } 50% { transform:scale(1.06); } }
.title-wrap { flex:1; min-width:0; }
.title {
  font-size:18px; font-weight:800; letter-spacing:.5px;
  font-family:'Playfair Display','Hind Siliguri',serif;
  background:linear-gradient(90deg,var(--g1),var(--g2),var(--g3));
  -webkit-background-clip:text; background-clip:text; color:transparent;
  white-space:nowrap; overflow:hidden; text-overflow:ellipsis;
}
.subtitle { font-size:11.5px; color:var(--dim); margin-top:1px; }
.status {
  font-size:11.5px; font-weight:600; color:var(--dim);
  padding:6px 12px; border-radius:999px;
  background:var(--chip-bg); border:1px solid var(--stroke);
  display:flex; align-items:center; gap:6px; white-space:nowrap;
}
.status .dot { width:8px; height:8px; border-radius:50%; background:#f59e0b;
               animation:blink 1.6s ease-in-out infinite; }
.status.ok .dot { background:#10b981; }
@keyframes blink { 50% { opacity:.35; } }
#menubtn { display:none; }
#scrim { display:none; position:fixed; inset:0; background:rgba(0,0,0,.55); z-index:55;
  backdrop-filter:blur(2px); }
#scrim.show { display:block; }
#quickacts { display:flex; flex-direction:column; gap:2px; margin-top:16px; width:100%; max-width:430px; }
.qa { display:flex; align-items:center; gap:13px; padding:13px 16px; border-radius:14px;
  border:none; background:transparent; color:var(--text); font-size:15px; font-family:inherit;
  cursor:pointer; text-align:left; transition:background .2s ease, transform .15s ease; }
.qa:hover { background:var(--chip-bg); transform:translateX(3px); }
.qa .qi { font-size:19px; width:26px; text-align:center; }
.sidelabel { font-size:10.5px; font-weight:800; color:var(--dim); letter-spacing:1px;
  padding:10px 14px 3px; text-transform:uppercase; }
.iconbtn {
  width:40px; height:40px; border-radius:12px; border:1px solid var(--stroke);
  background:var(--chip-bg); color:var(--text); cursor:pointer;
  display:flex; align-items:center; justify-content:center; flex-shrink:0;
  transition:transform .2s cubic-bezier(.34,1.56,.64,1), box-shadow .2s ease;
}
.iconbtn:hover { transform:translateY(-2px) scale(1.06); box-shadow:var(--card-shadow); }
.iconbtn:active { transform:scale(.92); }
.iconbtn svg { width:19px; height:19px; }

/* ══════════ SETTINGS PANEL ══════════ */
#settingspanel {
  position:relative; z-index:5; display:none; flex-direction:column; gap:9px;
  padding:13px 20px;
  background:var(--panel); backdrop-filter:blur(18px);
  border-bottom:1px solid var(--stroke);
  animation:slideDown .35s ease both;
}
@keyframes slideDown { from { opacity:0; transform:translateY(-12px); } }
.setrow { display:flex; gap:9px; }
.setrow input {
  flex:1; padding:11px 15px; border-radius:12px; min-width:0;
  border:1px solid var(--bot-stroke); background:var(--field); color:var(--text);
  font-size:13.5px; font-family:inherit; outline:none;
  transition:border .25s ease, box-shadow .25s ease;
}
.setrow input:focus { border-color:var(--g2); box-shadow:0 0 0 4px rgba(139,92,246,.15); }
.gbtn {
  padding:11px 19px; border:none; border-radius:12px; cursor:pointer;
  font-size:13.5px; font-weight:700; font-family:inherit; color:#fff; flex-shrink:0;
  background:linear-gradient(135deg,var(--g1),var(--g2));
  box-shadow:var(--glow);
  transition:transform .2s cubic-bezier(.34,1.56,.64,1), filter .2s ease;
}
.gbtn:hover { transform:translateY(-2px) scale(1.03); filter:brightness(1.1); }
.gbtn:active { transform:scale(.95); }

/* ══════════ SHELL: SIDEBAR + MAIN ══════════ */
#shell { position:relative; flex:1; display:flex; min-height:0; }

#sidebar {
  width:190px; flex-shrink:0; padding:14px 10px;
  display:flex; flex-direction:column; gap:5px;
  background:var(--side-bg);
  backdrop-filter:blur(20px); -webkit-backdrop-filter:blur(20px);
  border-right:1px solid var(--stroke);
  overflow-y:auto;
}
.nav {
  display:flex; align-items:center; gap:11px;
  padding:11px 14px; border-radius:13px; border:1px solid transparent;
  background:transparent; color:var(--text); cursor:pointer;
  font-size:14px; font-weight:600; font-family:inherit; text-align:left;
  transition:transform .18s cubic-bezier(.34,1.56,.64,1), background .2s, border .2s, box-shadow .2s;
  white-space:nowrap;
}
.nav .ic { font-size:17px; width:22px; text-align:center; flex-shrink:0; }
.nav:hover { background:var(--chip-bg); border-color:var(--stroke); transform:translateX(3px); }
.nav:active { transform:scale(.96); }
.nav.active {
  background:linear-gradient(135deg,var(--g1),var(--g2));
  color:#fff; box-shadow:var(--glow); border-color:transparent;
}
.side-foot { margin-top:auto; padding:10px 6px 4px; font-size:10.5px; color:var(--dim);
             text-align:center; }

#main { flex:1; display:flex; flex-direction:column; min-width:0; min-height:0; }

/* ══════════ HERO (welcome screen) ══════════ */
#hero {
  display:flex; flex-direction:column; align-items:center; justify-content:center;
  text-align:center; padding:34px 24px 10px; gap:12px;
  animation:rise .5s ease both;
}
.hero-badge {
  width:72px; height:72px; border-radius:24px; overflow:hidden;
  background:linear-gradient(135deg,var(--g1),var(--g2),var(--g3));
  box-shadow:var(--glow); display:flex; align-items:center; justify-content:center;
  font-size:36px; animation:pulse 3.5s ease-in-out infinite;
}
.hero-badge img { width:100%; height:100%; object-fit:cover; }
#hero h1 {
  font-size:clamp(22px,4.5vw,32px); font-weight:800; letter-spacing:.5px;
  font-family:'Playfair Display','Hind Siliguri',serif;
  background:linear-gradient(90deg,var(--g1),var(--g2),var(--g3));
  -webkit-background-clip:text; background-clip:text; color:transparent;
}
#hero p { font-size:14.5px; color:var(--dim); max-width:460px; line-height:1.7; white-space:pre-wrap; }

/* ══════════ CHAT ══════════ */
#chat {
  flex:1; overflow-y:auto; overflow-x:hidden;
  padding:18px 20px 10px;
  display:flex; flex-direction:column; gap:12px;
  scroll-behavior:smooth;
}
#chat::-webkit-scrollbar { width:6px; }
#chat::-webkit-scrollbar-thumb { background:var(--bot-stroke); border-radius:99px; }
.msg {
  max-width:74%; padding:13px 17px;
  border-radius:20px; line-height:1.65; font-size:15px;
  white-space:pre-wrap; word-wrap:break-word;
  animation:rise .38s cubic-bezier(.21,1.02,.73,1) both;
  box-shadow:var(--card-shadow);
}
@keyframes rise { from { opacity:0; transform:translateY(14px) scale(.97); } }
.user {
  align-self:flex-end;
  background:linear-gradient(135deg,var(--g1),var(--g2));
  color:#fff; border-bottom-right-radius:6px;
  box-shadow:var(--glow);
}
.bot {
  align-self:flex-start;
  background:var(--bot-bg); border:1px solid var(--bot-stroke);
  border-bottom-left-radius:6px;
  backdrop-filter:blur(14px); -webkit-backdrop-filter:blur(14px);
  white-space:normal;
}
.bot-row { display:flex; align-items:flex-end; gap:9px; align-self:flex-start; max-width:82%;
           animation:rise .38s cubic-bezier(.21,1.02,.73,1) both; }
.bot-row .msg { animation:none; max-width:100%; }
.avatar {
  width:30px; height:30px; border-radius:50%; flex-shrink:0;
  background:linear-gradient(135deg,var(--g1),var(--g2)); overflow:hidden;
  box-shadow:0 3px 10px rgba(99,102,241,.35);
  display:flex; align-items:center; justify-content:center; font-size:16px;
}
.avatar img { width:100%; height:100%; object-fit:cover; }
.tool {
  align-self:flex-start; max-width:86%;
  background:var(--tool-bg); border:1px solid var(--tool-stroke); color:var(--tool-text);
  font-size:12.5px; font-family:'SF Mono',Consolas,monospace;
  border-radius:14px; padding:9px 14px;
}
.err {
  align-self:center; text-align:center;
  background:var(--err-bg); border:1px solid var(--err-stroke); color:var(--err-text);
  font-size:13px; border-radius:14px;
}
.chatimg {
  max-width:74%; border-radius:20px; align-self:flex-start;
  border:1px solid var(--bot-stroke); box-shadow:var(--card-shadow);
  animation:rise .45s ease both;
}
.typing-row { display:flex; align-items:center; gap:10px; }
.tlabel { color:var(--dim); font-size:13.5px; }
.dots { display:flex; gap:5px; }
.dots i {
  width:7px; height:7px; border-radius:50%;
  background:linear-gradient(135deg,var(--g1),var(--g2));
  animation:bounce 1.2s ease-in-out infinite;
}
.dots i:nth-child(2) { animation-delay:.15s; }
.dots i:nth-child(3) { animation-delay:.3s; }
@keyframes bounce { 0%,60%,100% { transform:translateY(0); opacity:.5; }
                    30% { transform:translateY(-6px); opacity:1; } }
.think-ring { position:relative; }
.think-ring::after {
  content:''; position:absolute; inset:-4px; border-radius:50%;
  border:2px solid transparent;
  border-top-color:var(--g2); border-right-color:var(--g3);
  animation:spinring .9s linear infinite;
}
@keyframes spinring { to { transform:rotate(360deg); } }

/* ══════════ CHIPS + INPUT ══════════ */
#chips {
  display:flex; gap:9px; padding:4px 20px 10px; flex-wrap:wrap; justify-content:center;
}
.chip {
  padding:9px 16px; border-radius:999px; cursor:pointer;
  font-size:13px; font-weight:600; font-family:inherit; color:var(--text);
  background:var(--chip-bg); border:1px solid var(--stroke);
  backdrop-filter:blur(12px);
  transition:transform .2s cubic-bezier(.34,1.56,.64,1), border-color .2s, box-shadow .2s;
  animation:rise .5s ease both;
}
.chip:hover { transform:translateY(-2px) scale(1.04); border-color:var(--g2); box-shadow:var(--card-shadow); }
.chip:active { transform:scale(.94); }

#inputbar {
  position:relative; z-index:5;
  display:flex; gap:10px; align-items:center;
  padding:13px 20px calc(13px + env(safe-area-inset-bottom));
  background:var(--panel);
  backdrop-filter:blur(22px); -webkit-backdrop-filter:blur(22px);
  border-top:1px solid var(--stroke);
}
#msginput {
  flex:1; padding:14px 18px; border-radius:16px;
  border:1px solid var(--bot-stroke); background:var(--field); color:var(--text);
  font-size:15px; font-family:inherit; outline:none; min-width:0;
  transition:border .25s ease, box-shadow .25s ease;
}
#msginput:focus { border-color:var(--g2); box-shadow:0 0 0 4px rgba(139,92,246,.15); }
#msginput::placeholder { color:var(--dim); }
#sendbtn {
  width:50px; height:50px; border:none; border-radius:16px; cursor:pointer;
  display:flex; align-items:center; justify-content:center; flex-shrink:0;
  background:linear-gradient(135deg,var(--g1),var(--g2),var(--g3));
  color:#fff; box-shadow:var(--glow);
  transition:transform .2s cubic-bezier(.34,1.56,.64,1);
}
#sendbtn:hover { transform:translateY(-2px) scale(1.07); }
#sendbtn:active { transform:scale(.9); }
#sendbtn:disabled { filter:grayscale(.5) opacity(.7); cursor:wait; transform:none; }
#sendbtn svg { width:22px; height:22px; }

/* ══════════ DASHBOARD ══════════ */
#dashboard, #command, #control, #skills, #adminpanel { display:none; flex:1; flex-direction:column; overflow-y:auto; padding:18px 20px; gap:14px; }
#dashboard::-webkit-scrollbar, #command::-webkit-scrollbar { width:6px; }
#dashboard::-webkit-scrollbar-thumb, #command::-webkit-scrollbar-thumb { background:var(--bot-stroke); border-radius:99px; }
.dashhead { display:flex; align-items:center; gap:10px; }
.dashhead h2 { font-size:18px; font-weight:800;
  background:linear-gradient(90deg,var(--g1),var(--g2),var(--g3));
  -webkit-background-clip:text; background-clip:text; color:transparent; flex:1; }
.dashtime { font-size:11.5px; color:var(--dim); }
.dashgrid { display:grid; grid-template-columns:repeat(auto-fill,minmax(250px,1fr)); gap:14px; padding-bottom:12px; }
.dcard {
  background:var(--bot-bg); border:1px solid var(--bot-stroke); border-radius:18px;
  padding:15px 16px; box-shadow:var(--card-shadow);
  backdrop-filter:blur(14px); -webkit-backdrop-filter:blur(14px);
  animation:rise .45s cubic-bezier(.21,1.02,.73,1) both;
  transition:transform .2s ease, box-shadow .2s ease;
}
.dcard:hover { transform:translateY(-3px); box-shadow:var(--glow); }
.dcard h3 { font-size:13.5px; font-weight:700; display:flex; align-items:center; gap:8px; margin-bottom:9px; }
.dcard h3 .dic {
  width:28px; height:28px; border-radius:9px; display:flex; align-items:center; justify-content:center;
  background:linear-gradient(135deg,var(--g1),var(--g2)); font-size:14px; flex-shrink:0;
  box-shadow:0 3px 10px rgba(99,102,241,.3);
}
.dline { font-size:12.5px; color:var(--text); padding:5px 2px; border-bottom:1px dashed var(--bot-stroke);
         line-height:1.5; word-wrap:break-word; }
.dline:last-child { border-bottom:none; }
.dempty { font-size:12.5px; color:var(--dim); font-style:italic; padding:4px 2px; line-height:1.5; }

.bigbtn {
  width:100%; padding:15px 20px; border:none; border-radius:16px; cursor:pointer;
  font-size:15px; font-weight:800; font-family:inherit; color:#fff;
  background:linear-gradient(135deg,var(--g1),var(--g2),var(--g3));
  box-shadow:var(--glow); letter-spacing:.2px;
  transition:transform .2s cubic-bezier(.34,1.56,.64,1), filter .2s ease;
}
.bigbtn:hover { transform:translateY(-2px) scale(1.015); filter:brightness(1.1); }
.bigbtn:active { transform:scale(.97); }
.statbar { display:flex; align-items:center; gap:8px; padding:6px 2px; }
.statbar .lbl { font-size:12.5px; width:86px; flex-shrink:0; font-weight:600; }
.statbar .bar { flex:1; height:9px; border-radius:99px; background:var(--bot-stroke); overflow:hidden; }
.statbar .fill { height:100%; border-radius:99px;
  background:linear-gradient(90deg,var(--g1),var(--g2)); transition:width .6s ease; }
.statbar .val { font-size:12px; color:var(--dim); width:38px; text-align:right; flex-shrink:0; }

.crow { display:flex; gap:8px; align-items:center; padding:6px 2px; border-bottom:1px dashed var(--bot-stroke); }
.crow:last-child { border-bottom:none; }
.crow .ctext { flex:1; font-size:12.5px; line-height:1.5; word-wrap:break-word; min-width:0; }
.mini {
  padding:6px 10px; border-radius:9px; border:1px solid var(--bot-stroke); cursor:pointer;
  background:var(--chip-bg); color:var(--text); font-size:11.5px; font-weight:700; font-family:inherit;
  flex-shrink:0; transition:transform .15s ease, border-color .2s;
}
.mini:hover { transform:scale(1.06); border-color:var(--g2); }
.mini.danger { color:var(--err-text); border-color:var(--err-stroke); }
.minifield {
  flex:1; padding:9px 12px; border-radius:10px; min-width:0;
  border:1px solid var(--bot-stroke); background:var(--field); color:var(--text);
  font-size:12.5px; font-family:inherit; outline:none;
}
.toggle { position:relative; width:42px; height:24px; border-radius:99px; border:none; cursor:pointer;
  background:var(--bot-stroke); transition:background .25s ease; flex-shrink:0; }
.toggle.on { background:linear-gradient(90deg,var(--g1),var(--g2)); }
.toggle::after { content:''; position:absolute; top:3px; left:3px; width:18px; height:18px;
  border-radius:50%; background:#fff; transition:left .25s cubic-bezier(.34,1.56,.64,1); }
.toggle.on::after { left:21px; }
.prio-red { border-left:3px solid #ef4444; padding-left:8px; }
.prio-yel { border-left:3px solid #f59e0b; padding-left:8px; }
.prio-grn { border-left:3px solid #10b981; padding-left:8px; }
#micbtn.listening { background:linear-gradient(135deg,#ef4444,#f97316) !important; color:#fff;
  animation:pulse 1s ease-in-out infinite; }
.pbar { height:8px; border-radius:99px; background:var(--bot-stroke); overflow:hidden; margin-top:4px; }
.pbar .pfill { height:100%; background:linear-gradient(90deg,var(--g1),var(--g3)); border-radius:99px; transition:width .5s ease; }

.skillbtn {
  display:block; width:100%; text-align:left; padding:9px 12px; margin:3px 0;
  border-radius:11px; border:1px solid var(--bot-stroke); cursor:pointer;
  background:var(--field); color:var(--text); font-size:12.5px; font-weight:600; font-family:inherit;
  transition:transform .15s ease, border-color .2s ease;
}
.skillbtn:hover { transform:translateX(3px); border-color:var(--g2); }
.skillbtn:active { transform:scale(.97); }

#imgpreview {
  display:none; align-items:center; gap:10px; padding:8px 20px;
  background:var(--panel); border-top:1px solid var(--stroke);
}
#imgpreview img { height:52px; border-radius:10px; border:1px solid var(--bot-stroke); }
#imgpreview .xbtn { padding:5px 11px; border-radius:9px; border:1px solid var(--err-stroke);
  background:var(--err-bg); color:var(--err-text); cursor:pointer; font-size:12px; font-family:inherit; }
#imgbtn.has-img { background:linear-gradient(135deg,var(--g1),var(--g2)) !important; color:#fff; }
.speakbtn {
  margin-top:6px; padding:5px 12px; border-radius:9px; cursor:pointer;
  border:1px solid var(--bot-stroke); background:var(--chip-bg); color:var(--dim);
  font-size:11.5px; font-weight:700; font-family:inherit;
  transition:transform .15s ease, border-color .2s;
}
.speakbtn:hover { transform:scale(1.05); border-color:var(--g2); color:var(--text); }
.speakbtn.playing { background:linear-gradient(135deg,var(--g1),var(--g2)); color:#fff; }

.selfield {
  width:100%; padding:10px 12px; border-radius:10px; margin:3px 0;
  border:1px solid var(--bot-stroke); background:var(--field); color:var(--text);
  font-size:12.5px; font-family:inherit; outline:none;
}
textarea.selfield { resize:vertical; min-height:54px; }
.alabel { font-size:11px; font-weight:700; color:var(--dim); margin-top:8px; display:block; }
.colorrow { display:flex; gap:8px; align-items:center; margin:4px 0; }
.colorrow input[type=color] { width:42px; height:34px; border:none; border-radius:8px;
  background:transparent; cursor:pointer; }
.provrow { border:1px solid var(--bot-stroke); border-radius:12px; padding:9px 11px; margin:5px 0; }
.provrow.active { border-color:var(--g2); box-shadow:0 0 0 3px rgba(139,92,246,.12); }
.badge { font-size:10px; font-weight:800; padding:3px 8px; border-radius:99px; }
.badge.on { background:rgba(16,185,129,.15); color:#10b981; }
.badge.off { background:var(--bot-stroke); color:var(--dim); }

/* ══════════ RESPONSIVE ══════════ */
@media (max-width:720px) {
  #shell { flex-direction:column; }
  #menubtn { display:flex; }
  #sidebar {
    position:fixed; top:0; left:-280px; width:262px; height:100%;
    flex-direction:column; padding:18px 12px; gap:4px; z-index:60;
    background:var(--bg-base); border-right:1px solid var(--stroke);
    transition:left .28s ease; overflow-y:auto;
  }
  #sidebar.open { left:0; box-shadow:20px 0 60px rgba(0,0,0,.55); }
  header { padding:10px 12px; gap:9px; }
  .subtitle { display:none; }
  .title { font-size:15.5px; }
  .status { font-size:10.5px; padding:5px 9px; }
  #chat { padding:14px 12px 8px; }
  .msg { max-width:86%; font-size:14.5px; }
  .bot-row { max-width:92%; }
  .chatimg { max-width:86%; }
  #inputbar { padding:10px 12px calc(10px + env(safe-area-inset-bottom)); gap:8px; }
  #chips { padding:4px 12px 8px; }
  #hero { padding:22px 16px 6px; }
  #dashboard, #command, #control, #skills, #adminpanel { padding:12px 12px; }
  .hero-badge { width:58px; height:58px; border-radius:19px; font-size:29px; }
  .blob { filter:blur(60px); }
}
@media (prefers-reduced-motion:reduce) {
  *, .blob { animation:none !important; transition:none !important; }
}
</style>
</head>
<body>
<div class="blob" id="blob1"></div>
<div class="blob" id="blob2"></div>
<div class="blob" id="blob3"></div>

<div id="scrim" onclick="toggleDrawer(false)"></div>
<header>
  <button class="iconbtn" id="menubtn" onclick="toggleDrawer()" title="Menu">
    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round"><line x1="3" y1="6" x2="21" y2="6"></line><line x1="3" y1="12" x2="21" y2="12"></line><line x1="3" y1="18" x2="21" y2="18"></line></svg>
  </button>
  <div class="logo" id="logobox"><img id="logoimg" src="logo.png" alt="" style="width:100%;height:100%;object-fit:cover" onerror="this.style.display='none';document.getElementById('logobox').textContent=String.fromCodePoint(0x2726)"></div>
  <div class="title-wrap">
    <div class="title" id="brandname">Sabbir AI Agent</div>
    <div class="subtitle" id="brandtag">Personal Intelligence System</div>
  </div>
  <span class="status" id="status"><span class="dot"></span><span id="statustext">চেক হচ্ছে...</span></span>
  <button class="iconbtn" id="themebtn" onclick="toggleTheme()" title="Dark / Light mode">
    <svg id="iconMoon" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"></path></svg>
    <svg id="iconSun" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round" style="display:none"><circle cx="12" cy="12" r="5"></circle><line x1="12" y1="1" x2="12" y2="3"></line><line x1="12" y1="21" x2="12" y2="23"></line><line x1="4.22" y1="4.22" x2="5.64" y2="5.64"></line><line x1="18.36" y1="18.36" x2="19.78" y2="19.78"></line><line x1="1" y1="12" x2="3" y2="12"></line><line x1="21" y1="12" x2="23" y2="12"></line><line x1="4.22" y1="19.78" x2="5.64" y2="18.36"></line><line x1="18.36" y1="5.64" x2="19.78" y2="4.22"></line></svg>
  </button>
  <button class="iconbtn" onclick="toggleSettings()" title="Settings">
    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="3"></circle><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"></path></svg>
  </button>
  <button class="iconbtn" onclick="resetChat()" title="নতুন কথোপকথন">
    <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 4 23 10 17 10"></polyline><path d="M20.49 15a9 9 0 1 1-2.12-9.36L23 10"></path></svg>
  </button>
</header>

<div id="settingspanel">
  <div class="setrow">
    <input type="password" id="keyinput" placeholder="🔑 Gemini API key (aistudio.google.com/apikey থেকে ফ্রি)">
    <button class="gbtn" onclick="saveKey()">Key সংরক্ষণ</button>
  </div>
  <div class="setrow">
    <input type="password" id="codeinput" placeholder="🔐 Access Code (লাইভ সার্ভারে সেট করা পাসওয়ার্ড)" onkeydown="if(event.key==='Enter')saveCode()">
    <button class="gbtn" onclick="saveCode()">প্রবেশ</button>
  </div>
</div>

<div id="shell">
  <aside id="sidebar"></aside>
  <div id="main">
    <div id="dashboard">
      <div class="dashhead">
        <h2>📊 Agent Dashboard</h2>
        <span class="dashtime" id="dashtime"></span>
        <button class="iconbtn" onclick="loadDashboard()" title="Refresh">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 4 23 10 17 10"></polyline><path d="M20.49 15a9 9 0 1 1-2.12-9.36L23 10"></path></svg>
        </button>
      </div>
      <div class="dashgrid" id="dashgrid"></div>
    </div>
    <div id="command">
      <div class="dashhead">
        <h2>🎛️ Command Center</h2>
        <span class="dashtime" id="cmdtime"></span>
        <button class="iconbtn" onclick="loadCommand()" title="Refresh">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 4 23 10 17 10"></polyline><path d="M20.49 15a9 9 0 1 1-2.12-9.36L23 10"></path></svg>
        </button>
      </div>
      <div class="dashgrid" id="cmdgrid"></div>
      <button class="bigbtn" onclick="analyzeWeek()">🔮 Analyze my week</button>
    </div>
    <div id="control">
      <div class="dashhead">
        <h2>🛠️ Control Center</h2>
        <button class="iconbtn" onclick="loadControl()" title="Refresh">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 4 23 10 17 10"></polyline><path d="M20.49 15a9 9 0 1 1-2.12-9.36L23 10"></path></svg>
        </button>
      </div>
      <div class="dashgrid" id="ctrlgrid"></div>
    </div>
    <div id="adminpanel">
      <div class="dashhead">
        <h2>🔧 Admin Panel</h2>
        <span class="dashtime">সবকিছু নিজের মতো সাজান — সাথে সাথে কার্যকর</span>
        <button class="iconbtn" onclick="loadAdmin()" title="Refresh">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="23 4 23 10 17 10"></polyline><path d="M20.49 15a9 9 0 1 1-2.12-9.36L23 10"></path></svg>
        </button>
      </div>
      <div class="dashgrid" id="admingrid" onclick="if(event.target.closest('[data-act]'))adminAction(event.target.closest('[data-act]').getAttribute('data-act'),event.target.closest('[data-act]'))"></div>
    </div>
    <div id="skills">
      <div class="dashhead">
        <h2>🧰 Skills Library</h2>
        <span class="dashtime">এক ক্লিকে expert কাজ — [ব্র্যাকেট] অংশ বদলে পাঠান</span>
      </div>
      <div class="dashgrid" id="skillgrid"></div>
    </div>
    <div id="chat">
      <div id="hero">
        <div class="hero-badge" id="herobadge"><img src="logo.png" alt="" onerror="this.style.display='none';document.getElementById('herobadge').textContent=String.fromCodePoint(0x1F916)"></div>
        <h1 id="herogreet">শুভ দিন 👋</h1>
        <p id="herosub">আজ আপনার জন্য কী করতে পারি?</p>
        <div id="quickacts">
          <button class="qa" onclick="qa('image')"><span class="qi">🎨</span><span>ছবি বা logo বানাও</span></button>
          <button class="qa" onclick="qa('write')"><span class="qi">✍️</span><span>লিখে দাও বা এডিট করো</span></button>
          <button class="qa" onclick="qa('web')"><span class="qi">🌐</span><span>Web-এ search করো</span></button>
          <button class="qa" onclick="qa('biz')"><span class="qi">🧾</span><span>Order নাও / হিসাব দেখাও</span></button>
        </div>
      </div>
    </div>
    <div id="imgpreview"><img id="previmg" src=""><span style="font-size:12px;color:var(--dim)">ছবি যুক্ত — প্রশ্ন লিখে পাঠান</span><button class="xbtn" onclick="clearImage()">✕ বাদ</button></div>
    <div id="chips"></div>
    <div id="inputbar">
      <input type="file" id="fileinput" style="display:none" accept=".pdf,.txt,.md,.csv,.py,.html,.json,.xlsx" onchange="uploadFile(this.files[0])">
      <input type="file" id="imginput" style="display:none" accept="image/*" onchange="attachImage(this.files[0])">
      <button class="iconbtn" id="imgbtn" onclick="document.getElementById('imginput').click()" title="ছবি পাঠান — Sabbir ছবি দেখে উত্তর দেবে">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="18" rx="2" ry="2"></rect><circle cx="8.5" cy="8.5" r="1.5"></circle><polyline points="21 15 16 10 5 21"></polyline></svg>
      </button>
      <button class="iconbtn" id="attachbtn" onclick="document.getElementById('fileinput').click()" title="ফাইল যোগ করুন">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21.44 11.05l-9.19 9.19a6 6 0 0 1-8.49-8.49l9.19-9.19a4 4 0 0 1 5.66 5.66l-9.2 9.19a2 2 0 0 1-2.83-2.83l8.49-8.48"></path></svg>
      </button>
      <button class="iconbtn" id="micbtn" onclick="toggleVoice()" title="Voice command (বাংলায় বলুন)">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M12 1a3 3 0 0 0-3 3v8a3 3 0 0 0 6 0V4a3 3 0 0 0-3-3z"></path><path d="M19 10v2a7 7 0 0 1-14 0v-2"></path><line x1="12" y1="19" x2="12" y2="23"></line><line x1="8" y1="23" x2="16" y2="23"></line></svg>
      </button>
      <input type="text" id="msginput" placeholder="যা খুশি জিজ্ঞেস করুন..." onkeydown="if(event.key==='Enter')send()">
      <button id="sendbtn" onclick="send()" title="পাঠান">
        <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="22" y1="2" x2="11" y2="13"></line><polygon points="22 2 15 22 11 13 2 9 22 2"></polygon></svg>
      </button>
    </div>
  </div>
</div>

<script>
const chat = document.getElementById('chat');
let accessCode = localStorage.getItem('sabbir_access_code') || '';
let BRAND = {};
let MODE = 'chat';
let hasMessages = false;

/* ══════════ MODES — sidebar-এর প্রতিটা mode ══════════ */
const MODES = {
  dashboard: {
    icon:'📊', label:'Dashboard',
    greet:'', sub:'', placeholder:'', prefix:'', chips:[]
  },
  command: {
    icon:'🎛️', label:'Command',
    greet:'', sub:'', placeholder:'', prefix:'', chips:[]
  },
  control: {
    icon:'🛠️', label:'Control',
    greet:'', sub:'', placeholder:'', prefix:'', chips:[]
  },
  admin: {
    icon:'🔧', label:'Admin',
    greet:'', sub:'', placeholder:'', prefix:'', chips:[]
  },
  skills: {
    icon:'🧰', label:'Skills',
    greet:'', sub:'', placeholder:'', prefix:'', chips:[]
  },
  chat: {
    icon:'💬', label:'Chat',
    greet:'', sub:'আজ আপনার জন্য কী করতে পারি?',
    placeholder:'যা খুশি জিজ্ঞেস করুন...',
    prefix:'',
    chips:null
  },
  think: {
    icon:'🧠', label:'Think',
    greet:'গভীরভাবে ভাবা যাক 🧠', sub:'জটিল সমস্যা দিন — ধাপে ধাপে বিশ্লেষণ করে যুক্তিসহ সিদ্ধান্ত দেবো।',
    placeholder:'কোন সমস্যাটা নিয়ে ভাবতে হবে...',
    prefix:'[Think mode: গভীরভাবে ধাপে ধাপে ভাবো। সমস্যাটা ছোট অংশে ভাঙো, প্রতিটা দিকের সুবিধা-অসুবিধা দেখাও, তারপর যুক্তিসহ চূড়ান্ত সিদ্ধান্ত/পরামর্শ দাও।] ',
    chips:['🤔 ব্যবসা নাকি চাকরি — কোনটা ভালো?','⚖️ এই সিদ্ধান্তের সুবিধা-অসুবিধা দেখাও','🧩 একটা ধাঁধা সমাধান করো','🎯 আমার লক্ষ্যে পৌঁছানোর কৌশল দাও']
  },
  research: {
    icon:'🔎', label:'Research',
    greet:'গভীর রিসার্চ শুরু করি 🔎', sub:'একাধিক উৎস ঘেঁটে, যাচাই করে, উৎসসহ পূর্ণাঙ্গ রিপোর্ট দেবো।',
    placeholder:'কী নিয়ে রিসার্চ করতে হবে...',
    prefix:'[Research mode: web_search ও fetch_webpage দিয়ে একাধিক উৎস থেকে গভীর রিসার্চ করো। তথ্য যাচাই করো, উৎস উল্লেখ করো। বড় research হলে save_report দিয়ে report save করো। কিছু নজরে রাখতে চাইলে watch_tool, watchlist check করতে check_watchlist।] ',
    chips:['👁️ আমার watchlist check করো','📌 ডলারের দাম নজরে রাখো','🌍 আজকের গুরুত্বপূর্ণ খবর','📑 Research report বানাও']
  },
  code: {
    icon:'💻', label:'Coding',
    greet:'চলুন code লিখি 💻', sub:'Python code লিখে চালিয়ে ফলাফল দেখাবো — chart, হিসাব, সব!',
    placeholder:'কী বানাতে/হিসাব করতে চান...',
    prefix:'[Coding mode: run_python_code tool দিয়ে Python code লিখে চালাও। Code-টাও দেখাও, ফলাফলও ব্যাখ্যা করো।] ',
    chips:['📊 একটা bar chart বানাও','🔢 প্রথম ২০টা মৌলিক সংখ্যা','🎲 একটা dice simulator','📈 সুদের হিসাব graph সহ']
  },
  study: {
    icon:'📚', label:'Study',
    greet:'চলুন কিছু শেখা যাক! 📚', sub:'যেকোনো বিষয় — আমি ধাপে ধাপে সহজ করে বুঝিয়ে দেবো।',
    placeholder:'কী শিখতে/বুঝতে চান লিখুন...',
    prefix:'[Study mode: তুমি এখন একজন ধৈর্যশীল শিক্ষক। বিষয়টা ধাপে ধাপে, সহজ উদাহরণ দিয়ে বুঝিয়ে দাও। শেষে ছোট্ট একটা quiz প্রশ্ন করো।] ',
    chips:['📐 পিথাগোরাসের উপপাদ্য বুঝিয়ে দাও','🧪 Photosynthesis কী?','🇬🇧 English tense-এর নিয়ম শেখাও','✍️ আমাকে একটা quiz দাও']
  },
  business: {
    icon:'💼', label:'Business',
    greet:'ব্যবসার কথা হোক 💼', sub:'হিসাব, marketing, পরিকল্পনা — বাস্তবসম্মত পরামর্শ দেবো।',
    placeholder:'ব্যবসা নিয়ে প্রশ্ন করুন...',
    prefix:'[Business mode: তুমি ব্যবসার সহকারী। order নিতে order_tool, পণ্য/দাম/stock-এ product_tool, customer তথ্যে customer_tool, রসিদে make_invoice, অভিযোগে ticket_tool, আয়-ব্যয়ে finance_tool ব্যবহার করো। পরামর্শ লাগলে consultant-এর মতো বাস্তবসম্মত পরামর্শ দাও।] ',
    chips:['🧾 নতুন order নাও','📦 পণ্য তালিকা দেখাও','💰 এ মাসের লাভ-ক্ষতি রিপোর্ট','🎫 Ticket গুলো দেখাও']
  },
  files: {
    icon:'📁', label:'Files',
    greet:'আপনার ফাইল নিয়ে কাজ করি 📁', sub:'PDF/TXT/MD upload করুন (📎 বাটন) — পড়ে সারাংশ বা উত্তর দেবো।',
    placeholder:'ফাইল নিয়ে প্রশ্ন করুন...',
    prefix:'[Files mode: search_documents ও list_documents tool ব্যবহার করে user-এর document থেকে উত্তর দাও। উৎস ফাইলের নাম বলো।] ',
    chips:['📄 আমার কী কী ফাইল আছে?','📝 ফাইলটার সারাংশ দাও','🔍 ফাইলে খুঁজে দাও']
  },
  email: {
    icon:'📧', label:'Email',
    greet:'Email লিখে দিই ✉️', sub:'কাকে, কী বিষয়ে লিখবেন বলুন — professional ড্রাফট বানিয়ে দেবো।',
    placeholder:'যেমন: বসের কাছে ছুটির আবেদন...',
    prefix:'[Email mode: user-এর জন্য সুন্দর email/চিঠি draft করো। Subject সহ, formal ভাষায়। User চাইলে draft-টা emails.txt ফাইলে save_note দিয়ে সংরক্ষণ করো (আগের content পড়ে তার সাথে যোগ করে)।] ',
    chips:['🏢 ছুটির আবেদন লিখে দাও','🤝 Business proposal email','🎓 ভর্তির আবেদনপত্র','🙏 ধন্যবাদ জানানোর email']
  },
  tasks: {
    icon:'📅', label:'Tasks',
    greet:'কাজের তালিকা সামলাই 📅', sub:'Todo, event, plan — লিখে রাখবো, মনে করিয়ে দেবো।',
    placeholder:'যেমন: আজকের কাজের তালিকায় যোগ করো...',
    prefix:'[Tasks mode: save_note/read_note/list_notes tool দিয়ে কাজ manage করো। কাজের তালিকা tasks.txt ফাইলে রাখো (প্রতি লাইনে একটা কাজ)। Event/appointment গুলো events.txt ফাইলে রাখো (তারিখ সহ, প্রতি লাইনে একটা)। নতুন যোগ করার আগে আগের content পড়ে নিয়ে তার সাথে যোগ করো।] ',
    chips:['➕ নতুন কাজ যোগ করো','🗓️ নতুন event যোগ করো','📋 আমার তালিকা দেখাও','✅ একটা কাজ শেষ হয়েছে']
  },
  agents: {
    icon:'⚡', label:'Agents',
    greet:'বড় কাজ? Plan করে করবো ⚡', sub:'জটিল কাজ দিন — ভেঙে ভেঙে ধাপে ধাপে শেষ করবো।',
    placeholder:'বড় কাজটা বর্ণনা করুন...',
    prefix:'[Agent mode: বড় কাজে create_plan দিয়ে plan বানাও, ধাপে ধাপে execute করো। বিশেষজ্ঞ কাজে delegate_to_agent দিয়ে sub-agent-কে দাও (team_tool list দেখায়)। Website লাগলে build_website। শেষে গুছিয়ে ফলাফল।] ',
    chips:['👥 আমার agent team দেখাও','🌐 আমার দোকানের website বানাও','🤖 নতুন sub-agent বানাও','📊 রিসার্চ করে report বানাও']
  }
};

/* ══════════ Theme ══════════ */
function setTheme(t) {
  document.documentElement.setAttribute('data-theme', t);
  localStorage.setItem('sabbir_theme', t);
  document.getElementById('iconSun').style.display  = (t === 'dark') ? 'block' : 'none';
  document.getElementById('iconMoon').style.display = (t === 'dark') ? 'none' : 'block';
}
function toggleTheme() {
  setTheme(document.documentElement.getAttribute('data-theme') === 'dark' ? 'light' : 'dark');
}
const savedTheme = localStorage.getItem('sabbir_theme');
const prefersDark = window.matchMedia && window.matchMedia('(prefers-color-scheme: dark)').matches;
setTheme(savedTheme || (prefersDark ? 'dark' : 'light'));

/* ══════════ Settings panel ══════════ */
function toggleSettings() {
  const p = document.getElementById('settingspanel');
  p.style.display = (p.style.display === 'flex') ? 'none' : 'flex';
}

/* ══════════ Brand ══════════ */
function applyBrand(b) {
  if (!b) return;
  BRAND = b;
  document.title = b.name;
  document.getElementById('brandname').textContent = b.name;
  document.getElementById('brandtag').textContent = b.tagline;
  const r = document.documentElement.style;
  if (b.color1) r.setProperty('--g1', b.color1);
  if (b.color2) r.setProperty('--g2', b.color2);
  if (b.color3) r.setProperty('--g3', b.color3);
  updateHero();
}

/* ══════════ Sidebar (ChatGPT-style drawer) ══════════ */
function toggleDrawer(force) {
  const sb = document.getElementById('sidebar');
  const sc = document.getElementById('scrim');
  const open = (force !== undefined) ? force : !sb.classList.contains('open');
  sb.classList.toggle('open', open);
  sc.classList.toggle('show', open);
}
function qa(kind) {
  const inp = document.getElementById('msginput');
  if (kind === 'image') { setMode('chat'); inp.value = 'একটা ছবি বানাও: [কীসের ছবি চান লিখুন]'; }
  else if (kind === 'write') { setMode('chat'); inp.value = 'লিখে দাও: [কী লিখতে হবে]'; }
  else if (kind === 'web') { setMode('research'); inp.value = ''; }
  else if (kind === 'biz') { setMode('business'); inp.value = ''; }
  inp.focus();
}
function buildSidebar() {
  const sb = document.getElementById('sidebar');
  sb.innerHTML = '';
  const labels = {dashboard: '🎛️ প্যানেল', chat: '💬 কথা ও কাজ'};
  Object.keys(MODES).forEach(function(key) {
    if (labels[key]) {
      const l = document.createElement('div');
      l.className = 'sidelabel';
      l.textContent = labels[key];
      sb.appendChild(l);
    }
    const m = MODES[key];
    const b = document.createElement('button');
    b.className = 'nav' + (key === MODE ? ' active' : '');
    b.id = 'nav-' + key;
    b.innerHTML = '<span class="ic">' + m.icon + '</span><span>' + m.label + '</span>';
    b.onclick = function() { setMode(key); };
    sb.appendChild(b);
  });
  const foot = document.createElement('div');
  foot.className = 'side-foot';
  foot.textContent = '⚡ 18 tools active';
  sb.appendChild(foot);
}
function setMode(key) {
  MODE = key;
  document.querySelectorAll('.nav').forEach(function(n) { n.classList.remove('active'); });
  const btn = document.getElementById('nav-' + key);
  if (btn) btn.classList.add('active');
  const isDash = (key === 'dashboard');
  const isCmd = (key === 'command');
  const isCtrl = (key === 'control');
  const isSkill = (key === 'skills');
  const isAdmin = (key === 'admin');
  const isPanel = isDash || isCmd || isCtrl || isSkill || isAdmin;
  document.getElementById('dashboard').style.display = isDash ? 'flex' : 'none';
  document.getElementById('command').style.display = isCmd ? 'flex' : 'none';
  document.getElementById('control').style.display = isCtrl ? 'flex' : 'none';
  document.getElementById('skills').style.display = isSkill ? 'flex' : 'none';
  document.getElementById('adminpanel').style.display = isAdmin ? 'flex' : 'none';
  document.getElementById('chat').style.display = isPanel ? 'none' : 'flex';
  document.getElementById('chips').style.display = isPanel ? 'none' : 'flex';
  document.getElementById('inputbar').style.display = isPanel ? 'none' : 'flex';
  if (isDash) { toggleDrawer(false); loadDashboard(); return; }
  if (isCmd) { toggleDrawer(false); loadCommand(); return; }
  if (isCtrl) { toggleDrawer(false); loadControl(); return; }
  if (isSkill) { toggleDrawer(false); buildSkills(); return; }
  if (isAdmin) { toggleDrawer(false); loadAdmin(); return; }
  document.getElementById('msginput').placeholder = MODES[key].placeholder;
  updateHero();
  showChips();
  toggleDrawer(false);
  document.getElementById('msginput').focus();
}

let lastWeekData = null;
async function loadCommand() {
  const grid = document.getElementById('cmdgrid');
  grid.innerHTML = '<div class="dempty">লোড হচ্ছে...</div>';
  try {
    const r = await fetch('api/command', {method:'POST', headers:authHeaders()});
    const j = await r.json();
    if (j.error) { grid.innerHTML = '<div class="dempty">❌ ' + j.error + '</div>'; return; }
    lastWeekData = j;
    document.getElementById('cmdtime').textContent = 'আপডেট: ' + (j.generated || '');
    grid.innerHTML = '';

    // Card 1: Weekly Overview (bars)
    const c1 = document.createElement('div');
    c1.className = 'dcard';
    c1.innerHTML = '<h3><span class="dic">📊</span>Weekly Overview</h3>';
    const maxH = Math.max(0.1, ...(j.overview || []).map(function(o){ return o.hours; }));
    (j.overview || []).forEach(function(o) {
      const row = document.createElement('div');
      row.className = 'statbar';
      row.innerHTML = '<span class="lbl">' + o.label + '</span>' +
        '<span class="bar"><span class="fill" style="width:' + Math.round(o.hours/maxH*100) + '%"></span></span>' +
        '<span class="val">' + o.hours + 'h</span>';
      c1.appendChild(row);
    });
    grid.appendChild(c1);

    // Card 2: Tasks completed
    const c2 = document.createElement('div');
    c2.className = 'dcard';
    c2.style.animationDelay = '.05s';
    c2.innerHTML = '<h3><span class="dic">✅</span>Tasks completed</h3>' +
      '<div style="font-size:34px;font-weight:800;padding:6px 2px;' +
      'background:linear-gradient(90deg,var(--g1),var(--g2),var(--g3));' +
      '-webkit-background-clip:text;background-clip:text;color:transparent">' + (j.tasks_completed || 0) + '</div>' +
      '<div class="dempty">গত ৭ দিনে সম্পন্ন কাজ ও automation ধাপ</div>';
    grid.appendChild(c2);

    // Card 3: Daily activity
    const c3 = document.createElement('div');
    c3.className = 'dcard';
    c3.style.animationDelay = '.1s';
    c3.innerHTML = '<h3><span class="dic">📅</span>Daily Activity</h3>';
    if (j.days && j.days.length) {
      j.days.forEach(function(l) {
        const d = document.createElement('div');
        d.className = 'dline';
        d.textContent = l;
        c3.appendChild(d);
      });
    } else {
      c3.innerHTML += '<div class="dempty">এখনো কোনো activity নেই</div>';
    }
    grid.appendChild(c3);

    // Priority card (🔴🟡🟢)
    const cp = document.createElement('div');
    cp.className = 'dcard';
    cp.style.animationDelay = '.12s';
    cp.innerHTML = '<h3><span class="dic">🚦</span>Priorities</h3>';
    const pr = j.priorities || {};
    (pr.urgent || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline prio-red'; d.textContent='🔴 '+t; cp.appendChild(d); });
    (pr.important || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline prio-yel'; d.textContent='🟡 '+t; cp.appendChild(d); });
    (pr.later || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline prio-grn'; d.textContent='🟢 '+t; cp.appendChild(d); });
    if (!(pr.urgent||[]).length && !(pr.important||[]).length && !(pr.later||[]).length)
      cp.innerHTML += '<div class="dempty">Tasks-এ 🔴/🟢 emoji দিয়ে priority দিন</div>';
    grid.appendChild(cp);

    // Upcoming + Emails + Alerts card
    const cu = document.createElement('div');
    cu.className = 'dcard';
    cu.style.animationDelay = '.14s';
    cu.innerHTML = '<h3><span class="dic">🔔</span>Upcoming & Alerts</h3>';
    (j.upcoming || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline'; d.textContent='📅 '+t; cu.appendChild(d); });
    (j.emails || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline'; d.textContent='📧 '+t; cu.appendChild(d); });
    (j.alerts || []).forEach(function(t) { const d=document.createElement('div'); d.className='dline'; d.textContent=t; cu.appendChild(d); });
    if (!(j.upcoming||[]).length && !(j.emails||[]).length && !(j.alerts||[]).length)
      cu.innerHTML += '<div class="dempty">সব শান্ত — কোনো alert নেই ✨</div>';
    grid.appendChild(cu);
    if ((j.alerts||[]).length) notifyUser('Sabbir Alert', j.alerts[0]);

    // Card 4: Top mode
    const c4 = document.createElement('div');
    c4.className = 'dcard';
    c4.style.animationDelay = '.15s';
    c4.innerHTML = '<h3><span class="dic">🏆</span>This Week</h3>';
    (j.highlights || []).forEach(function(l) {
      const d = document.createElement('div');
      d.className = 'dline';
      d.textContent = l;
      c4.appendChild(d);
    });
    if (!(j.highlights || []).length) c4.innerHTML += '<div class="dempty">Data জমলে এখানে highlights আসবে</div>';
    grid.appendChild(c4);
  } catch (e) { grid.innerHTML = '<div class="dempty">❌ ' + e + '</div>'; }
}

/* ══════════ 🔧 ADMIN PANEL ══════════ */
async function loadAdmin() {
  const grid = document.getElementById('admingrid');
  grid.innerHTML = '<div class="dempty">লোড হচ্ছে...</div>';
  try {
    const j = await api('api/admin', {action:'get'});
    if (j.error) { grid.innerHTML = '<div class="dempty">❌ ' + j.error + '</div>'; return; }
    grid.innerHTML = '';
    const b = j.brand || {};

    // ── 🎨 Branding ──
    const bc = card('Branding & Identity', '🎨', 0);
    bc.innerHTML += '<span class="alabel">নাম</span><input class="selfield" id="ad_name" value="' + esc(b.name||'') + '">' +
      '<span class="alabel">Tagline</span><input class="selfield" id="ad_tag" value="' + esc(b.tagline||'') + '">' +
      '<span class="alabel">Brand Colors</span>' +
      '<div class="colorrow"><input type="color" id="ad_c1" value="' + esc(b.color1||'#6366f1') + '">' +
      '<input type="color" id="ad_c2" value="' + esc(b.color2||'#8b5cf6') + '">' +
      '<input type="color" id="ad_c3" value="' + esc(b.color3||'#ec4899') + '">' +
      '<span style="font-size:11px;color:var(--dim)">৩টা gradient রঙ</span></div>' +
      '<span class="alabel">Welcome message</span><textarea class="selfield" id="ad_wel">' + esc(b.welcome||'') + '</textarea>' +
      '<span class="alabel">ভাবার সময়ের লেখা</span><input class="selfield" id="ad_think" value="' + esc(b.thinking_text||'') + '">';
    const brow = crow('<button class="mini" data-act="adminsavebrand">💾 Save</button>' +
      '<button class="mini danger" data-act="adminresetbrand">↺ Reset default</button>');
    bc.appendChild(brow);
    grid.appendChild(bc);

    // ── 🧠 Personality ──
    const pc = card('Personality & Style', '🧠', 0.05);
    pc.innerHTML += '<span class="alabel">Personality (চরিত্র)</span><textarea class="selfield" id="ad_pers">' + esc(b.personality||'') + '</textarea>' +
      '<span class="alabel">Tone (কথার ধরন)</span><textarea class="selfield" id="ad_tone">' + esc(b.tone||'') + '</textarea>' +
      '<span class="alabel">Response style</span><textarea class="selfield" id="ad_style">' + esc(b.response_style||'') + '</textarea>' +
      '<span class="alabel">Suggestion chips (প্রতি লাইনে একটা)</span><textarea class="selfield" id="ad_chips">' + esc((b.chips||[]).join(String.fromCharCode(10))) + '</textarea>';
    pc.appendChild(crow('<button class="mini" data-act="adminsavebrand">💾 Save</button>'));
    grid.appendChild(pc);

    // ── 🤖 AI Engine (multi-provider) ──
    const ec = card('AI Engine', '🤖', 0.1);
    (j.providers || []).forEach(function(p) {
      const d = document.createElement('div');
      d.className = 'provrow' + (p.active ? ' active' : '');
      let opts = '';
      p.models.forEach(function(mo) {
        opts += '<option value="' + esc(mo) + '"' + (p.active && j.engine.model === mo ? ' selected' : '') + '>' + esc(mo) + '</option>';
      });
      const delbtn = p.custom ? '<button class="mini danger" data-act="delprovider" data-i="' + esc(p.key) + '">🗑️</button>' : '';
      d.innerHTML = '<div style="display:flex;align-items:center;gap:8px">' +
        '<span class="ctext"><b>' + esc(p.label) + '</b>' + (p.custom ? ' <span class="badge on">CUSTOM</span>' : '') + '<br><span style="color:var(--dim);font-size:10.5px">' + esc(p.note) + '</span></span>' +
        '<span class="badge ' + (p.configured ? 'on' : 'off') + '">' + (p.configured ? 'KEY ✓' : 'KEY নেই') + '</span>' + delbtn + '</div>' +
        '<div style="display:flex;gap:6px;margin-top:6px">' +
        '<input class="minifield" style="flex:1" id="ad_model_' + esc(p.key) + '" value="' + esc(p.active ? j.engine.model : p.models[0]) + '" placeholder="model নাম" list="ml_' + esc(p.key) + '">' +
        '<datalist id="ml_' + esc(p.key) + '">' + opts + '</datalist>' +
        '<button class="mini" data-act="adminengine" data-i="' + esc(p.key) + '">' +
        (p.active ? '✓ চালু' : 'ব্যবহার করো') + '</button></div>';
      ec.appendChild(d);
    });
    // 🚀 ভবিষ্যতের AI যোগ করার form
    const addp = document.createElement('div');
    addp.className = 'provrow';
    addp.innerHTML = '<b style="font-size:12.5px">🚀 নতুন AI যোগ করুন (ভবিষ্যতের যেকোনো model!)</b>' +
      '<input class="selfield" id="np_label" placeholder="নাম — যেমন: ChatGPT-6">' +
      '<input class="selfield" id="np_base" placeholder="API URL — যেমন: https://api.openai.com/v1">' +
      '<input class="selfield" id="np_model" placeholder="Model নাম — যেমন: gpt-6">' +
      '<button class="mini" data-act="addprovider" style="margin-top:4px">➕ যোগ করো</button>';
    ec.appendChild(addp);
    ec.innerHTML += '<div class="dempty">OpenAI-compatible যেকোনো API চলবে (OpenAI/Groq/DeepSeek/Mistral/xAI/Together...)। যোগ করার পর নিচে API Keys-এ key বসান।</div>';
    grid.appendChild(ec);

    // ── 🔑 API Keys vault ──
    const kc = card('API Keys (নিরাপদ vault)', '🔑', 0.15);
    (j.providers || []).forEach(function(p) {
      kc.appendChild(crow('<span class="ctext">' + esc(p.label) + '</span>' +
        '<input class="minifield" type="password" id="ad_key_' + esc(p.env) + '" placeholder="' + (p.configured ? '••• সেট আছে (বদলাতে লিখুন)' : 'key বসান') + '">' +
        '<button class="mini" data-act="adminkey" data-i="' + esc(p.env) + '">💾</button>'));
    });
    (j.integrations || []).forEach(function(ig) {
      kc.appendChild(crow('<span class="ctext">' + esc(ig.label) + '</span>' +
        '<input class="minifield" type="password" id="ad_key_' + esc(ig.key ? '' : '') + esc(ig.env || '') + '" placeholder="' + (ig.configured ? '••• সেট আছে' : ig.note) + '">' +
        '<button class="mini" data-act="adminkey" data-i="' + esc(ig.env || '') + '">💾</button>'));
    });
    kc.innerHTML += '<div class="dempty">🔐 Key শুধু server-এ থাকে — UI-তে কখনো ফেরত আসে না। খালি রেখে 💾 দিলে key মুছে যায়।</div>';
    grid.appendChild(kc);

    // ── 🎨 Personalize (মোবাইলের মতো!) ──
    const pz = card('Personalize (থিম ও ছবি)', '🎨', 0.17);
    pz.innerHTML += '<span class="alabel">🎨 Ready থিম — এক চাপে বদলান</span>';
    const themes = [
      ['💜 Purple SaaS', '#7c3aed', '#a78bfa', '#6366f1'],
      ['💎 Gold Luxury', '#d4af37', '#f0d060', '#b8860b'],
      ['💚 Emerald', '#059669', '#34d399', '#10b981'],
      ['🌊 Ocean Blue', '#0284c7', '#38bdf8', '#6366f1'],
      ['🔥 Sunset', '#ea580c', '#f97316', '#ef4444'],
      ['🌸 Rose Pink', '#db2777', '#f472b6', '#a78bfa']
    ];
    const trow = document.createElement('div');
    trow.style.cssText = 'display:flex;flex-wrap:wrap;gap:7px;margin:6px 0';
    themes.forEach(function(t) {
      const b = document.createElement('button');
      b.className = 'mini';
      b.style.cssText = 'background:linear-gradient(135deg,' + t[1] + ',' + t[2] + ');color:#fff;border:none';
      b.textContent = t[0];
      b.onclick = async function() {
        await api('api/admin', {action:'save_brand', brand:{color1:t[1], color2:t[2], color3:t[3]}});
        const r = document.documentElement.style;
        r.setProperty('--g1', t[1]); r.setProperty('--g2', t[2]); r.setProperty('--g3', t[3]);
        notifyUser('🎨', t[0] + ' থিম চালু!');
      };
      trow.appendChild(b);
    });
    pz.appendChild(trow);
    pz.innerHTML += '<span class="alabel">🖼️ Profile ছবি / Logo বদলান</span>';
    pz.appendChild(crow('<input type="file" id="logoup" accept="image/*" style="display:none">' +
      '<span class="ctext" style="color:var(--dim)">Gallery থেকে ছবি বাছুন (২MB-এর কম)</span>' +
      '<button class="mini" data-act="uplogo">📤 Upload</button>'));
    grid.appendChild(pz);

    // ── ⏰ Reminders ──
    const rem = await api('api/reminders', {action:'list'});
    const rc = card('Reminders', '⏰', 0.19);
    (rem.reminders || []).forEach(function(r, i) {
      rc.appendChild(crow('<span class="ctext">🕐 <b>' + esc(r.when) + '</b><br>' + esc(r.text) + '</span>' +
        '<button class="mini danger" data-act="delrem" data-i="' + (i+1) + '">🗑️</button>'));
    });
    if (!(rem.reminders || []).length) rc.innerHTML += '<div class="dempty">কোনো reminder নেই — chat-এ বলুন: "কালকে ১টায় মনে করিয়ে দিও..."</div>';
    rc.innerHTML += '<div class="dempty">⚠️ Notification পেতে app-টা browser-এ খোলা থাকতে হবে</div>';
    grid.appendChild(rc);

    // ── ⚙️ System ──
    const sc = card('System Settings', '⚙️', 0.2);
    sc.innerHTML += '<span class="alabel">সর্বোচ্চ tool ধাপ (৩-২৫)</span>' +
      '<input class="selfield" type="number" min="3" max="25" id="ad_iters" value="' + esc(String((j.system||{}).max_iterations||15)) + '">';
    sc.appendChild(crow('<button class="mini" data-act="adminsystem">💾 Save</button>'));
    sc.innerHTML += '<div class="dempty">বেশি ধাপ = জটিল কাজ পারবে, কিন্তু কোটা বেশি খরচ হবে</div>';
    grid.appendChild(sc);
  } catch (e) { grid.innerHTML = '<div class="dempty">❌ ' + e + '</div>'; }
}

async function adminAction(act, el) {
  if (act === 'adminsavebrand') {
    const chips = (document.getElementById('ad_chips').value || '').split(String.fromCharCode(10))
      .map(function(s){ return s.trim(); }).filter(Boolean);
    const brand = {
      name: document.getElementById('ad_name').value,
      tagline: document.getElementById('ad_tag').value,
      color1: document.getElementById('ad_c1').value,
      color2: document.getElementById('ad_c2').value,
      color3: document.getElementById('ad_c3').value,
      welcome: document.getElementById('ad_wel').value,
      thinking_text: document.getElementById('ad_think').value,
      personality: document.getElementById('ad_pers').value,
      tone: document.getElementById('ad_tone').value,
      response_style: document.getElementById('ad_style').value,
      chips: chips
    };
    const j = await api('api/admin', {action:'save_brand', brand: brand});
    applyBrand(ui_brand_from(j)); loadAdmin();
    notifyUser('Admin', 'Branding save হয়েছে!');
  }
  else if (act === 'adminresetbrand') {
    if (!confirm('সব branding default-এ ফিরবে?')) return;
    const j = await api('api/admin', {action:'reset_brand'});
    applyBrand(ui_brand_from(j)); loadAdmin();
  }
  else if (act === 'adminengine') {
    const prov = el.getAttribute('data-i');
    const sel = document.getElementById('ad_model_' + prov);
    await api('api/admin', {action:'set_engine', provider: prov, model: sel ? sel.value : ''});
    loadAdmin();
  }
  else if (act === 'addprovider') {
    const label = document.getElementById('np_label').value.trim();
    const base = document.getElementById('np_base').value.trim();
    const model = document.getElementById('np_model').value.trim();
    if (!label || !base || !model) { alert('তিনটা ঘরই পূরণ করুন!'); return; }
    const j = await api('api/admin', {action:'add_provider', label:label, base:base, model:model});
    if (j.error) alert('❌ ' + j.error); else notifyUser('Admin', label + ' যোগ হয়েছে!');
    loadAdmin();
  }
  else if (act === 'delprovider') {
    if (!confirm('এই provider মুছবেন?')) return;
    await api('api/admin', {action:'remove_provider', key: el.getAttribute('data-i')});
    loadAdmin();
  }
  else if (act === 'uplogo') {
    const f = document.getElementById('logoup');
    f.onchange = function() {
      const file = f.files[0];
      if (!file) return;
      if (file.size > 2 * 1024 * 1024) { alert('ছবি 2MB-এর কম হতে হবে!'); return; }
      const rd = new FileReader();
      rd.onload = async function() {
        const j = await api('api/logo', {data: rd.result.split(',')[1]});
        if (j.ok) {
          document.querySelectorAll('img[src^="logo.png"], #logoimg').forEach(function(im) {
            im.src = 'logo.png?t=' + Date.now();
          });
          notifyUser('🖼️', 'নতুন ছবি বসে গেছে!');
          alert('✅ নতুন logo/ছবি সেট হয়েছে!');
        } else alert('❌ ' + (j.error || 'সমস্যা হলো'));
      };
      rd.readAsDataURL(file);
    };
    f.click();
  }
  else if (act === 'delrem') {
    await api('api/reminders', {action:'delete', number: parseInt(el.getAttribute('data-i'))});
    loadAdmin();
  }
  else if (act === 'adminkey') {
    const env = el.getAttribute('data-i');
    const inp = document.getElementById('ad_key_' + env);
    await api('api/admin', {action:'set_key', env: env, value: inp ? inp.value : ''});
    if (inp) inp.value = '';
    loadAdmin();
  }
  else if (act === 'adminsystem') {
    await api('api/admin', {action:'set_system', max_iterations: parseInt(document.getElementById('ad_iters').value || '15')});
    loadAdmin();
  }
}
function ui_brand_from(j) {
  const b = (j && j.brand) || {};
  return {name: b.name, tagline: b.tagline, welcome: b.welcome,
          thinking_text: b.thinking_text, chips: b.chips,
          color1: b.color1, color2: b.color2, color3: b.color3};
}

/* ══════════ 🧰 SKILLS LIBRARY ══════════ */
const SKILLS = [
 {cat:'✍️ লেখালেখি', icon:'✍️', items:[
  ['Article/Blog লেখা','[বিষয়] নিয়ে একটা আকর্ষণীয় article লেখো — heading, intro, মূল অংশ, উপসংহার সহ।'],
  ['Email/আবেদন','[কার কাছে], [কী বিষয়ে] একটা professional email/আবেদন লিখে দাও, subject সহ।'],
  ['CV/Resume','আমার CV বানাও। নাম: [নাম], পড়াশোনা: [ডিগ্রি], দক্ষতা: [দক্ষতা], অভিজ্ঞতা: [অভিজ্ঞতা]।'],
  ['Cover Letter','[পদের নাম] পদের জন্য cover letter লেখো। আমার যোগ্যতা: [যোগ্যতা]।'],
  ['Product Description','[পণ্যের নাম] এর আকর্ষণীয় product description লেখো — feature, উপকারিতা, call-to-action সহ।'],
  ['YouTube Script','[বিষয়] নিয়ে [X মিনিটের] YouTube video script লেখো — hook, মূল content, outro সহ।'],
  ['Story লেখা','[থিম] নিয়ে একটা ছোটগল্প লেখো।'],
  ['Rewrite/উন্নত করা','এই লেখাটা আরো সুন্দর করে rewrite করো: [লেখা]'],
  ['Grammar ঠিক করা','এই লেখার ভুল ঠিক করে দাও আর কী ভুল ছিল বুঝিয়ে দাও: [লেখা]']]},
 {cat:'📚 পড়াশোনা', icon:'📚', items:[
  ['Topic বুঝে নেওয়া','[topic] টা একদম সহজ ভাষায়, উদাহরণ দিয়ে ধাপে ধাপে বুঝিয়ে দাও।'],
  ['Quiz/MCQ','[topic] থেকে ৫টা MCQ বানাও। আমি উত্তর দিলে check করে ভুল বুঝিয়ে দেবে।'],
  ['Study Plan','[পরীক্ষা/লক্ষ্য] এর জন্য [X দিনের] study plan বানাও, প্রতিদিনের রুটিন সহ, tasks.txt-এ save করো।'],
  ['Chapter Summary','এই অধ্যায়ের সারাংশ ও মূল পয়েন্ট বানাও: [লেখা বা 📎 ফাইল upload করুন]'],
  ['Flashcard','[topic] এর ১০টা flashcard বানাও (প্রশ্ন-উত্তর জোড়া) আর notes-এ save করো।'],
  ['Practice Problem','[topic] এর ৫টা practice problem দাও, সমাধান ধাপে ধাপে।']]},
 {cat:'💼 ব্যবসা', icon:'💼', items:[
  ['Business Plan','[ব্যবসার ধারণা] এর জন্য সম্পূর্ণ business plan বানাও — টার্গেট customer, খরচ, দাম, marketing সহ।'],
  ['SWOT Analysis','আমার ব্যবসার SWOT analysis করো। ব্যবসা: [বর্ণনা]'],
  ['Pricing Strategy','[পণ্য] এর দাম কত রাখা উচিত? খরচ: [খরচ], বাজার research করে পরামর্শ দাও।'],
  ['লাভ-ক্ষতি হিসাব','এই মাসের হিসাব করো: আয় [টাকা], খরচ [টাকা] — বিশ্লেষণ ও পরামর্শ দাও।'],
  ['Competitor Analysis','[প্রতিযোগীর নাম/ধরন] নিয়ে web research করে তুলনামূলক বিশ্লেষণ দাও।'],
  ['Sales Script','[পণ্য] বিক্রির জন্য কার্যকর sales script লেখো — আপত্তি সামলানো সহ।']]},
 {cat:'📣 Marketing & Social', icon:'📣', items:[
  ['Facebook Ad Copy','[পণ্য/অফার] এর জন্য ৩টা ভিন্ন style-এর Facebook ad copy লেখো।'],
  ['Facebook Post','[বিষয়/অফার] নিয়ে engaging Facebook post লেখো — caption + hashtag সহ।'],
  ['Content Calendar','[ব্যবসা/পেজ] এর জন্য ৭ দিনের content calendar বানাও আর notes-এ save করো।'],
  ['SEO Keywords','[বিষয়/ব্যবসা] এর জন্য SEO keyword research করো (web search ব্যবহার করে)।'],
  ['Customer Persona','[ব্যবসা] এর ideal customer persona বানাও — বয়স, চাহিদা, সমস্যা, কোথায় পাবো।'],
  ['Reel/Short Idea','[বিষয়] নিয়ে ৫টা viral reel/short idea দাও, hook সহ।'],
  ['Campaign Idea','[উপলক্ষ/অফার] এর জন্য promotional campaign idea দাও — offer, লেখা, টাইমিং।']]},
 {cat:'💬 Customer কথাবার্তা', icon:'💬', items:[
  ['Complaint Reply','এই complaint-এর ভদ্র ও সমাধানমুখী উত্তর লেখো: [complaint]'],
  ['Follow-up Message','[customer/কাজ] এর জন্য নরম follow-up message লেখো।'],
  ['FAQ বানানো','আমার [ব্যবসা] এর জন্য ১০টা FAQ ও উত্তর বানাও, notes-এ save করো।'],
  ['Negotiation সাহায্য','[পরিস্থিতি] — এই দর কষাকষিতে কী বলবো? কৌশল দাও।'],
  ['Tone বদলানো','এই message টা [formal/friendly] tone-এ বদলে দাও: [message]']]},
 {cat:'🔎 Research', icon:'🔎', items:[
  ['Product Research','[পণ্য] কেনার আগে research করো — দাম, রিভিউ, বিকল্প, সেরা কোনটা।'],
  ['Price Research','[পণ্য] এর বর্তমান বাজার দাম web search করে জানাও।'],
  ['Market Research','[পণ্য/সেবা] এর বাজার কেমন? চাহিদা, প্রতিযোগিতা, সুযোগ research করে report দাও।'],
  ['Fact-check','এই তথ্যটা সত্যি কিনা একাধিক উৎস থেকে যাচাই করো: [তথ্য]'],
  ['News Summary','আজকের [বিষয়/দেশ] এর গুরুত্বপূর্ণ খবরগুলোর সারাংশ দাও।'],
  ['Research Report','[বিষয়] নিয়ে গভীর research করে উৎসসহ পূর্ণাঙ্গ report বানাও — plan করে ধাপে ধাপে।']]},
 {cat:'🌍 অনুবাদ ও ভাষা', icon:'🌍', items:[
  ['বাংলা → English','English-এ অনুবাদ করো: [লেখা]'],
  ['English → বাংলা','বাংলায় অনুবাদ করো: [লেখা]'],
  ['Document অনুবাদ','📎 বাটনে ফাইল upload করে বলুন: আমার document-টা [ভাষা]-এ অনুবাদ করো।'],
  ['English শেখা','আমাকে English শেখাও — আজ [tense/vocabulary/spoken] নিয়ে ছোট lesson + practice দাও।']]},
 {cat:'🧑‍💻 ব্যক্তিগত', icon:'🧑‍💻', items:[
  ['Daily Plan','আজকের জন্য একটা কার্যকর daily plan বানাও। আমার কাজ: [কাজের তালিকা]'],
  ['Travel Plan','[জায়গা] ভ্রমণের plan বানাও — খরচ, রুট, দর্শনীয় স্থান, checklist সহ।'],
  ['সিদ্ধান্ত সাহায্য','[সিদ্ধান্ত] নিয়ে ভাবছি — pros & cons বিশ্লেষণ করে পরামর্শ দাও।'],
  ['Shopping Research','[পণ্য] কিনতে চাই, budget [টাকা] — research করে সেরা option বলো।']]}
];

function buildSkills() {
  const grid = document.getElementById('skillgrid');
  if (grid.dataset.built) return;
  grid.dataset.built = '1';
  SKILLS.forEach(function(group, gi) {
    const c = document.createElement('div');
    c.className = 'dcard';
    c.style.animationDelay = (gi * 0.05) + 's';
    c.innerHTML = '<h3><span class="dic">' + group.icon + '</span>' + group.cat + '</h3>';
    group.items.forEach(function(it) {
      const b = document.createElement('button');
      b.className = 'skillbtn';
      b.textContent = it[0];
      b.onclick = function() {
        setMode('chat');
        hideHero();
        const inp = document.getElementById('msginput');
        inp.value = it[1];
        inp.focus();
        addMsg('🧰 Skill: ' + it[0] + ' — নিচের বক্সে [ব্র্যাকেট] অংশ বদলে পাঠান!', 'tool');
      };
      c.appendChild(b);
    });
    grid.appendChild(c);
  });
}

/* ══════════ 🛠️ CONTROL CENTER ══════════ */
function card(title, icon, delay) {
  const c = document.createElement('div');
  c.className = 'dcard';
  if (delay) c.style.animationDelay = delay + 's';
  c.innerHTML = '<h3><span class="dic">' + icon + '</span>' + title + '</h3>';
  return c;
}
function crow(html) {
  const d = document.createElement('div');
  d.className = 'crow';
  d.innerHTML = html;
  return d;
}
function esc(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

async function api(path, body) {
  const r = await fetch(path, {method:'POST', headers:authHeaders(), body: JSON.stringify(body || {})});
  return r.json();
}

async function loadControl() {
  const grid = document.getElementById('ctrlgrid');
  grid.innerHTML = '<div class="dempty">লোড হচ্ছে...</div>';
  try {
    const [mem, goals, perms, autos, logd, appr] = await Promise.all([
      api('api/memory'), api('api/goals'), api('api/perms'), api('api/autos'), api('api/log'),
      api('api/approvals')
    ]);
    grid.innerHTML = '';

    // ── 🧠 Memory Manager ──
    const mc = card('Memory Manager', '🧠', 0);
    (mem.facts || []).forEach(function(f, i) {
      const r = crow('<span class="ctext">' + esc(f.text) + '</span>' +
        '<button class="mini" data-act="editmem" data-i="' + i + '">✏️</button>' +
        '<button class="mini danger" data-act="delmem" data-i="' + i + '">🗑️</button>');
      mc.appendChild(r);
    });
    if (!(mem.facts || []).length) mc.innerHTML += '<div class="dempty">কোনো স্মৃতি নেই</div>';
    const addm = crow('<input class="minifield" id="newmem" placeholder="নতুন তথ্য লিখুন...">' +
      '<button class="mini" data-act="addmem">➕</button>');
    mc.appendChild(addm);
    if ((mem.facts || []).length) {
      const fa = crow('<span class="ctext" style="color:var(--dim)">সব স্মৃতি মুছুন</span>' +
        '<button class="mini danger" data-act="forgetall">🧹 Forget all</button>');
      mc.appendChild(fa);
    }
    grid.appendChild(mc);

    // ── 🎯 Goals ──
    const gc = card('Goals', '🎯', 0.05);
    (goals.goals || []).forEach(function(g, i) {
      const dl = g.deadline ? ' ⏰' + esc(g.deadline) : '';
      const r = crow('<span class="ctext">' + (g.done ? '✅ ' : '') + esc(g.title) + dl +
        '<div class="pbar"><div class="pfill" style="width:' + (g.progress||0) + '%"></div></div></span>' +
        '<button class="mini" data-act="goalup" data-i="' + (i+1) + '">+10%</button>' +
        '<button class="mini danger" data-act="goaldel" data-i="' + i + '">🗑️</button>');
      gc.appendChild(r);
    });
    if (!(goals.goals || []).length) gc.innerHTML += '<div class="dempty">কোনো লক্ষ্য নেই</div>';
    gc.appendChild(crow('<input class="minifield" id="newgoal" placeholder="নতুন লক্ষ্য...">' +
      '<button class="mini" data-act="addgoal">➕</button>'));
    grid.appendChild(gc);

    // ── 🛡️ Permissions ──
    const pc = card('Permissions', '🛡️', 0.1);
    const pdesc = {web:'🌐 Web access (search/browse)', code:'💻 Code execution',
                   files:'📁 File/notes access', memory:'🧠 Memory লেখা/মোছা',
                   business:'🛒 Business data (order/customer/টাকা)',
                   approval_mode:'✋ Approval mode (sensitive কাজে অনুমতি লাগবে)'};
    Object.keys(pdesc).forEach(function(k) {
      const on = (perms.perms || {})[k];
      pc.appendChild(crow('<span class="ctext">' + pdesc[k] + '</span>' +
        '<button class="toggle' + (on ? ' on' : '') + '" data-act="perm" data-k="' + k + '" data-v="' + (on?1:0) + '"></button>'));
    });
    pc.innerHTML += '<div class="dempty">বন্ধ করলে agent সেই কাজ করতে পারবে না ⛔</div>';
    grid.appendChild(pc);

    // ── 🔄 Automations ──
    const ac = card('Automations', '🔄', 0.15);
    (autos.autos || []).forEach(function(a, i) {
      ac.appendChild(crow('<span class="ctext">' + esc(a.name) +
        '<span style="color:var(--dim)"> (' + (a.runs||0) + ' বার' + (a.last_run ? ', শেষ: ' + esc(a.last_run) : '') + ')</span></span>' +
        '<button class="mini" data-act="autorun" data-i="' + i + '">▶️ Run</button>' +
        '<button class="mini danger" data-act="autodel" data-i="' + i + '">🗑️</button>'));
    });
    if (!(autos.autos || []).length) ac.innerHTML += '<div class="dempty">কোনো automation নেই</div>';
    ac.appendChild(crow('<input class="minifield" id="autoname" placeholder="নাম (যেমন: সকালের খবর)">'));
    ac.appendChild(crow('<input class="minifield" id="autoprompt" placeholder="কী করবে (যেমন: আজকের খবর search করে দাও)">' +
      '<button class="mini" data-act="addauto">➕</button>'));
    grid.appendChild(ac);

    // ── ✋ Pending Approvals ──
    if ((appr.approvals || []).length) {
      const apc = card('Pending Approvals', '✋', 0.18);
      (appr.approvals || []).forEach(function(a) {
        apc.appendChild(crow('<span class="ctext">' + esc(a.id) + ': <b>' + esc(a.tool) + '</b> ' +
          esc(JSON.stringify(a.args).slice(0, 60)) + '<br><span style="color:var(--dim)">' + esc(a.time) + '</span></span>' +
          '<button class="mini" data-act="approve" data-i="' + esc(a.id) + '">✅</button>' +
          '<button class="mini danger" data-act="reject" data-i="' + esc(a.id) + '">❌</button>'));
      });
      grid.appendChild(apc);
    }

    // ── 🔌 Integrations (ভবিষ্যৎ) ──
    const ic = card('Integrations (ভবিষ্যৎ)', '🔌', 0.22);
    const INTG = [['📨 Telegram Bot','ফ্রি — token দিলেই চলবে'],
                  ['📧 Gmail inbox','Google OAuth লাগবে'],
                  ['📊 Google Sheets','Service account লাগবে'],
                  ['💬 WhatsApp Business','Meta API (paid)'],
                  ['💳 Payment gateway','Merchant account লাগবে']];
    INTG.forEach(function(x) {
      ic.appendChild(crow('<span class="ctext">' + x[0] + '<br><span style="color:var(--dim);font-size:11px">' +
        x[1] + '</span></span><span class="mini" style="cursor:default">🔒 বন্ধ</span>'));
    });
    ic.innerHTML += '<div class="dempty">Secrets-এ key দিলে ভবিষ্যতে চালু করা যাবে — কোডে জায়গা প্রস্তুত</div>';
    grid.appendChild(ic);

    // ── 💾 Backup ──
    const bc = card('Backup & Export', '💾', 0.2);
    bc.appendChild(crow('<span class="ctext">সব memory, goals, notes, history এক ফাইলে</span>' +
      '<button class="mini" data-act="export">⬇️ Export</button>'));
    bc.appendChild(crow('<span class="ctext">Backup ফাইল থেকে ফিরিয়ে আনুন</span>' +
      '<button class="mini" data-act="import">⬆️ Import</button>' +
      '<input type="file" id="importfile" accept=".json" style="display:none">'));
    grid.appendChild(bc);

    // ── 📜 Activity Log ──
    const lc = card('Activity Log', '📜', 0.25);
    const kindIcon = {tool:'🔧', chat:'💬', memory:'🧠', goal:'🎯', auto:'🔄',
                      perm:'🛡️', backup:'💾', blocked:'⛔'};
    (logd.log || []).slice(0, 25).forEach(function(e) {
      const d = document.createElement('div');
      d.className = 'dline';
      d.textContent = e.t + ' ' + (kindIcon[e.kind] || '•') + ' ' + e.text;
      lc.appendChild(d);
    });
    if (!(logd.log || []).length) lc.innerHTML += '<div class="dempty">এখনো কোনো activity নেই</div>';
    grid.appendChild(lc);

    grid.onclick = ctrlClick;
  } catch (e) { grid.innerHTML = '<div class="dempty">❌ ' + e + '</div>'; }
}

async function ctrlClick(ev) {
  const b = ev.target.closest('[data-act]');
  if (!b) return;
  const act = b.getAttribute('data-act');
  const i = parseInt(b.getAttribute('data-i') || '-1');
  if (act === 'delmem') { await api('api/memory', {action:'delete', index:i}); loadControl(); }
  else if (act === 'editmem') {
    const facts = (await api('api/memory')).facts || [];
    const nt = prompt('নতুন লেখা:', facts[i] ? facts[i].text : '');
    if (nt !== null && nt.trim()) { await api('api/memory', {action:'edit', index:i, text:nt}); loadControl(); }
  }
  else if (act === 'addmem') {
    const v = document.getElementById('newmem').value.trim();
    if (v) { await api('api/memory', {action:'add', text:v}); loadControl(); }
  }
  else if (act === 'forgetall') {
    if (confirm('সত্যিই সব স্মৃতি মুছে ফেলবেন?')) { await api('api/memory', {action:'forget_all'}); loadControl(); }
  }
  else if (act === 'addgoal') {
    const v = document.getElementById('newgoal').value.trim();
    if (v) { await api('api/goals', {action:'add', title:v}); loadControl(); }
  }
  else if (act === 'goalup') {
    const goals = (await api('api/goals')).goals || [];
    const g = goals[i-1];
    if (g) { await api('api/goals', {action:'progress', number:i, progress:Math.min(100,(g.progress||0)+10)}); loadControl(); }
  }
  else if (act === 'goaldel') { await api('api/goals', {action:'delete', index:i}); loadControl(); }
  else if (act === 'perm') {
    const k = b.getAttribute('data-k'), v = b.getAttribute('data-v') === '1';
    await api('api/perms', {set:{key:k, value:!v}});
    loadControl();
  }
  else if (act === 'addauto') {
    const n = document.getElementById('autoname').value.trim();
    const p = document.getElementById('autoprompt').value.trim();
    if (n && p) { await api('api/autos', {action:'add', name:n, prompt:p}); loadControl(); }
  }
  else if (act === 'autodel') { await api('api/autos', {action:'delete', index:i}); loadControl(); }
  else if (act === 'autorun') {
    b.textContent = '⏳...';
    const j = await api('api/autos', {action:'run', index:i});
    setMode('chat'); hideHero();
    (j.tool_calls || []).forEach(function(t) { addMsg('🔧 ' + t, 'tool'); showImages(t); });
    if (j.reply) { addMsg(j.reply, 'bot'); notifyUser('Automation শেষ!', j.reply.slice(0, 80)); }
  }
  else if (act === 'approve' || act === 'reject') {
    const j = await api('api/approvals', {action:'decide', id: b.getAttribute('data-i'), approve: act === 'approve'});
    if (j.result) { setMode('chat'); hideHero(); addMsg(j.result, 'bot'); }
    return;
  }
  else if (act === 'export') {
    const j = await api('api/backup');
    const blob = new Blob([JSON.stringify(j.backup, null, 1)], {type:'application/json'});
    const a = document.createElement('a');
    a.href = URL.createObjectURL(blob);
    a.download = 'sabbir_agent_backup_' + new Date().toISOString().slice(0,10) + '.json';
    a.click();
  }
  else if (act === 'import') {
    const f = document.getElementById('importfile');
    f.onchange = function() {
      const file = f.files[0];
      if (!file) return;
      const rd = new FileReader();
      rd.onload = async function() {
        try {
          const data = JSON.parse(rd.result);
          const j = await api('api/backup', {action:'import', data:data});
          alert(j.ok ? '✅ Restore সফল!' : '❌ Restore ব্যর্থ');
          loadControl();
        } catch (e) { alert('❌ ফাইলটা সঠিক backup না'); }
      };
      rd.readAsText(file);
    };
    f.click();
  }
}

/* ══════════ 📷 IMAGE ATTACH (Vision) ══════════ */
let pendingImage = null;
function attachImage(file) {
  if (!file) return;
  if (file.size > 4 * 1024 * 1024) { addMsg('❌ ছবি 4MB-এর বেশি বড়। ছোট ছবি দিন।', 'err'); return; }
  const reader = new FileReader();
  reader.onload = function() {
    pendingImage = { data: reader.result.split(',')[1], mime: file.type || 'image/jpeg' };
    document.getElementById('previmg').src = reader.result;
    document.getElementById('imgpreview').style.display = 'flex';
    document.getElementById('imgbtn').classList.add('has-img');
    document.getElementById('msginput').placeholder = 'ছবি নিয়ে কী জানতে চান?';
    document.getElementById('msginput').focus();
  };
  reader.readAsDataURL(file);
  document.getElementById('imginput').value = '';
}
function clearImage() {
  pendingImage = null;
  document.getElementById('imgpreview').style.display = 'none';
  document.getElementById('imgbtn').classList.remove('has-img');
  document.getElementById('msginput').placeholder = MODES[MODE].placeholder || 'যা খুশি জিজ্ঞেস করুন...';
}

/* ══════════ 🔊 TEXT-TO-SPEECH (উত্তর শোনা) ══════════ */
let currentUtter = null;
function speakText(text, btn) {
  try {
    if (!('speechSynthesis' in window)) { addMsg('❌ এই ব্রাউজারে voice নেই।', 'err'); return; }
    if (currentUtter) { speechSynthesis.cancel(); currentUtter = null;
      document.querySelectorAll('.speakbtn.playing').forEach(function(b){ b.classList.remove('playing'); b.textContent='🔊 শুনুন'; });
      if (btn && btn._wasPlaying) { btn._wasPlaying = false; return; } }
    const clean = text.split('').filter(function(ch){ return '*#`_>[]'.indexOf(ch) === -1; }).join('').slice(0, 1200);
    const u = new SpeechSynthesisUtterance(clean);
    u.lang = /[\u0980-\u09FF]/.test(clean) ? 'bn-BD' : 'en-US';
    u.rate = 1;
    u.onend = function() { if (btn) { btn.classList.remove('playing'); btn.textContent = '🔊 শুনুন'; } currentUtter = null; };
    if (btn) { btn.classList.add('playing'); btn.textContent = '⏹ থামান'; btn._wasPlaying = true; }
    currentUtter = u;
    speechSynthesis.speak(u);
  } catch (e) {}
}

/* ══════════ 🔔 SMART NOTIFICATION ══════════ */
function notifyUser(title, body) {
  try {
    if (!('Notification' in window)) return;
    if (Notification.permission === 'granted') new Notification(title, {body: body});
    else if (Notification.permission !== 'denied') Notification.requestPermission();
  } catch (e) {}
}

/* ══════════ 🎙️ VOICE COMMAND ══════════ */
let recog = null;
function toggleVoice() {
  const btn = document.getElementById('micbtn');
  const SR = window.SpeechRecognition || window.webkitSpeechRecognition;
  if (!SR) { addMsg('❌ এই ব্রাউজারে voice সাপোর্ট নেই। Chrome ব্যবহার করুন।', 'err'); return; }
  if (recog) { recog.stop(); return; }
  recog = new SR();
  recog.lang = 'bn-BD';
  recog.interimResults = false;
  btn.classList.add('listening');
  recog.onresult = function(e) {
    const text = e.results[0][0].transcript;
    document.getElementById('msginput').value = text;
    send();
  };
  recog.onend = function() { btn.classList.remove('listening'); recog = null; };
  recog.onerror = function() { btn.classList.remove('listening'); recog = null; };
  recog.start();
}

function analyzeWeek() {
  let summary = 'কোনো data নেই';
  if (lastWeekData) {
    const parts = (lastWeekData.overview || []).map(function(o){ return o.label + ': ' + o.hours + 'h'; });
    summary = parts.join(', ') + ' | Tasks completed: ' + (lastWeekData.tasks_completed || 0) +
              ' | ' + (lastWeekData.days || []).join(' | ');
  }
  setMode('chat');
  hideHero();
  const msg = 'আমার এই সপ্তাহের activity বিশ্লেষণ করো। Data: ' + summary +
    ' — কোথায় ভালো করছি, কোথায় সময় বাড়ানো উচিত, পরের সপ্তাহের জন্য ৩টা পরামর্শ দাও।';
  document.getElementById('msginput').value = msg;
  send();
}

async function loadDashboard() {
  const grid = document.getElementById('dashgrid');
  grid.innerHTML = '<div class="dempty">লোড হচ্ছে...</div>';
  try {
    const r = await fetch('api/dashboard', {method:'POST', headers:authHeaders()});
    const j = await r.json();
    if (j.error) { grid.innerHTML = '<div class="dempty">❌ ' + j.error + '</div>'; return; }
    document.getElementById('dashtime').textContent = 'আপডেট: ' + (j.generated || '');
    grid.innerHTML = '';
    (j.cards || []).forEach(function(cd, i) {
      const card = document.createElement('div');
      card.className = 'dcard';
      card.style.animationDelay = (i * 0.05) + 's';
      const h3 = document.createElement('h3');
      const ic = document.createElement('span');
      ic.className = 'dic';
      ic.textContent = cd.icon;
      h3.appendChild(ic);
      h3.appendChild(document.createTextNode(cd.title));
      card.appendChild(h3);
      if (cd.lines && cd.lines.length) {
        cd.lines.forEach(function(l) {
          const d = document.createElement('div');
          d.className = 'dline';
          d.textContent = l;
          card.appendChild(d);
        });
      } else {
        const d = document.createElement('div');
        d.className = 'dempty';
        d.textContent = cd.empty || 'কিছু নেই';
        card.appendChild(d);
      }
      grid.appendChild(card);
    });
  } catch (e) { grid.innerHTML = '<div class="dempty">❌ ' + e + '</div>'; }
}

/* ══════════ Hero (welcome) ══════════ */
function timeGreet() {
  const h = new Date().getHours();
  if (h < 5)  return 'শুভ রাত্রি 🌙';
  if (h < 12) return 'শুভ সকাল ☀️';
  if (h < 16) return 'শুভ দুপুর 👋';
  if (h < 18) return 'শুভ বিকেল 🌤️';
  return 'শুভ সন্ধ্যা ✨';
}
function updateHero() {
  const m = MODES[MODE];
  document.getElementById('herogreet').textContent = m.greet || timeGreet();
  document.getElementById('herosub').textContent =
    (MODE === 'chat' && BRAND.welcome) ? BRAND.welcome : m.sub;
}
function hideHero() {
  const h = document.getElementById('hero');
  if (h) h.style.display = 'none';
  hasMessages = true;
}
function showHero() {
  const h = document.getElementById('hero');
  if (h) h.style.display = 'flex';
  hasMessages = false;
  updateHero();
}

/* ══════════ Messages ══════════ */
function renderMd(text) {
  // regex-free markdown → HTML (Python-embedding-safe)
  const NL = String.fromCharCode(10);
  let s = esc(text);
  // code blocks (```)
  const parts = s.split('```');
  if (parts.length > 2) {
    let out = '';
    for (let i = 0; i < parts.length; i++) {
      if (i % 2 === 1) {
        out += '<pre style="background:rgba(0,0,0,.25);border:1px solid var(--bot-stroke);border-radius:10px;padding:10px;overflow-x:auto;font-size:12.5px;margin:6px 0">' + parts[i].trim() + '</pre>';
      } else out += parts[i];
    }
    s = out;
  }
  // inline code (`)
  const ic = s.split('`');
  if (ic.length > 2) {
    let out = '';
    for (let i = 0; i < ic.length; i++) {
      if (i % 2 === 1 && ic[i].indexOf(NL) === -1) {
        out += '<code style="background:rgba(0,0,0,.22);padding:1px 6px;border-radius:6px;font-size:.92em">' + ic[i] + '</code>';
      } else out += (i % 2 === 1 ? '`' + ic[i] + '`' : ic[i]);
    }
    s = out;
  }
  // bold (**)
  const bp = s.split('**');
  if (bp.length > 2) {
    let out = '';
    for (let i = 0; i < bp.length; i++) {
      out += (i % 2 === 1 && i < bp.length - (bp.length % 2 === 0 ? 1 : 0)) ? '<b>' + bp[i] + '</b>' : bp[i];
    }
    s = out;
  }
  // line-ভিত্তিক: heading, list, hr
  const lines = s.split(NL);
  const res = [];
  for (let i = 0; i < lines.length; i++) {
    let L = lines[i];
    const t = L.trim();
    if (t.indexOf('### ') === 0) res.push('<div style="font-weight:800;font-size:1.05em;margin:8px 0 3px;color:var(--g2)">' + t.slice(4) + '</div>');
    else if (t.indexOf('## ') === 0) res.push('<div style="font-weight:800;font-size:1.1em;margin:9px 0 3px;color:var(--g2)">' + t.slice(3) + '</div>');
    else if (t.indexOf('# ') === 0) res.push('<div style="font-weight:800;font-size:1.15em;margin:10px 0 4px;color:var(--g2)">' + t.slice(2) + '</div>');
    else if (t.indexOf('- ') === 0 || t.indexOf('• ') === 0) res.push('<div style="padding-left:14px">• ' + t.slice(2) + '</div>');
    else if (t.indexOf('* ') === 0) res.push('<div style="padding-left:14px">• ' + t.slice(2) + '</div>');
    else if (t === '---' || t === '----' || t === '-----') res.push('<hr style="border:none;border-top:1px solid var(--bot-stroke);margin:8px 0">');
    else if (t === '') res.push('<div style="height:6px"></div>');
    else res.push('<div>' + L + '</div>');
  }
  return res.join('');
}
function avatarHTML() {
  return '<img src="logo.png" onerror="this.parentNode.textContent=String.fromCodePoint(0x1F916)">';
}
function addMsg(text, cls) {
  hideHeroIfNeeded(cls);
  if (cls === 'bot') {
    const row = document.createElement('div');
    row.className = 'bot-row';
    const av = document.createElement('span');
    av.className = 'avatar';
    av.innerHTML = avatarHTML();
    const d = document.createElement('div');
    d.className = 'msg bot';
    d.innerHTML = renderMd(text);
    if (text.length > 20 && 'speechSynthesis' in window) {
      const sp = document.createElement('button');
      sp.className = 'speakbtn';
      sp.textContent = '🔊 শুনুন';
      sp.onclick = function() { speakText(text, sp); };
      d.appendChild(document.createElement('br'));
      d.appendChild(sp);
    }
    row.appendChild(av); row.appendChild(d);
    chat.appendChild(row);
    chat.scrollTop = chat.scrollHeight;
    return row;
  }
  const d = document.createElement('div');
  d.className = 'msg ' + cls;
  d.textContent = text;
  chat.appendChild(d);
  chat.scrollTop = chat.scrollHeight;
  return d;
}
function hideHeroIfNeeded(cls) {
  if (cls === 'user' || cls === 'tool') hideHero();
}
function addTyping() {
  hideHero();
  const d = document.createElement('div');
  d.className = 'bot-row';
  d.innerHTML = '<span class="avatar think-ring">' + avatarHTML() + '</span>' +
    '<div class="msg bot typing-row"><span class="tlabel">' + (BRAND.thinking_text || 'ভাবছি') +
    '</span><span class="dots"><i></i><i></i><i></i></span></div>';
  chat.appendChild(d);
  chat.scrollTop = chat.scrollHeight;
  return d;
}
function showLinks(t) {
  const idx = t.indexOf('.html');
  if (idx === -1) return;
  let start = idx;
  while (start > 0 && 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_-'.indexOf(t.charAt(start-1)) !== -1) start--;
  const fname = t.slice(start, idx) + '.html';
  if (!fname || fname === '.html') return;
  {
    const a = document.createElement('a');
    a.href = 'sandbox/' + fname + '?code=' + encodeURIComponent(accessCode);
    a.target = '_blank';
    a.className = 'msg tool';
    a.style.textDecoration = 'none';
    a.textContent = '🌐 ' + fname + ' — খুলতে ক্লিক করুন ↗';
    chat.appendChild(a);
    chat.scrollTop = chat.scrollHeight;
  }
}
function showImages(t) {
  const marker = 'তৈরি হয়েছে: ';
  const idx = t.indexOf(marker);
  if (idx === -1) return;
  let rest = t.slice(idx + marker.length);
  const paren = rest.indexOf(' (');
  if (paren !== -1) rest = rest.slice(0, paren);
  rest.split(',').forEach(function(s) {
    const name = s.trim();
    if (name.endsWith('.png') || name.endsWith('.jpg') || name.endsWith('.jpeg')) {
      const img = document.createElement('img');
      img.className = 'chatimg';
      img.src = 'sandbox/' + name + '?t=' + Date.now() + '&code=' + encodeURIComponent(accessCode);
      chat.appendChild(img);
      chat.scrollTop = chat.scrollHeight;
    }
  });
}

/* ══════════ Chips ══════════ */
function showChips() {
  const box = document.getElementById('chips');
  box.innerHTML = '';
  const list = MODES[MODE].chips || BRAND.chips || [];
  list.forEach(function(txt, i) {
    const b = document.createElement('button');
    b.className = 'chip';
    b.style.animationDelay = (i * 0.06) + 's';
    b.textContent = txt;
    b.onclick = function() {
      document.getElementById('msginput').value = txt.slice(txt.indexOf(' ') + 1);
      send();
    };
    box.appendChild(b);
  });
}

/* ══════════ Auth + API ══════════ */
function authHeaders() {
  return {'Content-Type':'application/json', 'X-Access-Code': accessCode};
}
function saveCode() {
  accessCode = document.getElementById('codeinput').value.trim();
  localStorage.setItem('sabbir_access_code', accessCode);
  document.getElementById('settingspanel').style.display = 'none';
  addMsg('🔓 Access Code সংরক্ষণ হয়েছে। এবার chat করুন!', 'bot');
}
async function saveKey() {
  const key = document.getElementById('keyinput').value.trim();
  if (!key) return;
  const r = await fetch('api/key', { method:'POST', headers: authHeaders(), body: JSON.stringify({key}) });
  const j = await r.json();
  if (j.ok) {
    document.getElementById('settingspanel').style.display = 'none';
    addMsg('✅ API key সংরক্ষণ হয়েছে! এখন chat শুরু করুন।', 'bot');
    checkStatus();
  } else {
    addMsg('Key সংরক্ষণে সমস্যা: ' + (j.error || ''), 'err');
    if (j.auth === false) document.getElementById('settingspanel').style.display = 'flex';
  }
}
async function checkStatus() {
  try {
    const r = await fetch('api/status');
    const j = await r.json();
    const st = document.getElementById('status');
    const sttext = document.getElementById('statustext');
    applyBrand(j.brand);
    if (j.needs_code && !accessCode) {
      document.getElementById('settingspanel').style.display = 'flex';
      addMsg('🔐 এই agent পাসওয়ার্ড-সুরক্ষিত। উপরে আপনার Access Code দিন।', 'bot');
    }
    if (j.has_key) {
      sttext.textContent = 'Online';
      st.className = 'status ok';
      showChips();
    } else {
      sttext.textContent = 'API key দিন';
      document.getElementById('settingspanel').style.display = 'flex';
      addMsg('স্বাগতম! 👋 শুরু করতে Settings-এ (উপরের ⚙️) আপনার ফ্রি Gemini API key দিন। ফ্রি key: aistudio.google.com/apikey', 'bot');
    }
  } catch (e) {
    document.getElementById('statustext').textContent = 'সংযোগ সমস্যা';
  }
}

/* ══════════ Send / Upload / Reset ══════════ */
async function send() {
  const inp = document.getElementById('msginput');
  const text = inp.value.trim();
  if (!text) return;
  inp.value = '';
  addMsg(text, 'user');
  const btn = document.getElementById('sendbtn');
  btn.disabled = true;
  const typing = addTyping();
  const payload = MODES[MODE].prefix ? (MODES[MODE].prefix + text) : text;
  const body = {message: payload, mode: MODE};
  if (pendingImage) { body.image = pendingImage.data; body.image_mime = pendingImage.mime; clearImage(); }
  try {
    const r = await fetch('api/chat', { method:'POST', headers: authHeaders(),
                                        body: JSON.stringify(body) });
    const j = await r.json();
    typing.remove();
    if (j.tool_calls && j.tool_calls.length) {
      const names = j.tool_calls.map(function(t) { return t.split('(')[0]; });
      addMsg('📋 Actions: ' + names.join(' → '), 'tool');
      for (const t of j.tool_calls) { addMsg('🔧 ' + t, 'tool'); showImages(t); showLinks(t); }
    }
    if (j.error) {
      addMsg('❌ ' + j.error, 'err');
      if (j.auth === false) document.getElementById('settingspanel').style.display = 'flex';
    }
    else addMsg(j.reply, 'bot');
  } catch (e) {
    typing.remove();
    addMsg('❌ Network error: ' + e, 'err');
  }
  btn.disabled = false;
  inp.focus();
}
async function uploadFile(file) {
  if (!file) return;
  if (file.size > 20 * 1024 * 1024) { addMsg('❌ ফাইল 20MB-এর বেশি বড়।', 'err'); return; }
  const note = addTyping();
  const reader = new FileReader();
  reader.onload = async () => {
    const base64 = reader.result.split(',')[1];
    try {
      const r = await fetch('api/upload', { method:'POST', headers: authHeaders(),
        body: JSON.stringify({filename: file.name, data: base64}) });
      const j = await r.json();
      note.remove();
      if (j.error) {
        addMsg('❌ ' + j.error, 'err');
        if (j.auth === false) document.getElementById('settingspanel').style.display = 'flex';
      }
      else addMsg('📚 ' + j.message, 'tool');
    } catch (e) { note.remove(); addMsg('❌ Upload error: ' + e, 'err'); }
  };
  reader.readAsDataURL(file);
  document.getElementById('fileinput').value = '';
}
async function resetChat() {
  await fetch('api/reset', {method:'POST', headers: authHeaders()});
  document.querySelectorAll('#chat .msg, #chat .bot-row, #chat .chatimg').forEach(function(el) { el.remove(); });
  if (['dashboard','command','control','skills','admin'].includes(MODE)) setMode('chat');
  showHero();
  showChips();
}

/* ══════════ Init ══════════ */
buildSidebar();
updateHero();
checkStatus();

/* ══════════ ⏰ REMINDER POLLING ══════════ */
function beep() {
  try {
    const ctx = new (window.AudioContext || window.webkitAudioContext)();
    const o = ctx.createOscillator(), g = ctx.createGain();
    o.connect(g); g.connect(ctx.destination);
    o.frequency.value = 880; g.gain.value = .12;
    o.start(); o.stop(ctx.currentTime + .45);
  } catch (e) {}
}
async function pollReminders() {
  try {
    const j = await api('api/reminders', {action: 'due'});
    (j.due || []).forEach(function(r) {
      notifyUser('⏰ Reminder!', r.text);
      beep();
      if (navigator.vibrate) navigator.vibrate([250, 100, 250]);
      hideHero();
      addMsg('⏰ 🔔 Reminder: ' + r.text + ' (সময়: ' + r.when + ')', 'bot');
    });
  } catch (e) {}
}
setInterval(pollReminders, 30000);
setTimeout(pollReminders, 4000);
if ('Notification' in window && Notification.permission === 'default') {
  setTimeout(function() { Notification.requestPermission(); }, 6000);
}
</script>
</body>
</html>"""

agent = Agent()


class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, content_type="application/json"):
        data = body.encode("utf-8") if isinstance(body, str) else body
        self.send_response(code)
        self.send_header("Content-Type", content_type + "; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(length) or b"{}")

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            self._send(200, HTML_PAGE, "text/html")
        elif self.path.startswith("/manifest.json"):
            self._send(200, json.dumps({
                "name": BRAND["name"], "short_name": BRAND["name"][:12],
                "start_url": "/", "display": "standalone",
                "background_color": "#0d0521", "theme_color": "#7c3aed",
                "icons": [{"src": "/logo.png", "sizes": "512x512", "type": "image/png"}],
            }), "application/manifest+json")
        elif self.path.startswith("/logo.png"):
            logo = Path(__file__).parent / "logo.png"
            if logo.exists():
                self._send(200, logo.read_bytes(), "image/png")
            else:
                self._send(404, json.dumps({"error": "no logo"}))
        elif self.path.startswith("/sandbox/"):
            if not self._check_access():
                # ব্রাউজার <img> ট্যাগে header পাঠাতে পারে না, তাই query দিয়েও চলবে
                from urllib.parse import parse_qs, urlparse
                q = parse_qs(urlparse(self.path).query)
                if q.get("code", [""])[0] != ACCESS_CODE:
                    self._send(401, json.dumps({"error": "unauthorized"}))
                    return
            name = Path(self.path.split("/sandbox/", 1)[1].split("?")[0]).name
            fpath = SANDBOX / name
            if fpath.exists() and fpath.is_file():
                ctype = ("image/png" if name.endswith(".png")
                         else "image/jpeg" if name.endswith((".jpg", ".jpeg"))
                         else "text/html" if name.endswith(".html")
                         else "application/zip" if name.endswith(".zip")
                         else "text/plain")
                self._send(200, fpath.read_bytes(), ctype)
            else:
                self._send(404, json.dumps({"error": "file not found"}))
        elif self.path.endswith("api/status"):
            self._send(200, json.dumps({"has_key": bool(GEMINI_API_KEY),
                                        "model": MODEL_NAME,
                                        "needs_code": bool(ACCESS_CODE),
                                        "brand": ui_brand()}, ensure_ascii=False))
        else:
            self._send(404, json.dumps({"error": "not found"}))

    def _check_access(self) -> bool:
        """পাসওয়ার্ড চেক — timing-safe compare + brute-force lockout।"""
        if not ACCESS_CODE:
            return True
        ip = self.client_address[0] if self.client_address else "?"
        now = time.time()
        attempts = [t for t in FAILED_ATTEMPTS.get(ip, []) if now - t < 300]
        FAILED_ATTEMPTS[ip] = attempts
        if len(attempts) >= 10:
            return False  # ৫ মিনিটে ১০ বার ভুল → সাময়িক block
        ok = hmac.compare_digest(self.headers.get("X-Access-Code", ""), ACCESS_CODE)
        if not ok:
            FAILED_ATTEMPTS[ip] = attempts + [now]
        return ok

    def do_POST(self):
        try:
            if not self._check_access():
                self._send(401, json.dumps({"error": "ভুল পাসওয়ার্ড! সঠিক Access Code দিন।", "auth": False},
                                           ensure_ascii=False))
                return
            if self.path.endswith("api/key"):
                body = self._json_body()
                key = body.get("key", "").strip()
                if not key:
                    self._send(400, json.dumps({"ok": False, "error": "খালি key"}))
                    return
                set_api_key(key)
                self._send(200, json.dumps({"ok": True}))

            elif self.path.endswith("api/chat"):
                body = self._json_body()
                message = body.get("message", "").strip()
                if not message:
                    self._send(400, json.dumps({"error": "খালি message"}))
                    return
                try:
                    track_activity(body.get("mode", "chat"))
                except Exception:
                    pass
                tool_log = []
                image = None
                if body.get("image"):
                    img_data = body["image"]
                    if len(img_data) > 6_000_000:
                        self._send(400, json.dumps({"error": "ছবি খুব বড় (সর্বোচ্চ ~4MB)"}, ensure_ascii=False))
                        return
                    image = {"data": img_data, "mime": body.get("image_mime", "image/jpeg")}
                    log_activity("vision", f"ছবিসহ প্রশ্ন: {message[:60]}")
                reply = agent.run(message, verbose=False, tool_log=tool_log, image=image)
                log_activity("chat", message[:80])
                self._send(200, json.dumps({"reply": reply, "tool_calls": tool_log},
                                           ensure_ascii=False))

            elif self.path.endswith("api/upload"):
                body = self._json_body()
                filename = Path(body.get("filename", "")).name
                data = body.get("data", "")
                if not filename or not data:
                    self._send(400, json.dumps({"error": "ফাইল বা নাম নেই"}))
                    return
                ext = Path(filename).suffix.lower()
                if ext not in SUPPORTED:
                    self._send(400, json.dumps(
                        {"error": f"'{ext}' সাপোর্টেড না। দিন: {', '.join(sorted(SUPPORTED))}"},
                        ensure_ascii=False))
                    return
                (DOCS_DIR / filename).write_bytes(base64.b64decode(data))
                result = build_index()
                self._send(200, json.dumps(
                    {"message": f"'{filename}' যোগ হলো!\n{result}\nএখন এই ফাইল নিয়ে প্রশ্ন করুন।"},
                    ensure_ascii=False))

            elif self.path.endswith("api/command"):
                data = command_data()
                data.update(command_extra())
                self._send(200, json.dumps(data, ensure_ascii=False))

            elif self.path.endswith("api/dashboard"):
                self._send(200, json.dumps(dashboard_data(), ensure_ascii=False))

            elif self.path.endswith("api/memory"):
                # 🧠 Personal Memory System: list / add / edit / delete
                body = self._json_body()
                action = body.get("action", "list")
                facts = load_facts()
                if action == "add" and body.get("text", "").strip():
                    add_fact(body["text"].strip())
                    log_activity("memory", f"যোগ: {body['text'][:50]}")
                elif action == "edit":
                    i = int(body.get("index", -1))
                    if 0 <= i < len(facts) and body.get("text", "").strip():
                        old = facts[i]["text"]
                        facts[i]["text"] = body["text"].strip()
                        save_facts(facts)
                        log_activity("memory", f"সম্পাদনা: {old[:30]} → {body['text'][:30]}")
                elif action == "delete":
                    i = int(body.get("index", -1))
                    if 0 <= i < len(facts):
                        removed = facts.pop(i)
                        save_facts(facts)
                        log_activity("memory", f"মুছে ফেলা: {removed['text'][:50]}")
                elif action == "forget_all":
                    save_facts([])
                    log_activity("memory", "সব স্মৃতি মুছে ফেলা হলো")
                self._send(200, json.dumps({"facts": load_facts()}, ensure_ascii=False))

            elif self.path.endswith("api/goals"):
                # 🎯 Goal System
                body = self._json_body()
                action = body.get("action", "list")
                if action == "add" and body.get("title", "").strip():
                    goal_add(body["title"], body.get("deadline", ""))
                elif action == "progress":
                    goal_update(int(body.get("number", 0)), progress=int(body.get("progress", 0)))
                elif action == "delete":
                    goals = load_goals()
                    i = int(body.get("index", -1))
                    if 0 <= i < len(goals):
                        g = goals.pop(i)
                        save_goals(goals)
                        log_activity("goal", f"লক্ষ্য মুছে ফেলা: {g['title'][:40]}")
                self._send(200, json.dumps({"goals": load_goals()}, ensure_ascii=False))

            elif self.path.endswith("api/perms"):
                # 🛡️ Permission System
                body = self._json_body()
                if "set" in body:
                    set_perm(body["set"].get("key", ""), body["set"].get("value", True))
                self._send(200, json.dumps({"perms": get_perms()}, ensure_ascii=False))

            elif self.path.endswith("api/log"):
                # 📜 Activity Log
                self._send(200, json.dumps({"log": load_log()[-80:][::-1]}, ensure_ascii=False))

            elif self.path.endswith("api/autos"):
                # 🔄 Automation Builder
                body = self._json_body()
                action = body.get("action", "list")
                autos = load_autos()
                if action == "add" and body.get("name", "").strip() and body.get("prompt", "").strip():
                    autos.append({"name": body["name"].strip(), "prompt": body["prompt"].strip(),
                                  "created": datetime.date.today().isoformat(), "runs": 0})
                    save_autos(autos)
                    log_activity("auto", f"নতুন automation: {body['name'][:40]}")
                elif action == "delete":
                    i = int(body.get("index", -1))
                    if 0 <= i < len(autos):
                        a = autos.pop(i)
                        save_autos(autos)
                        log_activity("auto", f"automation মুছে ফেলা: {a['name'][:40]}")
                elif action == "run":
                    i = int(body.get("index", -1))
                    if 0 <= i < len(autos):
                        autos[i]["runs"] = autos[i].get("runs", 0) + 1
                        autos[i]["last_run"] = datetime.datetime.now().strftime("%d/%m %H:%M")
                        save_autos(autos)
                        log_activity("auto", f"চালানো হলো: {autos[i]['name'][:40]}")
                        tool_log = []
                        reply = agent.run(autos[i]["prompt"], verbose=False, tool_log=tool_log)
                        self._send(200, json.dumps({"autos": autos, "reply": reply,
                                                    "tool_calls": tool_log}, ensure_ascii=False))
                        return
                self._send(200, json.dumps({"autos": load_autos()}, ensure_ascii=False))

            elif self.path.endswith("api/backup"):
                # 💾 Backup & Export / Import & Restore
                body = self._json_body()
                if body.get("action") == "import" and body.get("data"):
                    ok = restore_bundle(body["data"])
                    agent.history = load_history()
                    self._send(200, json.dumps({"ok": ok}, ensure_ascii=False))
                else:
                    self._send(200, json.dumps({"backup": backup_bundle()}, ensure_ascii=False))

            elif self.path.endswith("api/approvals"):
                body = self._json_body()
                if body.get("action") == "decide":
                    r = run_approval(body.get("id", ""), body.get("approve", True))
                    self._send(200, json.dumps({"approvals": [a for a in _approvals_load()
                                                              if a["status"] == "pending"],
                                                **r}, ensure_ascii=False))
                    return
                self._send(200, json.dumps({"approvals": [a for a in _approvals_load()
                                                          if a["status"] == "pending"]},
                                           ensure_ascii=False))

            elif self.path.endswith("api/admin"):
                body = self._json_body()
                action = body.get("action", "get")
                if action == "save_brand":
                    admin_save_brand(body.get("brand", {}))
                elif action == "reset_brand":
                    admin_reset_brand()
                elif action == "set_engine":
                    admin_set_engine(body.get("provider", "gemini"), body.get("model", ""))
                elif action == "set_system":
                    admin_set_system(body.get("max_iterations"))
                elif action == "add_provider":
                    ok = admin_add_provider(body.get("label", ""), body.get("base", ""), body.get("model", ""))
                    if not ok:
                        self._send(400, json.dumps({"error": "নাম, সঠিক URL (https://...) আর model নাম — তিনটাই দরকার"}, ensure_ascii=False))
                        return
                elif action == "remove_provider":
                    admin_remove_provider(body.get("key", ""))
                elif action == "set_key":
                    ok = save_runtime_key(body.get("env", ""), body.get("value", ""))
                    if not ok:
                        self._send(400, json.dumps({"error": "অজানা key নাম"}, ensure_ascii=False))
                        return
                self._send(200, json.dumps(admin_get(), ensure_ascii=False))

            elif self.path.endswith("api/reminders"):
                body = self._json_body()
                action = body.get("action", "list")
                if action == "due":
                    self._send(200, json.dumps({"due": due_reminders()}, ensure_ascii=False))
                    return
                if action == "add":
                    reminder_tool("add", text=body.get("text", ""), when=body.get("when", ""))
                elif action == "delete":
                    reminder_tool("delete", number=body.get("number"))
                self._send(200, json.dumps(
                    {"reminders": [r for r in _rem_load() if not r.get("notified")]},
                    ensure_ascii=False))

            elif self.path.endswith("api/logo"):
                body = self._json_body()
                data = body.get("data", "")
                if not data or len(data) > 3_000_000:
                    self._send(400, json.dumps({"error": "ছবি নেই বা 2MB-এর বেশি বড়"}, ensure_ascii=False))
                    return
                try:
                    raw = base64.b64decode(data)
                    (Path(__file__).parent / "logo.png").write_bytes(raw)
                    log_activity("admin", "নতুন logo/ছবি")
                    self._send(200, json.dumps({"ok": True}))
                except Exception as e:
                    self._send(400, json.dumps({"error": f"ছবি সমস্যা: {e}"}, ensure_ascii=False))

            elif self.path.endswith("api/reset"):
                agent.reset()
                self._send(200, json.dumps({"ok": True}))
            else:
                self._send(404, json.dumps({"error": "not found"}))

        except LLMError as e:
            self._send(200, json.dumps({"error": str(e)}, ensure_ascii=False))
        except Exception as e:
            self._send(500, json.dumps({"error": f"Server error: {e}"}, ensure_ascii=False))

    def log_message(self, *args):
        pass


# ═══════════════════════════════════════════════════════════════════
# ১০. MAIN — চালু করা (web UI অথবা terminal)
# ═══════════════════════════════════════════════════════════════════

def run_terminal():
    print(f"""
╔══════════════════════════════════════════╗
║   🤖 Sabbir AI Agent                     ║
║   Model: {MODEL_NAME:<32}║
╚══════════════════════════════════════════╝
কমান্ড:  /reset = নতুন কথোপকথন,  /quit = বের হওয়া
""")
    if not GEMINI_API_KEY:
        print("⚠️  API key নেই! .env ফাইলে লিখুন: GEMINI_API_KEY=আপনার_key")
        print("   ফ্রি key: https://aistudio.google.com/apikey\n")
        return
    while True:
        try:
            user_input = input("আপনি: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n👋 বিদায়!")
            break
        if not user_input:
            continue
        if user_input in ("/quit", "/exit", "exit", "quit"):
            print("👋 বিদায়!")
            break
        if user_input == "/reset":
            agent.reset()
            print("🔄 Memory মুছে ফেলা হয়েছে।\n")
            continue
        try:
            print(f"\n🤖 Sabbir: {agent.run(user_input)}\n")
        except LLMError as e:
            print(f"\n❌ {e}\n")


def run_server():
    port = int(os.environ.get("PORT", 7860))  # Hugging Face Spaces port
    server = ThreadingHTTPServer(("0.0.0.0", port), Handler)
    print(f"🌐 Sabbir AI Agent চালু: http://localhost:{port}")
    print("   বন্ধ করতে: Ctrl+C")
    server.serve_forever()


if __name__ == "__main__":
    if "--terminal" in sys.argv:
        run_terminal()
    else:
        run_server()
